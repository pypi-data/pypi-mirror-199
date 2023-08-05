# coding: utf-8

import click
import openpyxl
from bonnie.excel.common import select_sheet
from bonnie.utils.file import add_suffix_to_filename


def transpose():
    src_file = click.prompt('Please input source file').strip()
    dst_file = click.prompt(
        'Please input dest file',
        add_suffix_to_filename(src_file, '_1')).strip()

    # 打开输入文件和输出文件
    input_workbook = openpyxl.load_workbook(src_file)
    sheet_index = select_sheet(input_workbook)

    output_workbook = openpyxl.Workbook()

    # 获取要处理的工作表
    input_sheet = input_workbook.worksheets[sheet_index]
    output_sheet = output_workbook.active

    # 遍历每一行
    for row_index, row in enumerate(input_sheet.iter_rows()):
        # 遍历该行中的每个单元格
        for cell_index, cell in enumerate(row):
            # 将该单元格的值写入交换后的位置
            output_sheet.cell(row=cell_index+1,
                              column=row_index+1,
                              value=cell.value)

    # 保存输出文件
    output_workbook.save(dst_file)
