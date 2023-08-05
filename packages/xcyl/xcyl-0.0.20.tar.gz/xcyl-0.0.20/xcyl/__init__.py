import os, sys, requests, json, pandas as pd, sqlite3
from openpyxl import load_workbook
from copy import deepcopy as dc

class xcylobj(object):
	"""
	the new excel object
	"""
	def __init__(self, filename: str = "TEMP_VALUE", useIndex: bool = False, useSheets: bool = True):
		if not filename.endswith(".xlsx"):
			filename += ".xlsx"
		self.filename = filename
		self.cur_data_sets = {}
		self.useIndex = useIndex
		self.useSheetNames = useSheets

		if os.path.exists(self.filename):
			print("Loading existing file")
			print("[", end='', flush=True)
			for sheet_name in load_workbook(self.filename, read_only=True, keep_links=False).sheetnames:
				print(self.filename+"_"+sheet_name)
				self.cur_data_sets[self.filename+"_"+sheet_name] = pd.read_excel(self.filename, engine="openpyxl", sheet_name=sheet_name)
				print(".", end='', flush=True)
			print("]")

	def __enter__(self):
		return self

	def close(self):
		if os.path.exists(self.filename):
			os.system("mv {} {}".format(self.filename, self.filename.replace(".xlsx", "_backup.xlsx")))

		try:
			keys, sheet_names, temp_name = list(self.cur_data_sets.keys()), [], "gen_name"
			for itr, key in enumerate(keys):
				og_sheet_name, final_sheet_name = key, key
				if len(key) > 20:
					self[temp_name+"_"+str(itr)] = dc(self[key])
					del self[key]
					final_sheet_name = temp_name+"_"+str(itr)

				sheet_names += [{
					"og_sheet_name": og_sheet_name,
					"final_sheet_name": final_sheet_name
				}]

			def dyct_frame(raw_dyct):
				dyct = dc(raw_dyct)
				for key in list(raw_dyct.keys()):
					dyct[key] = [dyct[key]]
				return pd.DataFrame.from_dict(dyct)

			if self.useSheetNames:
				self['sheet_names'] = pd.concat(list(map(
					dyct_frame,
					sheet_names
				)), ignore_index=True)

			with pd.ExcelWriter(self.filename, engine="xlsxwriter") as writer:
				for key, value in self.cur_data_sets.items():
					value.to_excel(writer, sheet_name=key, startrow=1, header=False, index=self.useIndex)
					worksheet = writer.sheets[key]
					(max_row, max_col) = value.shape
					worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': [{'header': column} for column in value.columns]})
					worksheet.set_column(0, max_col - 1, 12)
		except Exception as e:
			print("Exception :> {}".format(e))
			for key, value in self.cur_data_sets.items():
				value.to_csv(str(key) + ".csv")

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.close()
		return self

	def addr(self, sheet_name, dataframe):
		while sheet_name in list(self.cur_data_sets.keys()):
			sheet_name += "_"
		self.cur_data_sets[sheet_name] = dataframe
		return self

	def add_frame(self, sheet_name, dataframe):
		self.addr(sheet_name, dataframe)

	def add_file(self, foil):
		if foil.endswith(".xlsx") and os.path.isfile(foil):
			with xcylr(foil) as x:
				for key, value in x.items():
					self.add_frame(key, value)
		elif foil.endswith(".csv") and os.path.isfile(foil):
			self.add_frame(foil, pd.read_csv(foil))
		elif foil.endswith(".json") and os.path.isfile(foil):
			self.add_frame(foil, pd.read_json(foil))

	def __iadd__(self, dataframe):
		if isinstance(dataframe, pd.DataFrame):
			self.add_frame("Addon", dataframe)
		else:
			print("The object is not a dataframe")
		return self

	def keys(self):
		return self.cur_data_sets.keys()

	def values(self):
		return self.cur_data_sets.values()

	def items(self):
		return self.cur_data_sets.items()

	def __getitem__(self, item):
		return self.cur_data_sets[item]

	def __setitem__(self, key, value):
		self.add_frame(key, value)

	def __delitem__(self, key):
		if key in self.keys():
			del self.cur_data_sets[key]

	def __iter__(self):
		return self.cur_data_sets.__iter__()
	
	def to_sqlobj(self):
		to_sql = sqlobj("temp.sqlite")
		for key, value in self.cur_data_sets.items():
			to_sql[key] = value
		return to_sql


class sqlobj(object):
	"""
	Sample usage:
	```
	with sqlobj("dataset.sqlite") as sql:
		container = pd.read_sql(sql.table_name, sql.connection_string)
	...
	with sqlobj("dataset.sqlite", threadLock=<x>) as sql:
		container = pd.read_sql(sql.table_name, sql.connection_string)
	...
	with sqlobj("dataset.sqlite") as sql:
		container.to_sql(sql.table_name, sql.connection, if_exists='replace')
	```
	"""

	def __init__(self, file_name: str, echo: bool = False, threadLock = None):
		self.file_name = file_name
		self.table_name = "dataset"
		self.echo = echo
		self.dataframes = {}
		self.exists = None
		self.lock = threadLock
		self.connection = None

	def just_enter(self):
		if self.exists is None:
			self.exists = os.path.exists(self.file_name)
			self.connection = sqlite3.connect(self.file_name)

	def enter(self):
		if self.lock:
			self.lock.acquire()

		if self.exists is None:
			self.just_enter()

			if self.exists:
				current_cursor = self.connection.cursor()
				current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
				for name in current_cursor.fetchall():
					self.dataframes[name[0]] = pd.read_sql_query("""SELECT * FROM "{0}";""".format(name[0]), self.connection)
				current_cursor = None

	def __enter__(self):
		self.enter()
		return self

	def exit(self):
		self.connection.close()
		self.exists = None

		if self.lock:
			self.lock.release()
		return self

	def __exit__(self, exc_type, exc_val, exc_tb):
		self.exit()
		return self

	def table_names(self):
		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		tables = []
		if just_enter:
			current_cursor = self.connection.cursor()
			current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

			for name in current_cursor.fetchall():
				tables += [name[0]]

			current_cursor = None
		else:
			tables = self.dataframes.keys()

		if just_enter:
			self.exit()

		return tables

	def load_table(self, table_name):
		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		output = pd.DataFrame()
		if just_enter:
			output = pd.read_sql_query("""SELECT * FROM "{0}";""".format(table_name), self.connection)
		else:
			output = self.dataframes[table_name]

		if just_enter:
			self.exit()

		return output

	def addr(self, sheet_name, og_dataframe, drop_columns={'level_0':'equals','Unnamed':'startswith'},if_exists='replace'):
		from copy import deepcopy as dc
		dataframe = dc(og_dataframe)
		while sheet_name in list(self.keys()):
			sheet_name += "_"

		for col in list(dataframe.columns):
			for drop_col, drop_match in drop_columns.items():
				if col.startswith(drop_col) if drop_match == 'startswith' else col == drop_col:
					dataframe.drop(columns=[col],inplace=True)
					break

		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		dataframe.to_sql(sheet_name, self.connection, if_exists=if_exists)

		if just_enter:
			self.exit()

		return self

	def add_frame(self, sheet_name, dataframe,drop_columns={'level_0':'equals','Unnamed':'startswith'}):
		self.addr(sheet_name, dataframe, drop_columns=drop_columns)

	def add_csv(self, filename):
		if os.path.exists(filename):
			tempdata = pd.read_csv(filename)
			self.addr(filename.split('/')[-1].replace('.csv',''), tempdata)

	def __iadd__(self, dataframe):
		if isinstance(dataframe, pd.DataFrame):
			self.add_frame("Addon", dataframe)
		else:
			print("The object is not a dataframe")
		return self

	def keys(self):
		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		tables = []
		if just_enter:
			current_cursor = self.connection.cursor()
			current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

			for name in current_cursor.fetchall():
				tables += [name[0]]

			current_cursor = None
		else:
			tables = self.dataframes.keys()

		if just_enter:
			self.exit()

		return tables

	def values(self):
		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()
		
		tables = []
		if just_enter:
			current_cursor = self.connection.cursor()
			current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
			for name in current_cursor.fetchall():
				tables  += [pd.read_sql_query("""SELECT * FROM "{0}";""".format(name[0]), self.connection)]
			current_cursor = None
		else:
			tables = self.dataframes.values()


		if just_enter:
			self.exit()

		return tables

	def items(self):
		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		tables = {}
		if just_enter:
			current_cursor = self.connection.cursor()
			current_cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
			for name in current_cursor.fetchall():
				tables[name[0]] =  [pd.read_sql_query("""SELECT * FROM "{0}";""".format(name[0]), self.connection)]
			current_cursor = None
		else:
			tables = self.dataframes.values()

		if just_enter:
			self.exit()

		return self.dataframes

	def __getitem__(self, item):
		return self.load_table(item)

	def __setitem__(self, key, value):
		self.add_frame(key, value)

	def tocsv(self, table_name):
		table_data = self[table_name] if table_name in list(self.table_names()) else None
		if table_data is not None:
			table_data.to_csv("{0}.csv".format(table_name))

	def drop_table(self, table_name, backup=True):
		table_names = list(self.table_names())
		backup_table_data = self[table_name] if backup and table_name in table_names else None

		just_enter = False
		if self.exists is None:
			just_enter = True
			self.just_enter()

		if table_name in table_names:
			if backup:
				backup_table_data.to_csv(".backup_{0}.csv".format(table_name))

			if just_enter:
				current_cursor = self.connection.cursor()
				current_cursor.execute("""DROP table IF EXISTS "{0}";""".format(table_name))
				current_cursor = None
			else:
				try:
					self.dataframes.pop(table_name)
				except: pass

		if just_enter:
			self.exit()

		return

	def to_xcylobj(self):
		output = xcylobj("temp.xlsx")
		for key, value in self.dataframes.items():
			output[key] = value
		return output
