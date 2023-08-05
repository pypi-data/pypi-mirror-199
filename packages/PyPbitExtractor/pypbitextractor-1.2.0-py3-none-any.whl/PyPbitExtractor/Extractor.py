import subprocess

my_libraries = ['openpyxl', 'codecs', 'json', 'zipfile', 'pathlib', 'openai', 'os', 'tkinter', 'shutil']
for module in my_libraries:
    try:
        __import__(module)
    except ImportError:
        subprocess.check_call(["pip", "install", module])

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.alignment import Alignment
import codecs
import json
import zipfile
import pathlib
from pathlib import Path
import openai
import os
from tkinter import filedialog, Tk
from tkinter.filedialog import askdirectory
import shutil
import time

openai.api_key = "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"


def xls_extract(data, file, base_file_name, json_path):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title="Measures")
    cnt = 0
    r = 0
    r1 = 0
    for t in data['model']['tables']:
        if 'measures' in t:
            list_measures = t['measures']

            for i in list_measures:
                if (r1 == 20):
                    time.sleep(30)
                    r1 = 0
                if (type(i['expression']) is list):
                    exp = " ".join(i['expression'][0:])
                else:
                    exp = str(i['expression'])
                prompt = "Explain the following calculation in a few sentences in simple business terms without using DAX function names: " + exp
                completion = openai.Completion.create(engine="text-davinci-002", prompt=prompt, max_tokens=64)
                r1 += 1
                i['description'] = completion.choices[0]['text'].strip()
                # i['description'] = "No Description"
                cnt += 1
                if (cnt == 1):
                    ws.append(['Measure Name', 'Measure Expression', 'Measure Data Type', 'Measure Folder',
                               'Measure Description'])
                # ws.append([i['name'], exp, i['dataType'] if 'dataType' in i else "", i['description']])
                exp = ""
                if (type(i['expression']) is list):
                    exp = "\n".join(i['expression'][0:])
                else:
                    exp = i['expression']
                # exp= i['name'] + "=" + exp
                prompt = "Format this dax query: " + exp

                if "\"" not in exp and (len(exp) > 1):
                    completion = openai.Completion.create(engine="text-davinci-002", prompt=prompt, max_tokens=256)
                    r1 += 1
                    if (r1 == 20):
                        time.sleep(15)
                        r1 = 0
                    if "formatted" not in completion.choices[0]['text'] or "Formatting" not in completion.choices[0][
                        'text'] or "properly" not in completion.choices[0]['text']:
                        exp = ("\n" + completion.choices[0]['text'].strip().replace(";", "")).split("\n")
                        exp_o = "\n" + completion.choices[0]['text'].strip().replace(";", "")

                    # i['expression'] = exp
                    # print("MEASURE ", i['description'])
                    if (type(i['expression']) is list):
                        exp_1 = [element.strip() for element in i['expression']]
                    else:
                        i['expression'] = "\n" + i['expression']
                        exp_1 = (" \n" + i['expression']).split("\n")
                        exp_1 = [element.strip() for element in exp_1]
                    exp_2 = [element.strip() for element in exp]
                    diff1 = set(str(exp_1)) - set(str(exp_2))
                    diff2 = set(str(exp_2)) - set(str(exp_1))
                    if (len(diff1) == 0) and (len(diff2) == 0):
                        i['expression'] = exp
                    if (len(exp_o) != 0):
                        ws.append([i['name'], exp_o, i['dataType'] if 'dataType' in i else "",
                                   i['displayFolder'] if 'displayFolder' in i else "No Folder Defined",
                                   i['description']])
                    else:
                        ws.append([i['name'], str(i['expression']), i['dataType'] if 'dataType' in i else "",
                                   i['displayFolder'] if 'displayFolder' in i else "No Folder Defined",
                                   i['description']])
                else:
                    # exp_o=str(i['expression'])
                    if (len(exp) != 0):
                        ws.append([i['name'], exp, i['dataType'] if 'dataType' in i else "No DataType Defined",
                                   i['displayFolder'] if 'displayFolder' in i else "No Folder Defined",
                                   i['description']])
                    else:
                        ws.append([i['name'], str(i['expression']),
                                   i['dataType'] if 'dataType' in i else "No DataType Defined",
                                   i['displayFolder'] if 'displayFolder' in i else "No Folder Defined",
                                   i['description']])

    if (cnt >= 1):
        table = Table(displayName="Table1", ref="A1:E" + str(cnt + 1))
        ws.add_table(table)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, wrapText=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 50
    else:
        ws.append(['No measures present in this file'])
        adjusted_width = (len('No measures present in this file') + 2) * 1.2
        ws.column_dimensions["A"].width = adjusted_width

    print(cnt)

    source = wb.create_sheet('Source Information')
    i = 0
    if 'tables' in data['model']:
        source.append(
            ['Table No', 'Table Name', 'Table Mode', 'Table Type', 'Table Source', 'Original Table Name', 'Table Query',
             'Modification', 'Modification Description'])
        for t in data['model']['tables']:
            if 'partitions' in t:
                list_partitions = t['partitions']
                name = list_partitions[0]['name'].split('-')[0]
                mode = str(list_partitions[0]['mode'])
                if (r == 60):
                    time.sleep(30)
                    r = 0
                if "DateTableTemplate" not in name and "LocalDateTable" not in name:
                    if list_partitions[0]['source']['type'] == 'entity':
                        Ttype = list_partitions[0]['source']['type']
                        TSource = list_partitions[0]['source']['expressionSource']
                        otname = list_partitions[0]['source']['entityName']
                        i += 1
                        source.append(
                            [i, name, mode, Ttype, TSource, otname, "NO Query", "No Modification", "No Description"])
                    elif list_partitions[0]['source']['type'] == 'calculated':
                        List_source = (list_partitions[0]['source']['expression'])
                        mode = str(list_partitions[0]['mode'])
                        Ttype = "Calculated"
                        TSource = "Calculated Table"
                        otname = name
                        if (type(List_source) is list):
                            TQuery = " ".join(List_source[0:])
                        else:
                            TQuery = str(List_source)
                        Tmodification = "No Modification"
                        Tdescription = "DAX Query Description: \n"
                        prompt = "Explain the following calculation in a few sentences in simple business terms without using DAX function names: " + TQuery
                        completion = openai.Completion.create(engine="text-ada-001", prompt=prompt, max_tokens=128)
                        r += 1
                        if (r == 60):
                            time.sleep(30)
                            r = 0
                        Tdescription += completion.choices[0]['text'].strip().replace("\n", "")
                        # t['description'] = 'Tdescription'
                        i += 1
                        source.append([i, name, mode, Ttype, TSource, otname, TQuery, Tmodification, Tdescription])
                        # print("MODIFICATION ", Tdescription)
                    else:
                        if 'expression' in list_partitions[0]['source']:
                            List_source = (list_partitions[0]['source']['expression'])
                            expression = []
                            if List_source[0] == 'let':
                                expression.append(List_source[0])
                                name = list_partitions[0]['name'].split('-')[0]
                                mode = str(list_partitions[0]['mode'])

                                p = List_source[1]
                                expression.append(List_source[1])
                                Ttype = ""
                                if '= ' in p.split("(")[0]:
                                    Ttype = p.split("(")[0].split('= ')[1]
                                elif '=' in p.split("(")[0]:
                                    Ttype = p.split("(")[0].split('=')[1]
                                else:
                                    pass
                                if '"' in p:
                                    if "Value.NativeQuery" in p:
                                        TSource = p.split("Value.NativeQuery(")[1].split("SELECT")[0].replace("#(lf)",
                                                                                                              " ")
                                    elif "Table.FromRows" in p:
                                        TSource = p.split("FromRows(")[1]
                                    else:
                                        st = 0
                                        ed = len(p) - 1
                                        while (p[st] != '"'):
                                            st += 1
                                        while (p[ed] != '"'):
                                            ed -= 1
                                        TSource = p[st: ed + 1]
                                    if "Query=" in TSource:
                                        TSource = TSource.split("[Query")[0]
                                    if "Delimiter=" in TSource:
                                        TSource = TSource.split("),[Delimiter")[0]
                                elif "Source" in p:
                                    TSource = p.split("=")[1]
                                else:
                                    TSource = List_source[2]
                            otname = ""
                            if "Query=" in p:
                                otname = p.replace("#(lf)", "")
                                otname = otname.split("FROM")[1]
                                if "WHERE" in otname:
                                    otname = otname.split("WHERE")[0]
                                if "INNER JOIN" in otname:
                                    otname = otname.split("INNER JOIN")
                                    n = []
                                    for tname in otname:
                                        if 'ON' in tname:
                                            n.append(tname.split("ON")[0])
                                        else:
                                            n.append(tname)
                                    otname = ",\n".join(n[0:])
                            elif "Sql." in p:
                                if "Name" in List_source[2]:
                                    otname = List_source[3].split("=")[0].strip()
                                else:
                                    otname = List_source[2].split("=")[0].strip()
                            elif "Excel." in p:
                                otname = List_source[2].split("=")[0].strip()
                                otname = otname.replace("#", "")
                            elif "Dataflows" in p:
                                otname = List_source[5].split("=")[0].split("\"")[1].strip()
                            else:
                                otname = name

                            TQuery = ""
                            if "NativeQuery" in p:
                                TQuery = "SELECT" + p.split("NativeQuery")[1].split("SELECT")[1]
                                TQuery = TQuery.replace("#(lf)", " ")
                            elif "Query=" in p:
                                TQuery = p.split("Query=")[1]
                                TQuery = TQuery.replace("#(lf)", " ")
                            else:
                                TQuery = "No Query"

                            if "import" in mode:
                                idx = -1
                                for i1 in range(2, len(List_source)):
                                    if "Sheet" not in List_source[i1]:
                                        if len(List_source[i1]) > 5 and List_source[i1][4] == '#':
                                            idx = i1
                                            break
                                for i1 in range(2, idx):
                                    expression.append(List_source[i1])
                                Tmodification = ""
                                Tdescription = ""
                                completion = "No Description"
                                if idx == -1:
                                    Tmodification = "No Modification"
                                    Tdescription = "No Description"
                                else:
                                    pr = 1
                                    for id in range(idx, len(List_source) - 2):
                                        try:
                                            p1 = List_source[id].split("    ")[1].strip()
                                        except:
                                            p1 = List_source[id].strip()
                                        if '//' not in p1:
                                            Tmodification += str(pr) + ". " + p1 + '\n\n'
                                            prompt = "Explain this in normal terms in one sentence: " + p1
                                            completion = openai.Completion.create(engine="text-ada-001", prompt=prompt,
                                                                                  max_tokens=64)
                                            r += 1
                                            if (r == 60):
                                                time.sleep(30)
                                                r = 0
                                            to = completion.choices[0]['text'].strip().replace("\n", "")
                                            Tdescription += str(
                                                pr) + ". " + to if completion != 'No Description' else completion
                                            Tdescription += '\n\n'
                                            # Tdescription = 'Tdescription'
                                            to = completion.choices[0]['text'].strip().replace("\n", "")
                                            to = "//" + to
                                            if completion != 'No Description':
                                                expression.append("    " + to)
                                            expression.append(List_source[id])
                                            pr += 1
                                    expression.append(List_source[len(List_source) - 2])
                                    expression.append(List_source[len(List_source) - 1])
                                    t['partitions'][0]['source']['expression'] = expression
                                    i += 1

                                    source.append(
                                        [i, name, mode, Ttype, TSource.replace("#(lf)", " "),
                                         otname.replace("#(lf)", " "), TQuery.replace("#(lf)", " "), Tmodification,
                                         Tdescription])
                                    # print("MODIFICATION ", Tdescription)
                            else:
                                i += 1
                                source.append(
                                    [i, name, mode, Ttype, TSource, otname, TQuery, "No Modification",
                                     "No Description"])

        with codecs.open(json_path, 'w', 'utf-16-le') as f:
            json.dump(data, f, indent=4)

        i += 1
        table = Table(displayName="Source", ref="A1:I" + str(i))
        source.add_table(table)

        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style

        for col in source.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        source.column_dimensions["A"].width = 20
        source.column_dimensions["B"].width = 20
        source.column_dimensions["C"].width = 20
        source.column_dimensions["D"].width = 30
        source.column_dimensions["E"].width = 40
        source.column_dimensions["F"].width = 30
        source.column_dimensions["G"].width = 50
        source.column_dimensions["H"].width = 80
        source.column_dimensions["I"].width = 80
    else:
        source.append(['No Source present in this file'])
        adjusted_width = (len('No Source present in this file') + 2) * 1.2
        source.column_dimensions["A"].width = adjusted_width

    relation = wb.create_sheet('Relationships')
    if 'relationships' in data['model']:
        cnt1 = 0
        for t in data['model']['relationships']:
            direction = ""
            cardinality = ""
            x11 = t['fromTable'].split('_')
            x12 = t['toTable'].split('_')
            if (x11[0] != 'LocalDateTable' and x12[0] != 'LocalDateTable') and "joinOnDateBehavior" not in t:
                cnt1 += 1
                if "crossFilteringBehavior" in t:
                    direction = "Both Directional"
                else:
                    direction = "Single Directional"

                if "toCardinality" in t:
                    if t['toCardinality'] == "one":
                        cardinality = 'One to one (1:1)'
                    elif t['toCardinality'] == "many":
                        cardinality = 'Many to many (*:*)'
                    else:
                        pass

                elif "fromCardinality" in t:
                    if t['fromCardinality'] == "one":
                        cardinality = 'One to one (1:1)'
                    elif t['fromCardinality'] == "many":
                        cardinality = 'Many to many (*:*)'
                    else:
                        pass
                else:
                    cardinality = 'Many to one (*:1)'

                if (cnt1 == 1):
                    relation.append(
                        ['From Table', 'From Column', 'To Table', 'To Column', 'State', 'Direction', 'Cardinality'])
                relation.append(
                    [t['fromTable'], t['fromColumn'], t['toTable'], t['toColumn'],
                     t['state'] if 'state' in t else "Inactive",
                     direction, cardinality])

        if (cnt1 >= 1):
            table = Table(displayName="Relationships", ref="A1:G" + str(cnt1 + 1))
            relation.add_table(table)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style

            for col in relation.columns:
                for cell in col:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            relation.column_dimensions["A"].width = 40
            relation.column_dimensions["B"].width = 20
            relation.column_dimensions["C"].width = 40
            relation.column_dimensions["D"].width = 20
            relation.column_dimensions["E"].width = 20
            relation.column_dimensions["F"].width = 20
            relation.column_dimensions["G"].width = 20

        else:
            relation.append(['No relationships present in this file'])
            adjusted_width = (len('No relationships present in this file') + 2) * 1.2
            relation.column_dimensions["A"].width = adjusted_width

    else:
        relation.append(['No relationships present in this file'])
        adjusted_width = (len('No relationships present in this file') + 2) * 1.2
        relation.column_dimensions["A"].width = adjusted_width

    dir_name = os.path.dirname(file)
    new_dir = pathlib.Path(dir_name, "EXCEL Output")
    new_dir.mkdir(parents=True, exist_ok=True)
    file_name = base_file_name + ".xlsx"
    save1 = str(new_dir) + "\\" + file_name
    wb.save(save1)
    wb = openpyxl.load_workbook(save1)
    sheet_to_remove = wb['Sheet']
    wb.remove(sheet_to_remove)
    wb.save(save1)
    print("Created :", file_name)


def json_extract(file):
    temp_dir = 'temporary'
    os.makedirs(temp_dir, exist_ok=True)
    dir_name = os.path.dirname(file)
    base_file_name = Path(file).stem

    x = list(file)
    x[-1] = 't'
    file = ''.join(x)
    pbit_path = Path(file)

    with zipfile.ZipFile(pbit_path, 'r') as zip_ref:
        zip_ref.extractall(temp_dir)

    json_path = os.path.join(temp_dir, 'DataModelSchema')
    with codecs.open(json_path, 'r', 'utf-16-le') as f:
        contents = f.read()

    data = json.loads(contents)

    new_dir = pathlib.Path(dir_name, "JSON Output")
    new_dir.mkdir(parents=True, exist_ok=True)
    xls_extract(data, file, base_file_name, json_path)
    with codecs.open(json_path, 'r', 'utf-16-le') as f:
        contents = f.read()

    data = json.loads(contents)
    file_name = base_file_name + ".json"

    save1 = str(new_dir) + "\\" + file_name
    out_file = open(Path(save1), "w")

    json.dump(data, out_file, indent=6)
    out_file.close()
    print("Created :", file_name)

    dir_name = os.path.dirname(file)
    new_dir = pathlib.Path(dir_name, "UPDATED pbit")
    new_dir.mkdir(parents=True, exist_ok=True)

    new_pbit_path = base_file_name + '_updated.pbit'
    save1 = str(new_dir) + "\\" + new_pbit_path
    print("Created :", new_pbit_path)
    print()
    with zipfile.ZipFile(save1, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                rel_path = os.path.relpath(file_path, temp_dir)
                zip_ref.write(file_path, rel_path)

    shutil.rmtree(temp_dir)


def api():
    key = input("Enter your OpenAI API Secret key: ")
    if (len(key) == len(openai.api_key)):
        openai.api_key = key


def main():
    print("Do you want to select file or folder\n1. File\n2. Multiple Files\n3. Folder")
    op = int(input("\nEnter Option: "))
    root = Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    if (op == 1):
        file = filedialog.askopenfilename(title="Select file")
        if (file != ""):
            print("Currently Processing {" + str(Path(file).stem) + ".pbit}...")
            json_extract(file)
        else:
            print("No File Selected")
    elif (op == 2):
        files = filedialog.askopenfilenames(title="Select files")
        if (files != ""):
            for f in files:
                print("Currently Processing {" + str(Path(f).stem) + ".pbit}...")
                json_extract(f)
        else:
            print("No Files Selected.")
    elif (op == 3):
        folder = askdirectory(title="Select folder")
        if (folder != ""):
            for f in os.listdir(folder):
                if f.endswith('.pbit'):
                    print("Currently Processing {" + str(f) + "}...")
                    file_path = f"{folder}/{f}"
                    json_extract(file_path)
        else:
            print("No Folder Selected.")
    else:
        print("Invalid Input")
