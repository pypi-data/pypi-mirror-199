"""
    ExcelReader：读取数据
    ExcelWriter：保存数据（可能会存在丢失长度问题）
    ExcelWriterPerfect：保存数据（不会丢失）

    使用：
        # 写入
        excel = ExcelWriter()
        writer = excel.writer(sheetname='Sheet')

        for i in range(10):
            writer.write_line([i])

        excel.save('Result.xlsx')


        # 写入2
        excel = ExcelWriterPerfect('Result.xlsx')
        writer = excel.writer(sheetname='Sheet')

        for i in range(10):
            writer.write_line([i])

        excel.save()


        # 读取
        reader = ExcelReader('Result.xlsx')
        for i in reader.read_lines():
            print(i)
"""
import xlsxwriter
from typing import List
from pathlib import Path
from threading import RLock
from openpyxl import load_workbook, Workbook


class _Writer:
    def __init__(self, actuator, sheetname: str):
        """

        :param actuator: ExcelWriter
        :param sheetname: sheetname
        """
        self.actuator = actuator
        self.sheetname = sheetname

        self.actuator.lock.acquire()
        try:
            if self.sheetname not in self.actuator.sheetnames_save:
                if self.sheetname == 'Sheet':
                    self.sh = self.actuator.wb.active
                else:
                    self.sh = self.actuator.wb.create_sheet(self.sheetname)
                self.actuator.sheetnames_save[self.sheetname] = self.sh
            else:
                self.sh = self.actuator.wb[self.sheetname]
        finally:
            self.actuator.lock.release()

    def write_line(self, line: list) -> None:
        """

        :param line: 保存的数据
        :return:
        """
        self.sh.append(line)

    def write_lines(self, lines: List[list]):
        """

        :param lines: 保存的数据列表包列表
        :return:
        """
        for line in lines:
            self.write_line(line)


class ExcelWriter:
    def __init__(self, filepath: str = None):
        """

        :param filepath: 文件保存路径
        """
        self.filepath = filepath

        if self.filepath:
            self.wb = load_workbook(self.filepath)
        else:
            self.wb = Workbook()

        self.lock = RLock()
        self.sheetnames_save = {}

    def writer(self, sheetname: str = None) -> _Writer:
        """
        writer 可使用多次，返回多个 sheet 表用于操作

        :param sheetname: 保存的 sheetname, 无则为默认或第一个
        :return:
        """
        sheetname = sheetname or (self.sheetnames[0] if self.sheetnames else 'Sheet')
        return _Writer(actuator=self, sheetname=sheetname)

    def save(self, filepath: str = None) -> None:
        """
        保存，有路径则存，无则为输入路径

        :param filepath: 文件保存路径
        :return:
        """
        if 'Sheet' not in self.sheetnames_save:
            del self.wb['Sheet']
        filepath = filepath or self.filepath or 'Result.xlsx'
        self.wb.save(filepath)

    @property
    def sheetnames(self) -> List[str]:
        """
        返回所有的 sheet name

        :return:
        """
        return self.wb.sheetnames

    def __del__(self):
        self.wb.close()


class _WriterPerfect:
    def __init__(self, actuator, sheetname: str):
        """

        :param actuator: ExcelWriterPerfect
        :param sheetname: sheetname
        """
        self.actuator = actuator
        self.sheetname = sheetname

        self.actuator.lock.acquire()
        try:
            if self.sheetname not in self.actuator.sheetnames:
                self.sh = self.actuator.wb.add_worksheet(self.sheetname)
                self.actuator.sheetnames[self.sheetname] = self.sh
            else:
                self.sh = self.actuator.sheetnames[self.sheetname]
        finally:
            self.actuator.lock.release()

        self._index = -1

    def write_line(self, line: list) -> None:
        """

        :param line: 保存的数据
        :return:
        """
        index = self._add_index()
        self.sh.write_row(row=index, col=0, data=line)

    def write_lines(self, lines: List[list]) -> None:
        """

        :param lines: 保存的数据列表包列表
        :return:
        """
        for line in lines:
            self.write_line(line)

    def _add_index(self) -> int:
        """
        修改索引序号并返回

        :return:
        """
        self.actuator.lock.acquire()
        try:
            self._index += 1

            return self._index
        finally:
            self.actuator.lock.release()


class ExcelWriterPerfect:
    def __init__(self, filepath: str):
        """

        :param filepath: 文件保存路径
        """
        self.filepath = Path(filepath)
        self.wb = xlsxwriter.Workbook(filepath)

        self.lock = RLock()
        self.sheetnames = {}

    def writer(self, sheetname: str = None) -> _WriterPerfect:
        """

        :param sheetname: 保存的 sheetname,只需给一次就行
        :return:
        """
        sheetname = sheetname or 'Sheet'
        return _WriterPerfect(actuator=self, sheetname=sheetname)

    def save(self) -> None:
        self.wb.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
