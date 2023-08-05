"""
    ExcelReader：读取数据（xlsx、xls）
    XlsxReader：读取 xlsx 数据
    XlsReader：读取 xls 数据

    使用：
        # 读取
        reader = ExcelReader('Result.xlsx')
        reader = XlsxReader('Result.xlsx')
        reader = XlsReader('Result.xls')
        for i in reader.read_lines():
            print(i)
"""
import xlrd
from pathlib import Path
from typing import List, Union
from openpyxl import load_workbook


class ExcelReader:
    """
        自适应读取 xlsx、xls 文件
        打开失败的唯一可能就是后缀和实际类型不匹配
    """

    def __init__(self, filepath: Union[str, Path]):
        """

        :param filepath: 读取文件路径
        """
        if not isinstance(filepath, Path):
            self.filepath = Path(filepath)
        else:
            self.filepath = filepath

        if not self.filepath.exists():
            raise FileExistsError(f'文件不存在：{self.filepath}')

        suffix = self.filepath.suffix
        if suffix == '.xlsx':
            self._reader = XlsxReader(self.filepath)
        elif suffix == '.xls':
            self._reader = XlsReader(self.filepath)
        else:
            raise Exception('不支持的文件类型，只支持 xlsx、xls！')

        self.wb = self._reader.wb

    def read_lines(self, sheetname: str = None, index: int = None):
        """
        读取 sheet 的行，默认第一个 sheet

        :param sheetname: 通过 名字 打开指定 sheet
        :param index: 通过 索引 打开指定 sheet
        :return:
        """
        return self._reader.read_lines(sheetname=sheetname, index=index)

    @property
    def sheetnames(self) -> List[str]:
        """
        返回所有的 sheet name

        :return:
        """
        return self._reader.sheetnames


class XlsxReader:
    """
        读取 xlsx 文件
    """

    def __init__(self, filepath: Union[str, Path]):
        """

        :param filepath: 读取文件路径
        """
        if not isinstance(filepath, Path):
            self.filepath = Path(filepath)
        else:
            self.filepath = filepath

        if not self.filepath.exists():
            raise FileExistsError(f'文件不存在：{self.filepath}')
        elif self.filepath.suffix != '.xlsx':
            raise Exception('不支持的文件类型，只支持 xlsx！')

        self.wb = load_workbook(self.filepath)

    def read_lines(self, sheetname: str = None, index: int = None):
        """
        读取 sheet 的行，默认第一个 sheet

        :param sheetname: 通过 名字 打开指定 sheet
        :param index: 通过 索引 打开指定 sheet
        :return:
        """
        if sheetname:
            sh = self.wb[sheetname]
        elif index:
            length = len(self.sheetnames)
            if index >= length:
                raise IndexError("该表格索引范围应在：0-{} 内！".format(length - 1))
            sh = self.wb[self.sheetnames[index]]
        else:
            sh = self.wb.active

        for row in sh.rows:
            row_list = []
            for col in row:
                row_list.append(col.value)
            yield row_list

    @property
    def sheetnames(self) -> List[str]:
        """
        返回所有的 sheet name

        :return:
        """
        return self.wb.sheetnames

    def __del__(self):
        self.wb.close()


class XlsReader:
    """
        读取 xls 文件
    """

    def __init__(self, filepath: Union[str, Path]):
        """

        :param filepath: 读取文件路径
        """
        if not isinstance(filepath, Path):
            self.filepath = Path(filepath)
        else:
            self.filepath = filepath

        if not self.filepath.exists():
            raise FileExistsError(f'文件不存在：{self.filepath}')
        elif self.filepath.suffix != '.xls':
            raise Exception('不支持的文件类型，只支持 xls！')

        self.wb = xlrd.open_workbook(self.filepath)

    def read_lines(self, sheetname: str = None, index: int = None):
        """
        读取 sheet 的行，默认第一个 sheet

        :param sheetname: 通过 名字 打开指定 sheet
        :param index: 通过 索引 打开指定 sheet
        :return:
        """
        if sheetname:
            sh = self.wb.sheet_by_name(sheetname)
        elif index:
            length = len(self.wb.nsheets)
            if index >= length:
                raise IndexError("该表格索引范围应在：0-{} 内！".format(length - 1))
            sh = self.wb.sheet_by_index(index)
        else:
            sh = self.wb.sheet_by_index(0)

        for row_index in range(sh.nrows):
            yield sh.row_values(row_index)

    @property
    def sheetnames(self) -> List[str]:
        """
        返回所有的 sheet name

        :return:
        """
        sheets = self.wb.sheets()
        if sheets:
            return [i.name for i in self.wb.sheets()]
        return []
