from jsonpath import jsonpath
from ddreport.exceptd import exceptContentObj
import os

class PytestFunctions:

    # 抛出异常
    def __execptions(self, err):
        exceptContentObj.raiseException({"错误详情": err})

    # 错误数据
    def __raiseAssert(self, d1, d2, n=None):
        str_content = f"\n实际值{type(d1)}：{d1}\n预期值{type(d2)}：{d2}\n"
        index = f"行数：{n + 1}" if n is not None else ''
        err_str = str_content + index
        self.__execptions(err_str)

    def __raiseLoseData(self, d2):
        err_str = f"\n缺少数据为：{d2}"
        self.__execptions(err_str)

    # 字典字段显示与隐藏
    def jsonKeysShow(self, data, keys, is_show=True):
        def dictHandle(data, is_show):
            d = dict()
            is_show and [d.update({k: v}) for k, v in data.items() if k in keys] or [d.update({k: v}) for k, v in data.items() if k not in keys]
            return d
        if isinstance(data, dict):
            return dictHandle(data, is_show)
        elif isinstance(data, list):
            return [isinstance(i, dict) and dictHandle(i, is_show) or i for i in data]

    # json排序
    def jsonSort(self, data):
        def dictSort(data):
            d = dict(sorted(data.items(), key=lambda x: x[0]))
            return d
        if isinstance(data, dict):
            return dictSort(data)
        elif isinstance(data, list):
            return [isinstance(item, dict) and dictSort(item) or item for item in data]
        else:
            return data

    # 有序数据对比
    def checkOrder(self, data1, data2):
        """
        data1: 实际数据
        data2：预期数据
        """
        if data1 != data2:
            if isinstance(data1, (list, dict)):
                data1 = self.jsonSort(data1)
            if isinstance(data2, (list, dict)):
                data2 = self.jsonSort(data2)
            if isinstance(data1, list) and isinstance(data2, list):
                if len(data1) > len(data2):
                    for n, i in enumerate(data1):
                        if n > len(data2) - 1:
                            self.__raiseAssert(i, None, n)
                        elif i != data2[n]:
                            self.__raiseAssert(i, data2[n], n)
                else:
                    for n, i in enumerate(data2):
                        if n > len(data1) - 1:
                            self.__raiseAssert(None, i, n)
                        elif i != data1[n]:
                            self.__raiseAssert(data1[n], i, n)
            else:
                self.__raiseAssert(data1, data2)

    # 无序数据对比
    def uncheckOrder(self, data1, data2, only_keys=None):
        """
        data1: 实际数据
        data2：预期数据
        only_keys： 字段列表---一般选取唯一字段
        """
        if isinstance(data1, (list, dict)):
            data1 = self.jsonSort(data1)
        if isinstance(data2, (list, dict)):
            data2 = self.jsonSort(data2)
        if data1:
            data1_copy = data1.copy()
        if data2:
            data2_copy = data2.copy()
        if isinstance(data1, list) and isinstance(data2, list):
            for i in data1:
                data1_copy.remove(i)
                if not only_keys:
                    if i not in data2:
                        self.__raiseAssert(i, None)
                    else:
                        data2_copy.remove(i)
                else:
                    exp = ''
                    for k in only_keys:
                        s = isinstance(i[k], str) and f'"{i[k]}"' or f"{i[k]}"
                        exp += f'@.{k}=={s}' + ' and '
                    get_yq_data = jsonpath(data2, f'$..[?({exp[:-5]})]')
                    if not get_yq_data:
                        self.__raiseAssert(i, None)
                    else:
                        data2_copy.remove(get_yq_data[0])
                        self.checkOrder(i, get_yq_data[0])
            if data1_copy != data2_copy:
                # 丢失数据
                self.__raiseLoseData(data2_copy)
        else:
            if data1 != data2:
                self.__raiseAssert(data1, data2)

    # 读取xlsx并转为LIST_DICT
    def readXlsx(self, file_path, sheet_name='Sheet1', head=True) -> list:
        """文件路径，sheet页"""
        import openpyxl
        myxls = openpyxl.load_workbook(file_path)
        activeSheet = myxls[sheet_name]
        if head:
            keys, xlsxData = list(), list()
            for row in range(1, activeSheet.max_row + 1):
                d = dict()
                for n, column in enumerate(range(1, activeSheet.max_column + 1)):
                    data = activeSheet.cell(row=row, column=column).value
                    if len(keys) < activeSheet.max_column:
                        keys.append(data)
                    else:
                        d[keys[n]] = data
                if d:
                    xlsxData.append(d)
        else:
            xlsxData = list()
            for row in range(1, activeSheet.max_row + 1):
                xlsxData.append(
                    [activeSheet.cell(row=row, column=column).value for column in range(1, activeSheet.max_column + 1)])
        return xlsxData

    # 写入xlsx
    def writeXlsx(self, file_path, data):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for n, row in enumerate(data):
            if isinstance(row, dict):
                if n == 0:
                    ws.append(list(row.keys()))
                ws.append(list(row.values()))
            else:
                ws.append(row)
        wb.save(file_path)

    # 读取sql语句
    def readSql(self, sql_path, parameter=None):
        with open(sql_path, 'r', encoding='utf-8')as f:
            sql = f.read()
        if isinstance(parameter, dict):
            sql = sql.format(**parameter)
        return sql

    # 日期处理
    def timeShift(self, strftime=None, **kwargs):
        '''
        支持：years, months, days, weeks, hours, minutes, seconds, microseconds
        例子： timeShift("%Y-%m-%d", days=1)
        '''
        from dateutil.relativedelta import relativedelta
        import datetime
        new_date = datetime.datetime.now() + relativedelta(**kwargs)
        if strftime:
            new_date = new_date.strftime(strftime)
        return new_date