# -*- coding: utf-8 -*-
"""
Created on Mon Dec 12 12:22:27 2022

@author: Franco, Bruno Agustín 

 AGREGAR DOCUMENTACION
 AGREGAR DOCUMENTACION
 AGREGAR DOCUMENTACION
"""

import os 
import pandas as pd

from time import gmtime, strftime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from openpyxl.drawing.image import Image

    #-----------------------------------------------------------------------
def gen_out_dp4(mode, inputs, parameters,  outputs, custom_mode=None): 
    '''generate output Excel file of a correlation calc
    Uses the input and output dictionaries for the construction
    At the end, adds signature and embellish the results using auxiliar functions
    '''
    #for programming porpuse it's recommended to get the output at desktop
    #desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    #cwf = os.getcwd().split('\\')[-1]   #current work folder 
    #time = strftime("%H%M", gmtime())
    #output_file = os.path.join(desktop,f'{cwf}_results_{time}.xlsx')
    
    if 'QM' in mode: 
        output_file = 'DP4plus_results.xlsx'
    elif 'MM' in mode: 
        output_file = 'MM-DP4plus_results.xlsx'
    elif 'Custom' in mode: 
        output_file = 'Custom_DP4plus_results.xlsx'
        
    with pd.ExcelWriter(output_file) as writer: 
    
        for sheet, matrix in outputs.items(): 
            if 'exp' in sheet : continue
            try: 
                matrix.to_excel(writer,sheet_name=sheet, index=True,float_format="%.2f", startrow=2)
            except: 
                matrix = pd.DataFrame(matrix,
                                      columns=inputs['isom'],
                                      index=outputs['exp'].index)
                matrix = pd.concat([outputs['exp'],matrix], axis=1)
                
                matrix.to_excel(writer,
                                sheet_name=sheet, 
                                index=True,
                                float_format="%.4f")
                
        if 'Custom' in mode: 
            custom_mode.to_excel(writer,sheet_name='results', index=True, startrow=18)
            parameters.to_excel(writer,sheet_name='results', index=True,float_format="%.2f", startrow=24)

    add_signs (output_file, 'results', mode ,  inputs)
                
    edit_appearance(output_file, 'results')
    
    return  

def add_signs(file, sheet, mode, inputs): 
    '''Adds signatures and sign to a given sheet in the .xlsx file
    It's design to embellish the 'results' sheet according the position 
    of its outputs. However, it's easy to modify for other presentations.
    
    Adds the signature in the first cell and the developer icon.
    At the end, adds labels of the theory level used in the calcs and warns 
    about G09 matching with the input. 
    '''
    wb = load_workbook(filename = file)
    ws = wb[sheet]
    
    ws.cell(1 , 1).value =f'DP4+ {mode}'
    
    #row = ws.max_row +3  #startrow
    row = 14  #startrow
    
    if 'Custom' not in mode: 
        inputs['the_lev'] = inputs['the_lev'].split('.')
        inputs['the_lev'] = '/'.join(inputs['the_lev'])
        
    ws.cell(row , 1).value = f'Theory Level selected : {inputs["the_lev"]}'
    ws.cell(row+1 , 1).value = f'G09 command lines: {inputs["G09command"]}'

    if 'Custom' in mode: 
        ws.cell(row+3 , 1).value = inputs['warn'].pop()
        
    elif inputs['warn']: 
        ws.cell(row+3 , 1).value ='¡Warning! Calcs and Theory Level do not match'
        ws.cell(row+4 , 1).value ='These are the inconsistencies:'
        for i,j in enumerate(inputs['warn'].most_common()):
            ws.cell(row+5+i , 1).value = '   x: '+j[0] 
    else: 
        ws.cell(row+3 , 1).value =u'Theory Level: OK \u2713'
    
    wb.save(file)
    
    return

def edit_appearance(file, sheet):
    '''Edits the appearance of the given sheet in the .xlsx file
    It's design to embellish the 'results' sheet according the position 
    of its outputs. However, it's easy to modify for other presentations. 
    '''
    wb = load_workbook(file)
    ws = wb[sheet]
    ws.column_dimensions['A'].width = 20
    for i in range(4, 7):
        ws[f'A{i}'].fill = PatternFill(start_color = 'DAF7A6', end_color = 'DAF7A6', fill_type = "solid")
    for i in range(7, 10):
        ws[f'A{i}'].fill = PatternFill(start_color = 'DAF7A6', end_color = 'DAF7A6', fill_type = "solid")
    for i in range(10, 13):
        ws[f'A{i}'].fill = PatternFill(start_color = '73C6B6', end_color = '73C6B6', fill_type = "solid")
    
    for row in range(4,13):
        ws[f'B{row}'].number_format = FORMAT_PERCENTAGE
        ws[f'C{row}'].number_format = FORMAT_PERCENTAGE
        
    for col in ['B','C','D']:
        ws[f'{col}14'].font = Font(name='Symbol')
    
    ws['A1'].font = Font(name='Lucida Sans', size=14, bold=True,
                         italic=False, vertAlign=None, underline='none',
                         strike=False, color='FF000000')
    
    #remember to: pip install Pillow (or add to dependencies)
    
    img1 =(Path(__file__).parent / 'logo_CONICET.png').as_posix()
    img1 = Image(img1)
    img1.height = img1.height/8
    img1.width= img1.width/8
    img1.anchor = 'E2'
    ws.add_image(img1)
    
    img2 =(Path(__file__).parent / 'logo_INGEBIO.png').as_posix()
    img2 = Image(img2)
    img2.height = img2.height/2
    img2.width= img2.width/2
    img2.anchor = 'E5'
    ws.add_image(img2)
    
    ws.sheet_view.showGridLines = False

    wb.save(file)
    return


    #-----------------------------------------------------------------------
    
def gen_out_custom(mode, name, C_TMS, H_TMS, param):
    ''' AGREGAR DOCUMENTACION
    AGREGAR DOCUMENTACION
    AGREGAR DOCUMENTACION
    '''
    data_base = (Path(__file__).parent / 'data_base_Custom.xlsx').as_posix()

    wb = load_workbook(filename = data_base)
    ws = wb['standard']
    levels = wb.sheetnames
    
    if name in levels: 
        for i in range (1, ws.max_row+1): 
            if name == ws.cell(i , 1).value :
                row = i
                wb.remove(wb[name])
                break
    else : 
        row = ws.max_row +1  #startrow
    
    ws.cell(row , 1).value = name
    ws.cell(row , 2).value = C_TMS
    ws.cell(row , 3).value = H_TMS
    ws.cell(row , 4).value = strftime("%m/%d/%Y", gmtime())
    ws.cell(row , 5).value = mode

    wb.save(data_base)
    
    
    with pd.ExcelWriter(data_base, engine = 'openpyxl', mode='a') as writer: 
            
        param.to_excel(writer, sheet_name= name)
    
    return  
    