from __future__ import annotations

import logging
from io import BytesIO
from typing import IO, Optional

from werkzeug.datastructures import FileStorage

import ckan.lib.munge as munge

from ..record import PackageRecord, ResourceRecord
from .base import ParsingExtras, ParsingStrategy

log = logging.getLogger(__name__)


class SeedExcelStrategy(ParsingStrategy):
    mimetypes = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

    def extract(self, source: FileStorage, extras: Optional[ParsingExtras] = None):
        from openpyxl import load_workbook

        doc = load_workbook(BytesIO(source.read()), read_only=True, data_only=True)

        md_name = "Dataset Metadata"
        res_name = "Resources"
        if md_name not in doc or res_name not in doc:
            log.warning(
                "Excel document does not contain '%s' or '%s' sheet",
                md_name,
                res_name,
            )
            return

        metadata_sheet = doc[md_name]
        resources_sheet = doc[res_name]

        rows = metadata_sheet.iter_rows(row_offset=1)
        data_dict = PackageRecord(_prepare_data_dict(rows))
        if not data_dict.data.get("name"):
            data_dict.data["name"] = munge.munge_title_to_name(data_dict.data["title"])

        yield data_dict

        for row in resources_sheet.iter_rows(row_offset=1):
            if not row[0].value:
                continue
            resource_title = row[0].value
            resource_from = row[1].value
            resource_format = row[2].value
            resource_desc = row[3].value

            if not resource_title:
                break

            if resource_from.startswith("http"):
                payload = {
                    "package_id": data_dict.data["name"],
                    "url": resource_from,
                    "name": resource_title,
                    "format": resource_format,
                    "description": resource_desc,
                }
            elif extras and "file_locator" in extras:
                fp = extras["file_locator"](resource_from)
                if not fp:
                    log.warning("Cannot locate file for resource %s", resource_title)
                    continue
                payload = {
                    "package_id": data_dict.data["name"],
                    # url must be provided, even for uploads
                    "url": resource_from,
                    "format": resource_format,
                    "name": resource_title,
                    "description": resource_desc,
                    "url_type": "upload",
                    "upload": FileStorage(fp, resource_from),
                }

            else:
                log.warning("Cannot determine source filesystem of %s", resource_title)
                continue

            yield ResourceRecord(payload)


def _prepare_data_dict(rows):
    """Parse .xlsx file and pushes data to dict."""
    raw = {}
    for row in rows:
        field = row[0].value
        value = row[1].value
        if not field:
            continue
        raw[field] = value

    return raw
