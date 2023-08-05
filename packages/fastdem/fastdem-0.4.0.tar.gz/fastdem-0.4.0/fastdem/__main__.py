#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time :   2023/2/10 14:22
# @Author   :   wanqiang.liu@freetech.com
# @File :   main.py
# @Software :   PyCharm 
# @Version  :   V1.0.0.0  
# @Description  :
# @config: poetry config
# @New: poetry new fastdem | cd ./fastdem
# @Install Package: poetry install
# @Build: poetry build
# @Test your wheel package: pip install --user c:/fastdem-0.1.0-py3-none-any.whl
# @install: pip install c:/fastdem-0.1.0-py3-none-any.whl
#           pip install c:/fastdem-0.1.0-py3-none-any.whl --force-reinstall
# @publish: poetry publish --build

import datetime
import os.path
import sys
import time
import typing
import xml
from xml.dom import minidom
import codecs

import numpy as np
import openpyxl
import pandas as pd

import typer
from rich import print
from typing import Optional
from typing import List, Any
from rich.progress import track, Progress, SpinnerColumn, TextColumn

"""
bit     | Intro
------------------------------------------------------------------------------------------
0         testFailed | 请求时刻测试结果为失败
          [0]:
          [1]:当前结果为故障状态
          --------------------------------------------------------------------------------
          通常来说,ECU内部以循环的方式不断地针对预先定义好的错误路径进行测试,如果在最近的一次测试中,在某个
          错误路径中发现了故障,则相应DTC的这一个状态位就要被置1,表征出错.此时DTC的testFailed位被置1,但
          是它不一定被ECU存储到non-volatile memory中,只有当pendingDTC或confirmedDTC被置1时DTC才会
          被存储.而pendingDTC或confirmedDTC被置1的条件应该是检测到错误出现的次数或时间满足某个预定义的门
          限.当错误消失或者诊断仪执行了清除DTC指令时,testFailed会再次被置为0.
          --------------------------------------------------------------------------------

1         testFailedThisOperationCycle | 在当前点火循环至少失败1次
          [0]:
          [1]:当前操作循环中至少检测到一次故障
          --------------------------------------------------------------------------------
          这个bit用于标识某个DTC在当前的operation cycle中是否出现过testFailed置1的情况,即是否出现过错误.
          operation cycle的起始点是ECU通过网络管理唤醒到ECU通过网络管理进入睡眠,对于没有网络管理的ECU,这个
          起始点就是KL15通断.通过bit 0我们无法判断某个DTC是否出现过,比如,当前testFailed = 0, 说明当前这个
          DTC没有出错,如果testFailedThisOperationCycle = 1的话,就说明这个DTC在当前这个operation cycle
          中出过错,但是当前错误又消失了.
          --------------------------------------------------------------------------------

2         pendingDTC | 在当前或者上一个点火循环测试结果不为失败
          [0]:
          [1]:当前操作循环或者上一个完成的操作循环期间至少检测到1次故障
          --------------------------------------------------------------------------------
          根据规范的解释,pendingDTC = 1表示某个DTC在当前或者上一个operation cycle中是否出现过.
          pendingDTC位其实是位于testFailed和confirmedDTC之间的一个状态,有的DTC被确认的判定条件比较严苛,
          需要在多个operation cycle中出现才可以被判定为confirmed的状态,此时就需要借助于pendingDTC位了.
          pendingDTC = 1的时候,DTC就要被存储下来了,如果接下来的两个operation cycle中这个DTC都还存在,那
          么confirmedDTC就要置1了.如果当前operation cycle中,故障发生,pendingDTC = 1,但是在下一个
          operation cycle中,故障没有了,pendingDTC 仍然为 1,再下一个operation cycle中,故障仍然不存在,
          那么pendingDTC 就可以置0了.
          --------------------------------------------------------------------------------

3         confirmedDTC | 请求时刻DTC被确认,一般确认是在一个点火周期内发生错误1次
          [0]:
          [1]:表示存在历史故障
          --------------------------------------------------------------------------------
          当confirmedDTC = 1时,则说明某个DTC已经被存储到ECU的non-volatile memory中,说明这个DTC曾经
          满足了被confirmed的条件.但是请注意,confirmedDTC = 1时,并不意味着当前这个DTC仍然出错,如果
          confirmedDTC = 1,但testFailed = 0,则说明这个DTC表示的故障目前已经消失了.将confirmedDTC 
          重新置0的方法只有删除DTC,UDS用0x14服务,OBD用0x04服务.
          --------------------------------------------------------------------------------

4         testNotCompleteSinceLastClear | 自上次清除DTC之后测试结果已完成,即测试结果为PASS或者FAIL
          [0]:自从清理DTC之后已经完成过针对该DTC的测试
          [1]:表示从上次进行清除诊断信息后,DTC检测尚未完成
          --------------------------------------------------------------------------------
          这个bit用于标识,自从上次调用了清理DTC的服务（UDS用0x14服务,OBD用0x04服务）之后,是否成功地执行了
          对某个DTC的测试（不管测试结果是什么,只关心是否测了）.因为很多DTC的测试也是需要满足某些边界条件的,并
          不是ECU上电就一定会对DTC进行检测.
          --------------------------------------------------------------------------------

5         testFailedSinceLastClear | 自上次清除DTC后测试结果都不是FAIL
          [0]:自从清理DTC之后该DTC没有出过错
          [1]:自从清理DTC之后该DTC出过至少一次错
          --------------------------------------------------------------------------------
          这个位与bit 1 :testFailedThisOperationCycle有些类似,后者标识某个DTC在当前的operation cycle
          中是否出现过testFailed置1的情况,而testFailedSinceLastClear标识的是在上次执行过清理DTC之后某个
          DTC是否出过错.
          --------------------------------------------------------------------------------

6         testNotCompletedThisOperationCycle | 在当前点火周期内测试结果已完成,即为PASS或FAIL状态
          [0]:在当前operation cycle中已经完成过针对该DTC的测试
          [1]:在当前operation cycle中还没在完成过针对该DTC的测试
          --------------------------------------------------------------------------------
          这个位与bit 4 : testNotCompletedSinceLastClear类似,后者标识自从上次调用了清理DTC的服务之后,
          是否成功地执行了对某个DTC的测试.而testNotCompletedThisOperationCycle则标识在当前operation
           cycle中是否成功地执行了对某个DTC的测试.
          --------------------------------------------------------------------------------

7         warningIndicatorRequested | ECU没有得到点亮警示灯请求
          [0]:ECU不请求激活警告指示
          [1]:表示该bit关联的特定DTC警告指示灯亮
          --------------------------------------------------------------------------------
          某些比较严重的DTC会与用户可见的警告指示相关联,比如仪表上的报警灯,或者是文字,或者是声音.这个
          warningIndicatorRequested就用于此类DTC.
          --------------------------------------------------------------------------------
"""


class ISO14229DTCSTATUS:
    @staticmethod
    def _bit0(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    testFailed')
        else:
            print(f'{flt_mask:02X}:                   当前结果 [不为] 故障状态')

    @staticmethod
    def _bit1(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    testFailedThisOperationCycle')
        else:
            print(f'{flt_mask:02X}:                   当前操作循环中 [没有] 检测到一次故障')

    @staticmethod
    def _bit2(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    pendingDTC')
        else:
            print(f'{flt_mask:02X}:                   当前操作循环或者上一个完成的操作循环期间 [没有] 检测到1次故障')

    @staticmethod
    def _bit3(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    confirmedDTC')
        else:
            print(f'{flt_mask:02X}:                   表示 [不存在] 历史故障')

    @staticmethod
    def _bit4(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    testNotCompleteSinceLastClear')
        else:
            print(f'{flt_mask:02X}:                   自从清理DTC之后 [已经完成] 过针对该DTC的测试')

    @staticmethod
    def _bit5(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    testFailedSinceLastClear')
        else:
            print(f'{flt_mask:02X}:                   自从清理DTC之后该DTC [没有] 出过错')

    @staticmethod
    def _bit6(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    testNotCompletedThisOperationCycle')
        else:
            print(f'{flt_mask:02X}:                   在当前operation cycle中 [已经完成] 过针对该DTC的测试')

    @staticmethod
    def _bit7(tf: bool, flt_mask: int):
        if tf is True:
            print(f'{flt_mask:02X}:           TRUE    warningIndicatorRequested')
        else:
            print(f'{flt_mask:02X}:                   ECU [不请求] 激活警告指示')

    @staticmethod
    def _default(tf: bool, flt_mask: int):
        print(f'{flt_mask:02X}:')
        pass

    @staticmethod
    def _bit(bit_func, tf: bool, flt_mask: int):
        return bit_func(tf, flt_mask)

    def set_status_code(self, b_faultStatus):
        switch_dict = {
            # mask | func_point
            0x01: self._bit0,
            0x02: self._bit1,
            0x04: self._bit2,
            0x08: self._bit3,
            0x10: self._bit4,
            0x20: self._bit5,
            0x40: self._bit6,
            0x80: self._bit7
        }
        # Notes: format(14,'#b') <=> f'{num:#b}'  =>> 0b1110
        #        format(14,'b') <=> f'{num:b}'    =>>   1110
        #        bin(14) , bin(-10)
        b_faultStatus = int(b_faultStatus.replace('0x', '').replace('0X', ''), 16)
        print(f'HEX: {b_faultStatus:02X}  |  Status  |  BIN: {b_faultStatus:08b}')
        print('----------------------------------------------------------------')
        masks = [0x80, 0x40, 0x20, 0x10, 0x08, 0x04, 0x02, 0x01]
        for mask in reversed(masks):
            tf = b_faultStatus & mask
            p_func = switch_dict.get(mask, self._default)
            p_func(bool(tf), mask)


class PrototypeComposition:
    def __init__(self, ComponentPrototype: str, ConnectorName: str, PortPath: str, Port: str, ConnectedComposition: str, ConnectedPortPath: str, ConnectedPort: str):
        self._itemComponentPrototype = ComponentPrototype.split("/")[-1]  # /Composition/Composition/
        self._itemConnectorName = ConnectorName
        self._itemPortPath = PortPath[:PortPath.rindex("/") + 1]
        self._itemPort = Port.split("/")[-1]
        self._itemConnectedComposition = ConnectedComposition.split("/")[-1]  # /Composition/Composition
        self._itemConnectedPortPath = ConnectedPortPath[:ConnectedPortPath.rindex("/") + 1]
        self._itemConnectedPort = ConnectedPort.split("/")[-1]
        self._path_math_port = dict()

    @property
    def get_path_math_port(self):
        self._path_math_port = {
            "PPortCom": "/Composition/Composition/",
            "RPortCom": "/Composition/Composition/",
            "PP_Event": "/AUTOSAR_Dem/SwComponentTypes/Dem/",
            "PP_EvtInfo": "/AUTOSAR_Dem/SwComponentTypes/Dem/",
            "PP_FltM": "/Pkg_SWCs/FltM_SWC/",
            "PP_EnableCondition": "/AUTOSAR_Dem/SwComponentTypes/Dem/",
        }
        return self._path_math_port

    @property
    def get_itemComponentPrototype(self):
        return self._itemComponentPrototype

    def set_itemComponentPrototype(self, itemComponentPrototype):
        self._itemComponentPrototype = itemComponentPrototype

    @property
    def get_itemConnectorName(self):
        return self._itemConnectorName

    def set_itemConnectorName(self, itemConnectorName):
        self._itemConnectorName = itemConnectorName

    @property
    def get_itemPortPath(self):
        return self._itemPortPath

    def set_itemPortPath(self, itemPortPath):
        self._itemPortPath = itemPortPath

    @property
    def get_itemPort(self):
        return self._itemPort

    def set_itemPort(self, itemPort):
        self._itemPort = itemPort

    @property
    def get_itemConnectedComposition(self):
        return self._itemConnectedComposition

    def set_itemConnectedComposition(self, itemConnectedComposition):
        self._itemConnectedComposition = itemConnectedComposition

    @property
    def get_itemConnectedPortPath(self):
        return self._itemConnectedPortPath

    def set_itemConnectedPortPath(self, itemConnectedPortPath):
        self._itemConnectedPortPath = itemConnectedPortPath

    @property
    def get_itemConnectedPort(self):
        return self._itemConnectedPort

    def set_itemConnectedPort(self, itemConnectedPort):
        self._itemConnectedPort = itemConnectedPort


class PrototypeFltMREServerCallPoint:
    def __init__(self, ServerCallPoint, ShortName, OperationPath: str, Operation: str, ClientPort: str):
        self._itemServerCallPoint = ServerCallPoint
        self._itemShortName = ShortName
        self._itemOperationPath = OperationPath.split("/")[-2]  # /AUTOSAR_Dem/PortInterfaces/ + OperationPath
        self._itemOperation = Operation.split("/")[-1]  # /AUTOSAR_Dem/PortInterfaces/ + OperationPath / + Operation
        self._itemClientPort = ClientPort.split("/")[-1]  # /Pkg_SWCs/FltM_SWC/

    @property
    def get_itemServerCallPoint(self):
        return self._itemServerCallPoint

    def set_itemServerCallPoint(self, itemServerCallPoint):
        self._itemServerCallPoint = itemServerCallPoint

    @property
    def get_itemShortName(self):
        return self._itemShortName

    def set_itemShortName(self, itemShortName):
        self._itemShortName = itemShortName

    @property
    def get_itemOperationPath(self):
        return self._itemOperationPath

    def set_itemOperationPath(self, itemOperationPath):
        self._itemOperationPath = itemOperationPath

    @property
    def get_itemOperation(self):
        return self._itemOperation

    def set_itemOperation(self, itemOperation):
        self._itemOperation = itemOperation

    @property
    def get_itemClientPort(self):
        return self._itemClientPort

    def set_itemClientPort(self, itemClientPort):
        self._itemClientPort = itemClientPort


class PrototypeFltMEvents:
    def __init__(self, RteEventType, EventName, StartRunnableEntity, TimingPeriod, EventTarget, EventPortPath: str, EventPort):
        self._itemRteEventType = RteEventType
        self._itemEventName = EventName
        self._itemStartRunnableEntity = str(StartRunnableEntity).split("/")[-1]  # /Pkg_SWCs/FltM_SWC/IB_FltM_SWC/
        self._itemTimingPeriod = TimingPeriod
        self._itemEventTarget = str(EventTarget).split("/")[-1]  # /Pkg_SWCs/FltM_SWC/
        self._itemEventPortPath = EventPortPath.split("/")[-2]  # /AUTOSAR_Dem/PortInterfaces/ + EventPortPath.split("/")[-2]
        self._itemEventPort = str(EventPort).split("/")[-1]

    @property
    def get_itemRteEventType(self):
        return self._itemRteEventType

    def set_itemRteEventType(self, itemRteEventType):
        self._itemRteEventType = itemRteEventType

    @property
    def get_itemEventName(self):
        return self._itemEventName

    def set_itemEventName(self, itemEventName):
        self._itemEventName = itemEventName

    @property
    def get_itemStartRunnableEntity(self):
        return self._itemStartRunnableEntity

    def set_itemStartRunnableEntity(self, itemStartRunnableEntity):
        self._itemStartRunnableEntity = itemStartRunnableEntity

    @property
    def get_itemTimingPeriod(self):
        return self._itemTimingPeriod

    def set_itemTimingPeriod(self, itemTimingPeriod):
        self._itemTimingPeriod = itemTimingPeriod

    @property
    def get_itemEventTarget(self):
        return self._itemEventTarget

    def set_itemEventTarget(self, itemEventTarget):
        self._itemEventTarget = itemEventTarget

    @property
    def get_itemEventPortPath(self):
        return self._itemEventPortPath

    def set_itemEventPortPath(self, itemEventPortPath):
        self._itemEventPortPath = itemEventPortPath

    @property
    def get_itemEventPort(self):
        return self._itemEventPort

    def set_itemEventPort(self, itemEventPort):
        self._itemEventPort = itemEventPort


class PrototypeFltMRE:
    def __init__(self, RunnableEntity, FunctionName):
        self._itemRunnableEntity = RunnableEntity
        self._itemFunctionName = FunctionName

    @property
    def get_itemRunnableEntity(self):
        return self._itemRunnableEntity

    def set_itemRunnableEntity(self, itemRunnableEntity):
        self._itemRunnableEntity = itemRunnableEntity

    @property
    def get_itemFunctionName(self):
        return self._itemFunctionName

    def set_itemFunctionName(self, itemFunctionName):
        self._itemFunctionName = itemFunctionName


class PrototypeFltMPRPort:
    def __init__(self, PortType, ShortName, Interface, PortInterfacePath, PortInterface, IsService):
        self._itemPortType = PortType
        self._itemShortName = ShortName
        self._itemInterface = Interface
        self._itemPortInterfacePath = PortInterfacePath
        self._itemPortInterface = PortInterface
        self._itemIsService = IsService

    @property
    def get_itemPortInterfacePath(self):
        return self._itemPortInterfacePath

    def set_itemPortInterfacePath(self, itemPortInterfacePath):
        self._itemPortInterfacePath = itemPortInterfacePath

    @property
    def get_itemPortType(self):
        return self._itemPortType

    def set_itemPortType(self, itemPortType):
        self._itemPortType = itemPortType

    @property
    def get_itemShortName(self):
        return self._itemShortName

    def set_itemShortName(self, itemShortName):
        self._itemShortName = itemShortName

    @property
    def get_itemInterface(self):
        return self._itemInterface

    def set_itemInterface(self, itemInterface):
        self._itemInterface = itemInterface

    @property
    def get_itemPortInterface(self):
        return self._itemPortInterface

    def set_itemPortInterface(self, itemPortInterface):
        self._itemPortInterface = itemPortInterface

    @property
    def get_itemIsService(self):
        return self._itemIsService

    def set_itemIsService(self, itemIsService):
        self._itemIsService = itemIsService


class PrototypeDemDebounceCounterBasedClasss:
    def __init__(self, DemDebounceCounterBasedClass, DemDebounceBehavior, DemDebounceCounterDecrementStepSize, DemDebounceCounterFailedThreshold, DemDebounceCounterIncrementStepSize,
                 DemDebounceCounterJumpDown, DemDebounceCounterJumpDownValue, DemDebounceCounterJumpUp, DemDebounceCounterJumpUpValue, DemDebounceCounterPassedThreshold,
                 DemDebounceCounterStorage, DemRbDebounceCounterFdcThresholdStorageValue, DemRbDebounceCounterJumpDownAlternative, DemRbDebounceCounterJumpUpAlternative):
        self._itemDemDebounceCounterBasedClass = DemDebounceCounterBasedClass
        self._itemDemDebounceBehavior = str(DemDebounceBehavior).upper()
        self._itemDemDebounceCounterDecrementStepSize = int(DemDebounceCounterDecrementStepSize)
        self._itemDemDebounceCounterFailedThreshold = int(DemDebounceCounterFailedThreshold)
        self._itemDemDebounceCounterIncrementStepSize = int(DemDebounceCounterIncrementStepSize)
        self._itemDemDebounceCounterJumpDown = str(DemDebounceCounterJumpDown).upper()
        self._itemDemDebounceCounterJumpDownValue = int(DemDebounceCounterJumpDownValue)
        self._itemDemDebounceCounterJumpUp = str(DemDebounceCounterJumpUp).upper()
        self._itemDemDebounceCounterJumpUpValue = int(DemDebounceCounterJumpUpValue)
        self._itemDemDebounceCounterPassedThreshold = int(DemDebounceCounterPassedThreshold)  # 负值
        self._itemDemDebounceCounterStorage = str(DemDebounceCounterStorage).upper()
        self._itemDemRbDebounceCounterFdcThresholdStorageValue = int(DemRbDebounceCounterFdcThresholdStorageValue)
        self._itemDemRbDebounceCounterJumpDownAlternative = str(DemRbDebounceCounterJumpDownAlternative).upper()
        self._itemDemRbDebounceCounterJumpUpAlternative = str(DemRbDebounceCounterJumpUpAlternative).upper()

    @property
    def get_itemDemDebounceCounterJumpDownValue(self):
        return self._itemDemDebounceCounterJumpDownValue

    def set_itemDemDebounceCounterJumpDownValue(self, itemDemDebounceCounterJumpDownValue):
        self._itemDemDebounceCounterJumpDownValue = itemDemDebounceCounterJumpDownValue

    @property
    def get_itemDemDebounceCounterBasedClass(self):
        return self._itemDemDebounceCounterBasedClass

    def set_itemDemDebounceCounterBasedClass(self, itemDemDebounceCounterBasedClass):
        self._itemDemDebounceCounterBasedClass = itemDemDebounceCounterBasedClass

    @property
    def get_itemDemDebounceBehavior(self):
        return self._itemDemDebounceBehavior

    def set_itemDemDebounceBehavior(self, itemDemDebounceBehavior):
        self._itemDemDebounceBehavior = itemDemDebounceBehavior

    @property
    def get_itemDemDebounceCounterDecrementStepSize(self):
        return self._itemDemDebounceCounterDecrementStepSize

    def set_itemDemDebounceCounterDecrementStepSize(self, itemDemDebounceCounterDecrementStepSize):
        self._itemDemDebounceCounterDecrementStepSize = itemDemDebounceCounterDecrementStepSize

    @property
    def get_itemDemDebounceCounterFailedThreshold(self):
        return self._itemDemDebounceCounterFailedThreshold

    def set_itemDemDebounceCounterFailedThreshold(self, itemDemDebounceCounterFailedThreshold):
        self._itemDemDebounceCounterFailedThreshold = itemDemDebounceCounterFailedThreshold

    @property
    def get_itemDemDebounceCounterIncrementStepSize(self):
        return self._itemDemDebounceCounterIncrementStepSize

    def set_itemDemDebounceCounterIncrementStepSize(self, itemDemDebounceCounterIncrementStepSize):
        self._itemDemDebounceCounterIncrementStepSize = itemDemDebounceCounterIncrementStepSize

    @property
    def get_itemDemDebounceCounterJumpDown(self):
        return self._itemDemDebounceCounterJumpDown

    def set_itemDemDebounceCounterJumpDown(self, itemDemDebounceCounterJumpDown):
        self._itemDemDebounceCounterJumpDown = itemDemDebounceCounterJumpDown

    @property
    def get_itemDemDebounceCounterJumpUp(self):
        return self._itemDemDebounceCounterJumpUp

    def set_itemDemDebounceCounterJumpUp(self, itemDemDebounceCounterJumpUp):
        self._itemDemDebounceCounterJumpUp = itemDemDebounceCounterJumpUp

    @property
    def get_itemDemDebounceCounterJumpUpValue(self):
        return self._itemDemDebounceCounterJumpUpValue

    def set_itemDemDebounceCounterJumpUpValue(self, itemDemDebounceCounterJumpUpValue):
        self._itemDemDebounceCounterJumpUpValue = itemDemDebounceCounterJumpUpValue

    @property
    def get_itemDemDebounceCounterPassedThreshold(self):
        return self._itemDemDebounceCounterPassedThreshold

    def set_itemDemDebounceCounterPassedThreshold(self, itemDemDebounceCounterPassedThreshold):
        self._itemDemDebounceCounterPassedThreshold = itemDemDebounceCounterPassedThreshold

    @property
    def get_itemDemDebounceCounterStorage(self):
        return self._itemDemDebounceCounterStorage

    def set_itemDemDebounceCounterStorage(self, itemDemDebounceCounterStorage):
        self._itemDemDebounceCounterStorage = itemDemDebounceCounterStorage

    @property
    def get_itemDemRbDebounceCounterFdcThresholdStorageValue(self):
        return self._itemDemRbDebounceCounterFdcThresholdStorageValue

    def set_itemDemRbDebounceCounterFdcThresholdStorageValue(self, itemDemRbDebounceCounterFdcThresholdStorageValue):
        self._itemDemRbDebounceCounterFdcThresholdStorageValue = itemDemRbDebounceCounterFdcThresholdStorageValue

    @property
    def get_itemDemRbDebounceCounterJumpDownAlternative(self):
        return self._itemDemRbDebounceCounterJumpDownAlternative

    def set_itemDemRbDebounceCounterJumpDownAlternative(self, itemDemRbDebounceCounterJumpDownAlternative):
        self._itemDemRbDebounceCounterJumpDownAlternative = itemDemRbDebounceCounterJumpDownAlternative

    @property
    def get_itemDemRbDebounceCounterJumpUpAlternative(self):
        return self._itemDemRbDebounceCounterJumpUpAlternative

    def set_itemDemRbDebounceCounterJumpUpAlternative(self, itemDemRbDebounceCounterJumpUpAlternative):
        self._itemDemRbDebounceCounterJumpUpAlternative = itemDemRbDebounceCounterJumpUpAlternative


class PrototypeDemEventParameters:
    def __init__(self, DemEventParameter, DemEventAvailable, DemEventFailureCycleCounterThreshold, DemEventKind, DemFFPrestorageSupported,
                 DemDTCRef, DemEnableConditionGroupRef, DemOperationCycleRef, DemDebounceCounterBasedClassRef):
        self._itemDemEventParameter = DemEventParameter
        self._itemDemEventAvailable = DemEventAvailable
        self._itemDemEventFailureCycleCounterThreshold = int(DemEventFailureCycleCounterThreshold)
        self._itemDemEventKind = DemEventKind
        self._itemDemFFPrestorageSupported = DemFFPrestorageSupported
        self._itemDemDTCRef = str(DemDTCRef).split('/')[-1]
        self._itemDemEnableConditionGroupRef = str(DemEnableConditionGroupRef).split('/')[-1]
        self._itemDemOperationCycleRef = str(DemOperationCycleRef).split('/')[-1]
        self._itemDemDebounceCounterBasedClassRef = str(DemDebounceCounterBasedClassRef).split('/')[-1]

    @property
    def get_itemDemEventParameter(self):
        return self._itemDemEventParameter

    def set_itemDemEventParameter(self, itemDemEventParameter):
        self._itemDemEventParameter = itemDemEventParameter

    @property
    def get_itemDemEventAvailable(self):
        return self._itemDemEventAvailable

    def set_itemDemEventAvailable(self, itemDemEventAvailable):
        self._itemDemEventAvailable = itemDemEventAvailable

    @property
    def get_itemDemEventFailureCycleCounterThreshold(self):
        return self._itemDemEventFailureCycleCounterThreshold

    def set_itemDemEventFailureCycleCounterThreshold(self, itemDemEventFailureCycleCounterThreshold):
        self._itemDemEventFailureCycleCounterThreshold = itemDemEventFailureCycleCounterThreshold

    @property
    def get_itemDemEventKind(self):
        return self._itemDemEventKind

    def set_itemDemEventKind(self, itemDemEventKind):
        self._itemDemEventKind = itemDemEventKind

    @property
    def get_itemDemFFPrestorageSupported(self):
        return self._itemDemFFPrestorageSupported

    def set_itemDemFFPrestorageSupported(self, itemDemFFPrestorageSupported):
        self._itemDemFFPrestorageSupported = itemDemFFPrestorageSupported

    @property
    def get_itemDemDTCRef(self):
        return self._itemDemDTCRef

    def set_itemDemDTCRef(self, itemDemDTCRef):
        self._itemDemDTCRef = itemDemDTCRef

    @property
    def get_itemDemEnableConditionGroupRef(self):
        return self._itemDemEnableConditionGroupRef

    def set_itemDemEnableConditionGroupRef(self, itemDemEnableConditionGroupRef):
        self._itemDemEnableConditionGroupRef = itemDemEnableConditionGroupRef

    @property
    def get_itemDemOperationCycleRef(self):
        return self._itemDemOperationCycleRef

    def set_itemDemOperationCycleRef(self, itemDemOperationCycleRef):
        self._itemDemOperationCycleRef = itemDemOperationCycleRef

    @property
    def get_itemDemDebounceCounterBasedClassRef(self):
        return self._itemDemDebounceCounterBasedClassRef

    def set_itemDemDebounceCounterBasedClassRef(self, itemDemDebounceCounterBasedClassRef):
        self._itemDemDebounceCounterBasedClassRef = itemDemDebounceCounterBasedClassRef


class PrototypeDemExtendedDataRecordClasss:
    def __init__(self, DemExtendedDataRecordClass, DemExtendedDataRecordNumber, DemExtendedDataRecordTrigger, DemExtendedDataRecordUpdate, DemDataElementClassRef):
        self._itemDemExtendedDataRecordClass = DemExtendedDataRecordClass
        self._itemDemExtendedDataRecordNumber = hex(int(DemExtendedDataRecordNumber)).upper()
        self._itemDemExtendedDataRecordTrigger = DemExtendedDataRecordTrigger
        self._itemDemExtendedDataRecordUpdate = DemExtendedDataRecordUpdate
        self._itemDemDataElementClassRef = str(DemDataElementClassRef).split('/')[-1]

    @property
    def get_itemDemExtendedDataRecordClass(self):
        return self._itemDemExtendedDataRecordClass

    def set_itemDemExtendedDataRecordClass(self, itemDemExtendedDataRecordClass):
        self._itemDemExtendedDataRecordClass = itemDemExtendedDataRecordClass

    @property
    def get_itemDemExtendedDataRecordNumber(self):
        return self._itemDemExtendedDataRecordNumber

    def set_itemDemExtendedDataRecordNumber(self, itemDemExtendedDataRecordNumber):
        self._itemDemExtendedDataRecordNumber = itemDemExtendedDataRecordNumber

    @property
    def get_itemDemExtendedDataRecordTrigger(self):
        return self._itemDemExtendedDataRecordTrigger

    def set_itemDemExtendedDataRecordTrigger(self, itemDemExtendedDataRecordTrigger):
        self._itemDemExtendedDataRecordTrigger = itemDemExtendedDataRecordTrigger

    @property
    def get_itemDemExtendedDataRecordUpdate(self):
        return self._itemDemExtendedDataRecordUpdate

    def set_itemDemExtendedDataRecordUpdate(self, itemDemExtendedDataRecordUpdate):
        self._itemDemExtendedDataRecordUpdate = itemDemExtendedDataRecordUpdate

    @property
    def get_itemDemDataElementClassRef(self):
        return self._itemDemDataElementClassRef

    def set_itemDemDataElementClassRef(self, itemDemDataElementClassRef):
        self._itemDemDataElementClassRef = itemDemDataElementClassRef


class PrototypeDemExtendedDataClasss:
    def __init__(self, DemExtendedDataClass, DemExtendedDataRecordClassRef):
        self._itemDemExtendedDataClass = DemExtendedDataClass
        self._itemDemExtendedDataRecordClassRef = str(DemExtendedDataRecordClassRef).replace('[', '').replace(']', '').replace("'", '').replace(',', '\n').replace(' ', '')

    @property
    def get_itemDemExtendedDataClass(self):
        return self._itemDemExtendedDataClass

    def set_itemDemExtendedDataClass(self, itemDemExtendedDataClass):
        self._itemDemExtendedDataClass = itemDemExtendedDataClass

    @property
    def get_itemDemExtendedDataRecordClassRef(self):
        return self._itemDemExtendedDataRecordClassRef

    def set_itemDemExtendedDataRecordClassRef(self, itemDemExtendedDataRecordClassRef):
        self._itemDemExtendedDataRecordClassRef = itemDemExtendedDataRecordClassRef


class PrototypeDemFreezeFrameClasss:
    def __init__(self, DemFreezeFrameClass, DemDidClassRef):
        self._itemDemFreezeFrameClass = DemFreezeFrameClass
        self._itemDemDidClassRef = str(DemDidClassRef).replace('[', '').replace(']', '').replace("'", '').replace(',', '\n').replace(' ', '')

    @property
    def get_itemDemFreezeFrameClass(self):
        return self._itemDemFreezeFrameClass

    def set_itemDemFreezeFrameClass(self, itemDemFreezeFrameClass):
        self._itemDemFreezeFrameClass = itemDemFreezeFrameClass

    @property
    def get_itemDemDidClassRef(self):
        return self._itemDemDidClassRef

    def set_itemDemDidClassRef(self, itemDemDidClassRef):
        self._itemDemDidClassRef = itemDemDidClassRef


class PrototypeDemFreezeFrameRecordClasss:
    def __init__(self, DemFreezeFrameRecordClass, DemFreezeFrameRecordNumber, DemFreezeFrameRecordTrigger, DemFreezeFrameRecordUpdate):
        self._itemDemFreezeFrameRecordClass = DemFreezeFrameRecordClass
        self._itemDemFreezeFrameRecordNumber = hex(int(DemFreezeFrameRecordNumber)).upper()
        self._itemDemFreezeFrameRecordTrigger = DemFreezeFrameRecordTrigger
        self._itemDemFreezeFrameRecordUpdate = DemFreezeFrameRecordUpdate

    @property
    def get_itemDemFreezeFrameRecordClass(self):
        return self._itemDemFreezeFrameRecordClass

    def set_itemDemFreezeFrameRecordClass(self, itemDemFreezeFrameRecordClass):
        self._itemDemFreezeFrameRecordClass = itemDemFreezeFrameRecordClass

    @property
    def get_itemDemFreezeFrameRecordNumber(self):
        return self._itemDemFreezeFrameRecordNumber

    def set_itemDemFreezeFrameRecordNumber(self, itemDemFreezeFrameRecordNumber):
        self._itemDemFreezeFrameRecordNumber = itemDemFreezeFrameRecordNumber

    @property
    def get_itemDemFreezeFrameRecordTrigger(self):
        return self._itemDemFreezeFrameRecordTrigger

    def set_itemDemFreezeFrameRecordTrigger(self, itemDemFreezeFrameRecordTrigger):
        self._itemDemFreezeFrameRecordTrigger = itemDemFreezeFrameRecordTrigger

    @property
    def get_itemDemFreezeFrameRecordUpdate(self):
        return self._itemDemFreezeFrameRecordUpdate

    def set_itemDemFreezeFrameRecordUpdate(self, itemDemFreezeFrameRecordUpdate):
        self._itemDemFreezeFrameRecordUpdate = itemDemFreezeFrameRecordUpdate


class PrototypeDemFreezeFrameRecNumClasss:
    def __init__(self, DemFreezeFrameRecNumClass, DemFreezeFrameRecordClassRef):
        self._itemDemFreezeFrameRecNumClass = DemFreezeFrameRecNumClass
        self._itemDemFreezeFrameRecordClassRef = str(DemFreezeFrameRecordClassRef).replace('[', '').replace(']', '').replace("'", '').replace(',', '\n').replace(' ', '')

    @property
    def get_itemDemFreezeFrameRecNumClass(self):
        return self._itemDemFreezeFrameRecNumClass

    def set_itemDemFreezeFrameRecNumClass(self, itemDemFreezeFrameRecNumClass):
        self._itemDemFreezeFrameRecNumClass = itemDemFreezeFrameRecNumClass

    @property
    def get_itemDemFreezeFrameRecordClassRef(self):
        return self._itemDemFreezeFrameRecordClassRef

    def set_itemDemFreezeFrameRecordClassRef(self, itemDemFreezeFrameRecordClassRef):
        self._itemDemFreezeFrameRecordClassRef = itemDemFreezeFrameRecordClassRef


class PrototypeDemEnableConditionGroups:
    def __init__(self, DemEnableConditionGroup, DemEnableConditionRef):
        self._itemEnableConditionGroup = DemEnableConditionGroup
        self._itemEnableConditionRef = str(DemEnableConditionRef).replace('[', '').replace(']', '').replace("'", '').replace(',', '\n').replace(' ', '')

    @property
    def get_itemEnableConditionGroup(self):
        return self._itemEnableConditionGroup

    def set_itemEnableConditionGroup(self, itemEnableConditionGroup):
        self._itemEnableConditionGroup = itemEnableConditionGroup

    @property
    def get_itemEnableConditionRef(self):
        return self._itemEnableConditionRef

    def set_itemEnableConditionRef(self, itemEnableConditionRef):
        self._itemEnableConditionRef = itemEnableConditionRef


class PrototypeDemEnableConditions:
    def __init__(self, DemEnableCondition, DemEnableConditionStatus: str = 'false'):
        self._itemEnaCondition = DemEnableCondition
        self._itemConditionStatus = DemEnableConditionStatus

    @property
    def get_itemEnaCondition(self):
        return self._itemEnaCondition

    def set_itemEnaCondition(self, itemEnaCondition):
        self._itemEnaCondition = itemEnaCondition

    @property
    def get_itemConditionStatus(self):
        return self._itemConditionStatus

    def set_itemConditionStatus(self, itemConditionStatus):
        self._itemConditionStatus = itemConditionStatus


class PrototypeDemDTCAttributess:
    def __init__(self, DemDTCAttributes, DemAgingAllowed, DemAgingCycleCounterThreshold,
                 DemDTCPriority, DemDTCSignificance, DemImmediateNvStorage,
                 DemMaxNumberFreezeFrameRecords, DemAgingCycleRef, DemExtendedDataClassRef,
                 DemFreezeFrameClassRef, DemFreezeFrameRecNumClassRef, DemMemoryDestinationRef):
        self._itemDTCAttributes = DemDTCAttributes
        self._itemAgingAllowed = DemAgingAllowed
        self._itemAgingCycleCounterThreshold = DemAgingCycleCounterThreshold
        self._itemDTCPriority = DemDTCPriority
        self._itemDTCSignificance = DemDTCSignificance
        self._itemImmediateNvStorage = DemImmediateNvStorage
        self._itemMaxNumberFreezeFrameRecords = DemMaxNumberFreezeFrameRecords
        self._itemAgingCycleRef = str(DemAgingCycleRef).split('/')[-1]
        self._itemExtendedDataClassRef = str(DemExtendedDataClassRef).split('/')[-1]
        self._itemFreezeFrameClassRef = str(DemFreezeFrameClassRef).split('/')[-1]
        self._itemFreezeFrameRecNumClassRef = str(DemFreezeFrameRecNumClassRef).split('/')[-1]
        self._itemMemoryDestinationRef = str(DemMemoryDestinationRef).split('/')[-1]

    @property
    def get_itemDTCAttributes(self):
        return self._itemDTCAttributes

    def set_itemDTCAttributes(self, itemDTCAttributes):
        self._itemDTCAttributes = itemDTCAttributes

    @property
    def get_itemAgingAllowed(self):
        return self._itemAgingAllowed

    def set_itemAgingAllowed(self, itemAgingAllowed):
        self._itemAgingAllowed = itemAgingAllowed

    @property
    def get_itemAgingCycleCounterThreshold(self):
        return self._itemAgingCycleCounterThreshold

    def set_itemAgingCycleCounterThreshold(self, itemAgingCycleCounterThreshold):
        self._itemAgingCycleCounterThreshold = itemAgingCycleCounterThreshold

    @property
    def get_itemDTCPriority(self):
        return self._itemDTCPriority

    def set_itemDTCPriority(self, itemDTCPriority):
        self._itemDTCPriority = itemDTCPriority

    @property
    def get_itemDTCSignificance(self):
        return self._itemDTCSignificance

    def set_itemDTCSignificance(self, itemDTCSignificance):
        self._itemDTCSignificance = itemDTCSignificance

    @property
    def get_itemImmediateNvStorage(self):
        return self._itemImmediateNvStorage

    def set_itemImmediateNvStorage(self, itemImmediateNvStorage):
        self._itemImmediateNvStorage = itemImmediateNvStorage

    @property
    def get_itemMaxNumberFreezeFrameRecords(self):
        return self._itemMaxNumberFreezeFrameRecords

    def set_itemMaxNumberFreezeFrameRecords(self, itemMaxNumberFreezeFrameRecords):
        self._itemMaxNumberFreezeFrameRecords = itemMaxNumberFreezeFrameRecords

    @property
    def get_itemAgingCycleRef(self):
        return self._itemAgingCycleRef

    def set_itemAgingCycleRef(self, itemAgingCycleRef):
        self._itemAgingCycleRef = itemAgingCycleRef

    @property
    def get_itemExtendedDataClassRef(self):
        return self._itemExtendedDataClassRef

    def set_itemExtendedDataClassRef(self, itemExtendedDataClassRef):
        self._itemExtendedDataClassRef = itemExtendedDataClassRef

    @property
    def get_itemFreezeFrameClassRef(self):
        return self._itemFreezeFrameClassRef

    def set_itemFreezeFrameClassRef(self, itemFreezeFrameClassRef):
        self._itemFreezeFrameClassRef = itemFreezeFrameClassRef

    @property
    def get_itemFreezeFrameRecNumClassRef(self):
        return self._itemFreezeFrameRecNumClassRef

    def set_itemFreezeFrameRecNumClassRef(self, itemFreezeFrameRecNumClassRef):
        self._itemFreezeFrameRecNumClassRef = itemFreezeFrameRecNumClassRef

    @property
    def get_itemMemoryDestinationRef(self):
        return self._itemMemoryDestinationRef

    def set_itemMemoryDestinationRef(self, itemMemoryDestinationRef):
        self._itemMemoryDestinationRef = itemMemoryDestinationRef


class PrototypeDemDTCs:
    def __init__(self, DemDTC, DemDTCSeverity, DemDtcValue, DemDTCAttributesRef):
        self._itemDtc = DemDTC
        self._itemDtcSeverity = DemDTCSeverity
        self._itemDtcId = hex(int(DemDtcValue)).upper()
        self._itemDtrAttrRef = DemDTCAttributesRef

    @property
    def get_itemDtc(self):
        return self._itemDtc

    def set_itemDtc(self, itemDtc):
        self._itemDtc = itemDtc

    @property
    def get_itemDtcSeverity(self):
        return self._itemDtcSeverity

    def set_itemDtcSeverity(self, itemDtcSeverity):
        self._itemDtcSeverity = itemDtcSeverity

    @property
    def get_itemDtcId(self):
        return self._itemDtcId

    def set_itemDtcId(self, itemDtcId):
        self._itemDtcId = itemDtcId

    @property
    def get_itemDtrAttrRef(self):
        return self._itemDtrAttrRef

    def set_itemDtrAttrRef(self, itemDtrAttrRef):
        self._itemDtrAttrRef = itemDtrAttrRef


class PrototypeDemDidClasss:
    def __init__(self, DemDidClass, DemDidIdentifier, DemDidDataElementClassRef):
        self._itemName = DemDidClass
        self._itemID = hex(int(DemDidIdentifier)).upper()
        self._itemRef = DemDidDataElementClassRef

    @property
    def get_itemName(self):
        return self._itemName

    def set_itemName(self, itemName):
        self._itemName = itemName

    @property
    def get_itemID(self):
        return self._itemID

    def set_itemID(self, itemID):
        self._itemID = itemID

    @property
    def get_itemRef(self):
        return self._itemRef

    def set_itemRef(self, itemRef):
        self._itemRef = itemRef


class PrototypeDemDataElementClasss:
    def __init__(self, DemDataElementClass, PortTypeName, PortType, DemDataElementDataSize, DemDataElementUsePort: str = 'N/A', DemInternalDataElement: str = 'N/A'):
        """
        Constructor.
        """
        self._itemName = DemDataElementClass
        self._itemTypeName = PortTypeName
        self._itemType = PortType
        self._itemSize = int(DemDataElementDataSize)
        self._itemPort = DemDataElementUsePort
        self._itemInternal = DemInternalDataElement

    def __repr__(self):
        return u'<{c} name={n}>'.format(c=self.__class__.__name__, n=self._itemName).encode('utf-8')

    def __iter__(self):
        return iter(self)

    @property
    def get_itemName(self):
        return self._itemName

    @property
    def set_itemName(self, itemName):
        self._itemName = itemName

    @property
    def get_itemTypeName(self):
        return self._itemTypeName

    @property
    def set_itemTypeName(self, itemTypeName):
        self._itemTypeName = itemTypeName

    @property
    def get_itemType(self):
        return self._itemType

    @property
    def set_itemType(self, itemType):
        self._itemType = itemType

    @property
    def get_itemSize(self):
        return self._itemSize

    @property
    def set_itemSize(self, itemSize):
        self._itemSize = itemSize

    @property
    def get_itemPort(self):
        return self._itemPort

    @property
    def set_itemPort(self, itemPort):
        self._itemPort = itemPort

    @property
    def get_itemInternal(self):
        return self._itemInternal

    @property
    def set_itemInternal(self, itemInternal):
        self._itemInternal = itemInternal


class EtasDem:
    _arxmlDemPath = ""
    _arxmlFltMPath = ""
    _arxmlCompositionPath = ""
    _xlsxPath = ""
    _eventId_path = ""
    _dtcId_path = ""
    _DemDataElementClasss = list()
    _DemDidClasss = list()
    _DemDTCs = list()
    _DemDTCAttributes = list()
    _DemEnableConditions = list()
    _DemEnableConditionGroups = list()
    _DemFreezeFrameRecNumClasss = list()
    _DemFreezeFrameRecordClasss = list()
    _DemFreezeFrameClasss = list()
    _DemExtendedDataClasss = list()
    _DemExtendedDataRecordClasss = list()
    _DemEventParameters = list()
    _DemDebounceCounterBasedClasss = list()
    _FltMPRPort = list()
    _FltMRE = list()
    _FltMEvents = list()
    _FltMREServerCallPoint = list()
    _Composition = list()

    def __init__(self, arxmlDemPath, arxmlFltMPath, arxmlCompositionPath, xlsx_path, eventId_path, dtcId_path):
        self._arxmlDemPath = arxmlDemPath
        self._arxmlFltMPath = arxmlFltMPath
        self._arxmlCompositionPath = arxmlCompositionPath
        self._xlsxPath = xlsx_path
        self._eventId_path = eventId_path
        self._dtcId_path = dtcId_path
        self.__pd_setting()

    @staticmethod
    def __pd_setting(full_display: bool = True):
        if full_display:
            pd.set_option('display.width', None)
            pd.set_option('display.max_rows', None)
            pd.set_option('display.max_columns', None)
            pd.set_option('display.max_colwidth', None)
            pd.set_option('display.max_info_rows', None)

    @staticmethod
    def cal_uuid(val_str: str = 'wanqiang.liu'):
        import uuid
        namespace = uuid.NAMESPACE_URL
        rtn_uuid = uuid.uuid5(namespace, val_str)
        return rtn_uuid

    def check_xlsx_is_open(self):
        RTN_OK = True
        RTN_NOK = False
        tempXlsxPath = str(self._xlsxPath)
        tempXlsxPath = tempXlsxPath.replace(tempXlsxPath.split("/")[-1], f"~${tempXlsxPath.split('/')[-1]}")
        if os.path.exists(tempXlsxPath):
            print(f"{self._xlsxPath} is opened,please close")
            return RTN_NOK
        else:
            return RTN_OK

    @staticmethod
    def fixed_writexml(self, writer, indent="", add_indent="", new_line=""):
        writer.write(indent + "<" + self.tagName)
        attrs = self._get_attributes()
        a_names = attrs.keys()
        # a_names.sort()
        for a_name in a_names:
            writer.write(" %s=\"" % a_name)
            minidom._write_data(writer, attrs[a_name].value)
            writer.write("\"")
        if self.childNodes:
            if len(self.childNodes) == 1 \
                    and self.childNodes[0].nodeType == minidom.Node.TEXT_NODE:
                writer.write(">")
                self.childNodes[0].writexml(writer, "", "", "")
                writer.write("</%s>%s" % (self.tagName, new_line))
                return
            writer.write(">%s" % new_line)
            for node in self.childNodes:
                if node.nodeType is not minidom.Node.TEXT_NODE:
                    node.writexml(writer, indent + add_indent, add_indent, new_line)
            writer.write("%s</%s>%s" % (indent, self.tagName, new_line))
        else:
            writer.write("/>%s" % new_line)

    def parser_Composition(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['Component Prototype'])
            g_uad_2 = str(row['Connector Name'])
            g_uad_3 = str(row['PortPath'])
            g_uad_4 = str(row['Port'])
            g_uad_5 = str(row['Connected Composition'])
            g_uad_6 = str(row['Connected PortPath'])
            g_uad_7 = str(row['Connected Port'])

        add_item_exist = False
        target = root.getElementsByTagName("CONNECTORS")[0]
        for node_1 in target.getElementsByTagName("ASSEMBLY-SW-CONNECTOR"):
            # SHORT-NAME
            node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
            node_1_2 = node_1.getElementsByTagName('PROVIDER-IREF')[0]
            node_1_3 = node_1.getElementsByTagName('REQUESTER-IREF')[0]
            node_1_2_1 = node_1_2.getElementsByTagName('CONTEXT-COMPONENT-REF')[0]
            node_1_2_2 = node_1_2.getElementsByTagName('TARGET-P-PORT-REF')[0]
            node_1_3_1 = node_1_3.getElementsByTagName('CONTEXT-COMPONENT-REF')[0]
            node_1_3_2 = node_1_3.getElementsByTagName('TARGET-R-PORT-REF')[0]

            node_1_1_text = str(node_1_1.firstChild.data)
            node_1_2_1_text = str(node_1_2_1.firstChild.data)
            node_1_2_2_text = str(node_1_2_2.firstChild.data)
            node_1_3_1_text = str(node_1_3_1.firstChild.data)
            node_1_3_2_text = str(node_1_3_2.firstChild.data)
            node_1_2_1_attr = node_1_2_1.getAttribute('DEST')  # SW-COMPONENT-PROTOTYPE
            node_1_2_2_attr = node_1_2_2.getAttribute('DEST')  # P-PORT-PROTOTYPE
            node_1_3_1_attr = node_1_2_1.getAttribute('DEST')  # SW-COMPONENT-PROTOTYPE
            node_1_3_2_attr = node_1_2_2.getAttribute('DEST')  # R-PORT-PROTOTYPE

            if method == "read":
                l_r_1 = node_1_2_1_text  # Component Prototype
                l_r_2 = node_1_1_text  # Connector Name
                l_r_3 = node_1_2_2_text  # PortPath
                l_r_4 = node_1_2_2_text  # Port
                l_r_5 = node_1_3_1_text  # Connected Composition
                l_r_6 = node_1_3_2_text  # Connected PortPath
                l_r_7 = node_1_3_2_text  # Connected Port
                atom = PrototypeComposition(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5, l_r_6, l_r_7)
                self._Composition += [atom]

            if method == "update" and node_1_1_text == g_uad_2:
                node_1_1.firstChild.data = f"ASC_{g_uad_1}_{g_uad_4}_{g_uad_5}_{g_uad_7}"
                node_1_2_1.firstChild.data = f"/Composition/Composition/{g_uad_1}"
                node_1_2_2.firstChild.data = f"{g_uad_3}{g_uad_4}"
                node_1_3_1.firstChild.data = f"/Composition/Composition/{g_uad_5}"
                node_1_3_2.firstChild.data = f"{g_uad_6}{g_uad_7}"

            if method == "remove" and node_1_1_text == g_uad_2:
                target.removeChild(node_1)

            if method == 'add' and node_1_1_text == g_uad_2:
                add_item_exist = True

        if method == 'add' and add_item_exist is False:
            node_1 = dom.createElement('ASSEMBLY-SW-CONNECTOR')
            node_1_1 = dom.createElement('SHORT-NAME')
            node_1_2 = dom.createElement('PROVIDER-IREF')
            node_1_3 = dom.createElement('REQUESTER-IREF')
            node_1_2_1 = dom.createElement('CONTEXT-COMPONENT-REF')
            node_1_2_2 = dom.createElement('TARGET-P-PORT-REF')
            node_1_3_1 = dom.createElement('CONTEXT-COMPONENT-REF')
            node_1_3_2 = dom.createElement('TARGET-R-PORT-REF')
            node_1_2_1.setAttribute('DEST', "SW-COMPONENT-PROTOTYPE")
            node_1_2_2.setAttribute('DEST', "P-PORT-PROTOTYPE")
            node_1_3_1.setAttribute('DEST', "SW-COMPONENT-PROTOTYPE")
            node_1_3_2.setAttribute('DEST', "R-PORT-PROTOTYPE")
            node_1_1.appendChild(dom.createTextNode(f"ASC_{g_uad_1}_{g_uad_4}_{g_uad_5}_{g_uad_7}"))
            node_1_2_1.appendChild(dom.createTextNode(f'/Composition/Composition/{g_uad_1}'))
            node_1_2_2.appendChild(dom.createTextNode(f'{g_uad_3}{g_uad_4}'))
            node_1_3_1.appendChild(dom.createTextNode(f'/Composition/Composition/{g_uad_5}'))
            node_1_3_2.appendChild(dom.createTextNode(f'{g_uad_6}{g_uad_7}'))
            tmp_node = [
                (node_1_3, [node_1_3_1, node_1_3_2]),
                (node_1_2, [node_1_2_1, node_1_2_2]),
                (node_1, [node_1_1, node_1_2, node_1_3]),
                (target, [node_1]),
            ]
            for item in tmp_node:
                for each in item[1]:
                    item[0].appendChild(each)
        return self._Composition

    def parser_FltMREServerCallPoint(self, root, dom, row: object = None, method: str = 'read', RE_SN: str = "RE_FltM_Main_10ms", auto_reindex: bool = False):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['Server Call Point']).split('/')
            g_uad_2 = str(row['Short Name']).split('/')
            g_uad_3 = str(row['OperationPath']).split('/')
            g_uad_4 = str(row['Operation']).split('/')
            g_uad_5 = str(row['Client Port']).split('/')
            try:
                for item in [g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'FltM_SWC':
                target = node.parentNode.getElementsByTagName("RUNNABLES")[0]
                for node_1 in target.getElementsByTagName("RUNNABLE-ENTITY"):
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    # 过滤要读取哪个RE的ServerCallPoint
                    if node_1_1_text == RE_SN:
                        target_node_1_2 = node_1.getElementsByTagName('SERVER-CALL-POINTS')[0]
                        cp_reindex = 0
                        for node_1_2_x in target_node_1_2.getElementsByTagName("SYNCHRONOUS-SERVER-CALL-POINT"):
                            node_1_2_x_1 = node_1_2_x.getElementsByTagName('SHORT-NAME')[0]
                            node_1_2_x_2 = node_1_2_x.getElementsByTagName('OPERATION-IREF')[0]
                            node_1_2_x_2_1 = node_1_2_x.getElementsByTagName('CONTEXT-R-PORT-REF')[0]
                            node_1_2_x_2_2 = node_1_2_x.getElementsByTagName('TARGET-REQUIRED-OPERATION-REF')[0]
                            node_1_2_x_1_text = str(node_1_2_x_1.firstChild.data)
                            node_1_2_x_2_1_text = str(node_1_2_x_2_1.firstChild.data)  # R-PORT
                            node_1_2_x_2_2_text = str(node_1_2_x_2_2.firstChild.data)  # OPERATION

                            if auto_reindex is True:
                                node_1_2_x_1.firstChild.data = "SynchronousServerCallPoint_" + str(cp_reindex)
                                cp_reindex = cp_reindex + 1

                            if method == 'read':
                                atom = PrototypeFltMREServerCallPoint("SynchronousServerCallPoint", node_1_2_x_1_text, node_1_2_x_2_2_text, node_1_2_x_2_2_text, node_1_2_x_2_1_text)
                                self._FltMREServerCallPoint += [atom]

                            if method == 'update' and node_1_2_x_1_text == g_uad_2[0]:
                                node_1_2_x_1.firstChild.data = g_uad_2[1]
                                node_1_2_x_2_1.firstChild.data = f"/Pkg_SWCs/FltM_SWC/{g_uad_5[1]}"
                                node_1_2_x_2_2.firstChild.data = f"/AUTOSAR_Dem/PortInterfaces/{g_uad_3[1]}/{g_uad_4[1]}"
                            if method == 'remove' and node_1_2_x_1_text == g_uad_2[0]:
                                target_node_1_2.removeChild(node_1_2_x)

                            if method == 'add' and node_1_2_x_1_text == g_uad_2[0]:
                                add_item_exist = True

                        if method == 'add' and add_item_exist is False:
                            tempUUID = self.cal_uuid(g_uad_2[1])
                            node_1 = dom.createElement("SYNCHRONOUS-SERVER-CALL-POINT")
                            node_1.setAttribute('UUID', f"{tempUUID}")
                            node_1_1 = dom.createElement('SHORT-NAME')
                            node_1_1.appendChild(dom.createTextNode(g_uad_2[1]))
                            node_1_2 = dom.createElement("OPERATION-IREF")
                            node_1_2_1 = dom.createElement("CONTEXT-R-PORT-REF")
                            node_1_2_1.setAttribute('DEST', f"R-PORT-PROTOTYPE")
                            node_1_2_1.appendChild(dom.createTextNode(f"/Pkg_SWCs/FltM_SWC/{g_uad_5[1]}"))

                            node_1_2_2 = dom.createElement("TARGET-REQUIRED-OPERATION-REF")
                            node_1_2_2.setAttribute('DEST', f"CLIENT-SERVER-OPERATION")
                            node_1_2_2.appendChild(dom.createTextNode(f"/AUTOSAR_Dem/PortInterfaces/{g_uad_3[1]}/{g_uad_4[1]}"))
                            tmp_node = [
                                (node_1_2, [node_1_2_1, node_1_2_2]),
                                (node_1, [node_1_1, node_1_2]),
                                (target_node_1_2, [node_1]),
                            ]
                            for item in tmp_node:
                                for each in item[1]:
                                    item[0].appendChild(each)

        return self._FltMREServerCallPoint

    def parser_FltMEvents(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['Rte Event Type']).split('/')
            g_uad_2 = str(row['Event Name']).split('/')
            g_uad_3 = str(row['Start Runnable Entity']).split('/')
            g_uad_4 = str(row['Timing Period(s)']).split('/')
            g_uad_5 = str(row['Target']).split('/')
            g_uad_6 = str(row['PortPath']).split('/')
            g_uad_7 = str(row['Port']).split('/')
            try:
                for item in [g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'FltM_SWC':
                target = node.parentNode.getElementsByTagName("EVENTS")[0]
                for node_1 in (target.getElementsByTagName("TIMING-EVENT") +
                               target.getElementsByTagName("OPERATION-INVOKED-EVENT") +
                               target.getElementsByTagName("INIT-EVENT")):
                    l_r_1 = ''
                    node_1_1 = object
                    node_1_2 = object
                    node_1_3 = object
                    node_1_3_1 = object
                    node_1_3_2 = object
                    node_1_1_text = ''
                    node_1_2_text = ''
                    node_1_3_text = ''
                    node_1_3_1_text = ""
                    node_1_3_2_text = "/"
                    # SHORT-NAME
                    if node_1.nodeName == "TIMING-EVENT":
                        l_r_1 = "TimingEvent"
                        node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                        node_1_2 = node_1.getElementsByTagName('START-ON-EVENT-REF')[0]
                        node_1_3 = node_1.getElementsByTagName("PERIOD")[0]
                        node_1_1_text = str(node_1_1.firstChild.data)
                        node_1_2_attr = node_1_2.getAttribute('DEST')
                        node_1_2_text = str(node_1_2.firstChild.data)
                        node_1_3_text = str(node_1_3.firstChild.data)

                    elif node_1.nodeName == "OPERATION-INVOKED-EVENT":
                        l_r_1 = "OperationInvokedEvent"
                        node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                        node_1_2 = node_1.getElementsByTagName('START-ON-EVENT-REF')[0]
                        node_1_3 = node_1.getElementsByTagName("OPERATION-IREF")[0]
                        node_1_3_1 = node_1_3.getElementsByTagName("CONTEXT-P-PORT-REF")[0]
                        node_1_3_2 = node_1_3.getElementsByTagName("TARGET-PROVIDED-OPERATION-REF")[0]
                        node_1_1_text = str(node_1_1.firstChild.data)
                        node_1_2_attr = node_1_2.getAttribute('DEST')
                        node_1_2_text = str(node_1_2.firstChild.data)
                        node_1_3_1_text = str(node_1_3_1.firstChild.data)
                        node_1_3_2_text = str(node_1_3_2.firstChild.data)

                    elif node_1.nodeName == "INIT-EVENT":
                        l_r_1 = "InitEvent"
                        node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                        node_1_2 = node_1.getElementsByTagName('START-ON-EVENT-REF')[0]
                        node_1_1_text = str(node_1_1.firstChild.data)
                        node_1_2_text = str(node_1_2.firstChild.data)
                    if method == 'read':
                        l_r_2 = node_1_1_text
                        l_r_3 = node_1_2_text
                        l_r_4 = node_1_3_text
                        l_r_5 = node_1_3_1_text
                        l_r_6 = node_1_3_2_text
                        l_r_7 = node_1_3_2_text
                        atom = PrototypeFltMEvents(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5, l_r_6, l_r_7)
                        self._FltMEvents += [atom]

                    if method == 'update' and node_1_1_text == g_uad_2[0]:
                        try:
                            if g_uad_1[1] == "TimingEvent":
                                node_1_1.firstChild.data = g_uad_2[1]
                                node_1_2.firstChild.data = f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"
                                node_1_3.firstChild.data = g_uad_4[1]
                            if g_uad_1[1] == "OperationInvokedEvent":
                                node_1_1.firstChild.data = g_uad_2[1]
                                node_1_2.firstChild.data = f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"
                                node_1_3_1.firstChild.data = f"/Pkg_SWCs/FltM_SWC/{g_uad_5[1]}"
                                node_1_3_2.firstChild.data = f"/AUTOSAR_Dem/PortInterfaces/{g_uad_6[1]}/{g_uad_7[1]}"
                            if g_uad_1[1] == "InitEvent":
                                node_1_1.firstChild.data = g_uad_2[1]
                                node_1_2.firstChild.data = f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"
                        except Exception as e:
                            pass

                    if method == 'remove' and node_1_1_text == g_uad_2[0]:
                        target.removeChild(node_1)

                    if method == 'add' and node_1_1_text == g_uad_2[0]:
                        add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    tempUUID = self.cal_uuid(g_uad_2[1])
                    tmp_node = []
                    if g_uad_1[1] == "TimingEvent":
                        node_1 = dom.createElement("TIMING-EVENT")
                        node_1.setAttribute('UUID', f"{tempUUID}")
                        node_1_1 = dom.createElement('SHORT-NAME')
                        node_1_1.appendChild(dom.createTextNode(g_uad_2[1]))
                        node_1_2 = dom.createElement("START-ON-EVENT-REF")
                        node_1_2.setAttribute('DEST', "RUNNABLE-ENTITY")
                        node_1_2.appendChild(dom.createTextNode(f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"))
                        node_1_3 = dom.createElement("PERIOD")
                        node_1_3.appendChild(dom.createTextNode(g_uad_4[1]))
                        tmp_node = [
                            (node_1, [node_1_1, node_1_2, node_1_3]),
                            (target, [node_1]),
                        ]

                    if g_uad_1[1] == "OperationInvokedEvent":
                        node_1 = dom.createElement("OPERATION-INVOKED-EVENT")
                        node_1.setAttribute('UUID', f"{tempUUID}")
                        node_1_1 = dom.createElement('SHORT-NAME')
                        node_1_1.appendChild(dom.createTextNode(g_uad_2[1]))
                        node_1_2 = dom.createElement("START-ON-EVENT-REF")
                        node_1_2.setAttribute('DEST', "RUNNABLE-ENTITY")
                        node_1_2.appendChild(dom.createTextNode(f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"))
                        node_1_3 = dom.createElement("OPERATION-IREF")
                        node_1_3_1 = dom.createElement('CONTEXT-P-PORT-REF')
                        node_1_3_1.setAttribute('DEST', "P-PORT-PROTOTYPE")
                        node_1_3_1.appendChild(dom.createTextNode(f"/Pkg_SWCs/FltM_SWC/{g_uad_5[1]}"))
                        node_1_3_2 = dom.createElement('TARGET-PROVIDED-OPERATION-REF')
                        node_1_3_2.setAttribute('DEST', "CLIENT-SERVER-OPERATION")
                        node_1_3_2.appendChild(dom.createTextNode(f"/AUTOSAR_Dem/PortInterfaces/{g_uad_6[1]}/{g_uad_7[1]}"))
                        tmp_node = [
                            (node_1_3, [node_1_3_1, node_1_3_2]),
                            (node_1, [node_1_1, node_1_2, node_1_3]),
                            (target, [node_1]),
                        ]

                    if g_uad_1[1] == "InitEvent":
                        node_1 = dom.createElement("INIT-EVENT")
                        node_1.setAttribute('UUID', f"{tempUUID}")
                        node_1_1 = dom.createElement('SHORT-NAME')
                        node_1_1.appendChild(dom.createTextNode(g_uad_2[1]))
                        node_1_2 = dom.createElement("START-ON-EVENT-REF")
                        node_1_2.setAttribute('DEST', "RUNNABLE-ENTITY")
                        node_1_2.appendChild(dom.createTextNode(f"/Pkg_SWCs/FltM_SWC/IB_FltM_SWC/{g_uad_3[1]}"))
                        tmp_node = [
                            (node_1, [node_1_1, node_1_2]),
                            (target, [node_1]),
                        ]

                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._FltMEvents

    def parser_FltMRE(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['Runnable Entity']).split('/')
            g_uad_2 = str(row['Function Name']).split('/')
            try:
                for item in [g_uad_1, g_uad_2]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'FltM_SWC':
                target = node.parentNode.getElementsByTagName("RUNNABLES")[0]
                for node_1 in target.getElementsByTagName("RUNNABLE-ENTITY"):
                    node_1_1 = object
                    node_1_2 = object
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('SYMBOL')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_text = str(node_1_2.firstChild.data)

                    if method == 'read':
                        l_r_1 = node_1_1_text
                        l_r_2 = node_1_2_text
                        atom = PrototypeFltMRE(l_r_1, l_r_2)
                        self._FltMRE += [atom]

                    if method == 'update' and node_1_1_text == g_uad_1[0]:
                        node_1_1.firstChild.data = g_uad_1[1]
                        node_1_2.firstChild.data = g_uad_2[1]

                    if method == 'remove' and node_1_1_text == g_uad_1[0]:
                        target.removeChild(node_1)

                    if method == 'add' and node_1_1_text == g_uad_1[0]:
                        add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    tempUUID = self.cal_uuid(g_uad_1[1])
                    node_1 = dom.createElement("RUNNABLE-ENTITY")
                    node_1.setAttribute('UUID', f"{tempUUID}")
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(g_uad_1[1]))
                    node_1_2 = dom.createElement("SYMBOL")
                    node_1_2.appendChild(dom.createTextNode(g_uad_2[1]))

                    tmp_node = [
                        (node_1, [node_1_1, node_1_2]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._FltMRE

    def parser_FltMPRPort(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['Type']).split('/')
            g_uad_2 = str(row['Short Name']).split('/')
            g_uad_3 = str(row['Interface']).split('/')
            g_uad_4 = str(row['Port InterfacePath'])
            g_uad_5 = str(row['Port Interface']).split('/')
            g_uad_6 = str(row['Is Service']).split('/')
            try:
                for item in [g_uad_1, g_uad_2, g_uad_3, g_uad_5, g_uad_6]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'FltM_SWC':
                target = node.parentNode.getElementsByTagName("PORTS")[0]
                for node_1 in (target.getElementsByTagName("R-PORT-PROTOTYPE") + target.getElementsByTagName("P-PORT-PROTOTYPE")):
                    node_1_1 = object
                    node_1_2 = object
                    node_1_1_text = ''
                    node_1_2_attr = ''
                    node_1_2_text = ''
                    l_r_1 = ""
                    l_r_6 = ""
                    # SHORT-NAME
                    if node_1.nodeName == "R-PORT-PROTOTYPE":
                        node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                        node_1_2 = node_1.getElementsByTagName('REQUIRED-INTERFACE-TREF')[0]
                        node_1_1_text = str(node_1_1.firstChild.data)
                        node_1_2_attr = node_1_2.getAttribute('DEST')
                        node_1_2_text = str(node_1_2.firstChild.data)
                        l_r_1 = "RPort"
                        if node_1_2_attr == "SENDER-RECEIVER-INTERFACE":
                            l_r_6 = "false"
                        else:
                            l_r_6 = "true"

                    elif node_1.nodeName == "P-PORT-PROTOTYPE":
                        node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                        node_1_2 = node_1.getElementsByTagName('PROVIDED-INTERFACE-TREF')[0]
                        node_1_1_text = str(node_1_1.firstChild.data)
                        node_1_2_attr = node_1_2.getAttribute('DEST')
                        node_1_2_text = str(node_1_2.firstChild.data)
                        l_r_1 = "PPort"
                        if node_1_2_attr == "SENDER-RECEIVER-INTERFACE":
                            l_r_6 = "false"
                        else:
                            l_r_6 = "true"
                    else:
                        pass

                    if method == 'read':
                        l_r_2 = node_1_1_text
                        l_r_3 = node_1_2_attr
                        l_r_4 = node_1_2_text[:node_1_2_text.rindex("/") + 1]
                        l_r_5 = node_1_2_text.split('/')[-1]
                        atom = PrototypeFltMPRPort(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5, l_r_6)
                        self._FltMPRPort += [atom]
                    if method == 'update' and node_1_1_text == g_uad_2[0]:
                        # 当前实现,仅能更改ShortName及PortInterface,其余不可更改
                        node_1_1.firstChild.data = g_uad_2[1]
                        node_1_2.firstChild.data = g_uad_4 + g_uad_5[1]

                    if method == 'remove' and node_1_1_text == g_uad_2[0]:
                        target.removeChild(node_1)

                    if method == 'add' and node_1_1_text == g_uad_2[0]:
                        add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    if g_uad_1[1] == "RPort":
                        element_r1 = 'R-PORT-PROTOTYPE'
                        element_r2 = 'REQUIRED-INTERFACE-TREF'
                    elif g_uad_1[1] == "PPort":
                        element_r1 = 'P-PORT-PROTOTYPE'
                        element_r2 = 'PROVIDED-INTERFACE-TREF'
                    else:
                        element_r1 = ''
                        element_r2 = ''
                        pass
                    tempUUID = self.cal_uuid(g_uad_2[1])
                    node_1 = dom.createElement(element_r1)
                    node_1.setAttribute('UUID', f"{tempUUID}")
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(g_uad_2[1]))
                    node_1_2 = dom.createElement(element_r2)
                    node_1_2.setAttribute('DEST', g_uad_3[1])
                    node_1_2.appendChild(dom.createTextNode(g_uad_4 + g_uad_5[1]))

                    tmp_node = [
                        (node_1, [node_1_1, node_1_2]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._FltMPRPort

    def parser_DemDebounceCounterBasedClasss(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7, g_uad_8, g_uad_9, g_uad_10, g_uad_11, g_uad_12, g_uad_13, g_uad_14
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['DemDebounceCounterBasedClass']).split('/')
            g_uad_2 = str(row['DemDebounceBehavior*']).split('/')
            g_uad_3 = str(row['DemDebounceCounterDecrementStepSize*']).split('/')
            g_uad_4 = str(row['DemDebounceCounterFailedThreshold']).split('/')
            g_uad_5 = str(row['DemDebounceCounterIncrementStepSize*']).split('/')
            g_uad_6 = str(row['DemDebounceCounterJumpDown']).split('/')
            g_uad_7 = str(row['DemDebounceCounterJumpDownValue']).split('/')
            g_uad_8 = str(row['DemDebounceCounterJumpUp']).split('/')
            g_uad_9 = str(row['DemDebounceCounterJumpUpValue']).split('/')
            g_uad_10 = str(row['DemDebounceCounterPassedThreshold']).split('/')
            g_uad_11 = str(row['DemDebounceCounterStorage*']).split('/')
            g_uad_12 = str(row['DemRbDebounceCounterFdcThresholdStorageValue']).split('/')
            g_uad_13 = str(row['DemRbDebounceCounterJumpDownAlternative']).split('/')
            g_uad_14 = str(row['DemRbDebounceCounterJumpUpAlternative']).split('/')
            try:
                for item in [g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7, g_uad_8, g_uad_9, g_uad_10, g_uad_11, g_uad_12, g_uad_13, g_uad_14]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemConfigSet':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    node_1_3_1_2_text = ''
                    node_1_3_2_2_text = ''
                    node_1_3_3_2_text = ''
                    node_1_3_4_2_text = ''
                    node_1_3_5_2_text = ''
                    node_1_3_6_2_text = ''
                    node_1_3_7_2_text = ''
                    node_1_3_8_2_text = ''
                    node_1_3_9_2_text = ''
                    node_1_3_10_2_text = ''
                    node_1_3_11_2_text = ''
                    node_1_3_1_2 = object
                    node_1_3_2_2 = object
                    node_1_3_3_2 = object
                    node_1_3_4_2 = object
                    node_1_3_5_2 = object
                    node_1_3_6_2 = object
                    node_1_3_7_2 = object
                    node_1_3_8_2 = object
                    node_1_3_9_2 = object
                    node_1_3_10_2 = object
                    node_1_3_11_2 = object
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemDebounceCounterBasedClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF' and \
                            node_1_2_text.find('/DemDebounceCounterBasedClass/') == -1 and node_1_2_text.find('/DemDebounceCounterBasedClassRef') == -1:

                        node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE')[0]
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceBehavior') != -1:
                                node_1_3_1_2 = node_1_3_x_2
                                node_1_3_1_2_text = str(node_1_3_x_2.firstChild.data)

                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE')[0]
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterDecrementStepSize') != -1:
                                node_1_3_2_2 = node_1_3_x_2
                                node_1_3_2_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterFailedThreshold') != -1:
                                node_1_3_3_2 = node_1_3_x_2
                                node_1_3_3_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterIncrementStepSize') != -1:
                                node_1_3_4_2 = node_1_3_x_2
                                node_1_3_4_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterJumpDown') != -1 and node_1_3_x_1_text.find('/DemDebounceCounterJumpDownValue') == -1:  # TRUE or FALSE
                                node_1_3_5_2 = node_1_3_x_2
                                node_1_3_5_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterJumpDownValue') != -1:
                                node_1_3_6_2 = node_1_3_x_2
                                node_1_3_6_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterJumpUp') != -1 and node_1_3_x_1_text.find('/DemDebounceCounterJumpUpValue') == -1:  # TRUE or FALSE
                                node_1_3_7_2 = node_1_3_x_2
                                node_1_3_7_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterJumpUpValue') != -1:
                                node_1_3_8_2 = node_1_3_x_2
                                node_1_3_8_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterPassedThreshold') != -1:
                                node_1_3_9_2 = node_1_3_x_2
                                node_1_3_9_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDebounceCounterStorage') != -1:
                                node_1_3_10_2 = node_1_3_x_2
                                node_1_3_10_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemRbDebounceCounterFdcThresholdStorageValue') != -1:
                                node_1_3_11_2 = node_1_3_x_2
                                node_1_3_11_2_text = str(node_1_3_x_2.firstChild.data)

                        if method == 'read':
                            l_r_1 = node_1_1_text
                            l_r_2 = node_1_3_1_2_text
                            l_r_3 = node_1_3_2_2_text
                            l_r_4 = node_1_3_3_2_text
                            l_r_5 = node_1_3_4_2_text
                            l_r_6 = node_1_3_5_2_text
                            l_r_7 = node_1_3_6_2_text
                            l_r_8 = node_1_3_7_2_text
                            l_r_9 = node_1_3_8_2_text
                            l_r_10 = node_1_3_9_2_text
                            l_r_11 = node_1_3_10_2_text
                            l_r_12 = node_1_3_11_2_text
                            atom = PrototypeDemDebounceCounterBasedClasss(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5, l_r_6, l_r_7, l_r_8, l_r_9, l_r_10, l_r_11, l_r_12, 'false', 'false')
                            self._DemDebounceCounterBasedClasss += [atom]

                        if method == 'update' and node_1_1_text == g_uad_1[0]:
                            node_1_1.firstChild.data = g_uad_1[1]
                            node_1_3_1_2.firstChild.data = str(g_uad_2[1]).upper()
                            node_1_3_2_2.firstChild.data = str(int(float(g_uad_3[1])))
                            node_1_3_3_2.firstChild.data = str(int(float(g_uad_4[1])))
                            node_1_3_4_2.firstChild.data = str(int(float(g_uad_5[1])))
                            node_1_3_5_2.firstChild.data = str(g_uad_6[1]).lower()
                            node_1_3_6_2.firstChild.data = str(int(float(g_uad_7[1])))
                            node_1_3_7_2.firstChild.data = str(g_uad_8[1]).lower()
                            node_1_3_8_2.firstChild.data = str(int(float(g_uad_9[1])))
                            node_1_3_9_2.firstChild.data = str(int(float(g_uad_10[1])))
                            node_1_3_10_2.firstChild.data = str(g_uad_11[1]).lower()
                            node_1_3_11_2.firstChild.data = str(int(float(g_uad_12[1])))

                        if method == 'remove' and node_1_1_text == g_uad_1[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == g_uad_1[0]:
                            add_item_exist = True
                if method == 'add' and add_item_exist is False:
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(g_uad_1[1]))
                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDebounceCounterBasedClass'))
                    node_1_3 = dom.createElement('PARAMETER-VALUES')
                    cycItem = [
                        ('ECUC-TEXTUAL-PARAM-VALUE', 'ECUC-ENUMERATION-PARAM-DEF', 'DemDebounceBehavior', str(g_uad_2[1]).upper()),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterDecrementStepSize', str(int(float(g_uad_3[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterFailedThreshold', str(int(float(g_uad_4[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterIncrementStepSize', str(int(float(g_uad_5[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-BOOLEAN-PARAM-DEF', 'DemDebounceCounterJumpDown', str(g_uad_6[1]).lower()),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterJumpDownValue', str(int(float(g_uad_7[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-BOOLEAN-PARAM-DEF', 'DemDebounceCounterJumpUp', str(g_uad_8[1]).lower()),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterJumpUpValue', str(int(float(g_uad_9[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemDebounceCounterPassedThreshold', str(int(float(g_uad_10[1])))),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-BOOLEAN-PARAM-DEF', 'DemDebounceCounterStorage', str(g_uad_11[1]).lower()),
                        ('ECUC-NUMERICAL-PARAM-VALUE', 'ECUC-INTEGER-PARAM-DEF', 'DemRbDebounceCounterFdcThresholdStorageValue', str(int(float(g_uad_12[1])))),
                    ]
                    for item in cycItem:
                        node_1_3_x = dom.createElement(f'{item[0]}')
                        node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                        node_1_3_x_1.setAttribute('DEST', f"{item[1]}")
                        node_1_3_x_1.appendChild(dom.createTextNode(f'/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDebounceCounterBasedClass/{item[2]}'))
                        node_1_3_x_2 = dom.createElement('VALUE')
                        node_1_3_x_2.appendChild(dom.createTextNode(f'{item[3]}'))
                        node_1_3_x.appendChild(node_1_3_x_1)
                        node_1_3_x.appendChild(node_1_3_x_2)
                        node_1_3.appendChild(node_1_3_x)

                    tmp_node = [
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)

        return self._DemDebounceCounterBasedClasss

    def parser_DemEventParameters(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7, g_uad_8, g_uad_9
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['DemEventParameter']).split('/')
            g_uad_2 = str(row['DemEventAvailable*']).split('/')
            g_uad_3 = str(row['DemEventFailureCycleCounterThreshold*']).split('/')
            g_uad_4 = str(row['DemEventKind*']).split('/')
            g_uad_5 = str(row['DemFFPrestorageSupported*']).split('/')
            g_uad_6 = str(row['DemDTCRef']).split('/')
            g_uad_7 = str(row['DemEnableConditionGroupRef']).split('/')
            g_uad_8 = str(row['DemOperationCycleRef*']).split('/')
            g_uad_9 = str(row['DemDebounceCounterBasedClassRef']).split('/')
            try:
                for item in [g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5, g_uad_6, g_uad_7, g_uad_8, g_uad_9]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemConfigSet':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    node_1_3_1_2_text = ''
                    node_1_3_2_2_text = ''
                    node_1_3_3_2_text = ''
                    node_1_3_4_2_text = ''
                    node_1_4_1_2_text = ''
                    node_1_4_2_2_text = ''
                    node_1_4_3_2_text = ''
                    node_1_5_1_2_text = ''

                    node_1_3_1_2 = object
                    node_1_3_2_2 = object
                    node_1_3_4_2 = object

                    node_1_4_1_2 = object
                    node_1_4_2_2 = object
                    node_1_4_3_2 = object
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemEventParameter') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF' and node_1_2_text.find('/DemEventParameter/') == -1:

                        node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE')[0]
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemEventAvailable') != -1:
                                node_1_3_1_2 = node_1_3_x_2
                                node_1_3_1_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemEventFailureCycleCounterThreshold') != -1:
                                node_1_3_2_2 = node_1_3_x_2
                                node_1_3_2_2_text = str(node_1_3_x_2.firstChild.data)
                            if node_1_3_x_1_text.find('/DemFFPrestorageSupported') != -1:
                                node_1_3_4_2 = node_1_3_x_2
                                node_1_3_4_2_text = str(node_1_3_x_2.firstChild.data)
                        node_1_3_3 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                        node_1_3_3_1 = node_1_3_3.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_3_2 = node_1_3_3.getElementsByTagName('VALUE')[0]
                        node_1_3_3_1_text = str(node_1_3_3_1.firstChild.data)
                        if node_1_3_3_1_text.find('/DemEventKind') != -1:
                            node_1_3_3_2_text = str(node_1_3_3_2.firstChild.data)

                        node_1_4 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        for node_1_4_x in node_1_4.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                            node_1_4_x_1 = node_1_4_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_4_x_2 = node_1_4_x.getElementsByTagName('VALUE-REF')[0]
                            node_1_4_x_1_attr = node_1_4_x_1.getAttribute('DEST')
                            node_1_4_x_1_text = str(node_1_4_x_1.firstChild.data)
                            if node_1_4_x_1_text.find('/DemDTCRef') != -1:
                                node_1_4_1_2 = node_1_4_x_2
                                node_1_4_1_2_text = str(node_1_4_x_2.firstChild.data)
                            if node_1_4_x_1_text.find('/DemEnableConditionGroupRef') != -1:
                                node_1_4_2_2 = node_1_4_x_2
                                node_1_4_2_2_text = str(node_1_4_x_2.firstChild.data)
                            if node_1_4_x_1_text.find('/DemOperationCycleRef') != -1:
                                node_1_4_3_2 = node_1_4_x_2
                                node_1_4_3_2_text = str(node_1_4_x_2.firstChild.data)

                        # Just debounce ref
                        node_1_5 = node_1.getElementsByTagName('SUB-CONTAINERS')[0]
                        node_1_5_1 = node_1_5.getElementsByTagName('ECUC-REFERENCE-VALUE')[0]
                        node_1_5_1_1 = node_1_5_1.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_5_1_2 = node_1_5_1.getElementsByTagName('VALUE-REF')[0]
                        node_1_5_1_1_text = str(node_1_5_1_1.firstChild.data)
                        if node_1_5_1_1_text.find('/DemDebounceCounterBasedClassRef') != -1:
                            node_1_5_1_2_text = str(node_1_5_1_2.firstChild.data)

                        if method == 'read':
                            l_r_1 = node_1_1_text
                            l_r_2 = node_1_3_1_2_text
                            l_r_3 = node_1_3_2_2_text
                            l_r_4 = node_1_3_3_2_text
                            l_r_5 = node_1_3_4_2_text
                            l_r_6 = node_1_4_1_2_text
                            l_r_7 = node_1_4_2_2_text
                            l_r_8 = node_1_4_3_2_text
                            l_r_9 = node_1_5_1_2_text
                            atom = PrototypeDemEventParameters(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5, l_r_6, l_r_7, l_r_8, l_r_9)
                            self._DemEventParameters += [atom]
                        if method == 'update' and node_1_1_text == g_uad_1[0]:
                            node_1_1.firstChild.data = g_uad_1[1]
                            node_1_3_1_2.firstChild.data = str(g_uad_2[1]).lower()
                            node_1_3_2_2.firstChild.data = int(float(g_uad_3[1]))
                            node_1_3_3_2.firstChild.data = g_uad_4[1]
                            node_1_3_4_2.firstChild.data = str(g_uad_5[1]).lower()
                            if g_uad_6[1] != 'nan':
                                node_1_4_1_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{g_uad_6[1]}'
                            if g_uad_7[1] != 'nan':
                                node_1_4_2_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_7[1]}'
                            if g_uad_8[1] != 'nan':
                                node_1_4_3_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_8[1]}'
                            if g_uad_9[1] != 'nan':
                                node_1_5_1_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{g_uad_9[1]}'
                        if method == 'remove' and node_1_1_text == g_uad_1[0]:
                            target.removeChild(node_1)
                        if method == 'add' and node_1_1_text == g_uad_1[0]:
                            add_item_exist = True
                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(g_uad_1[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter'))

                    node_1_3 = dom.createElement('PARAMETER-VALUES')

                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemEventAvailable'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(str(g_uad_2[1]).lower()))

                    node_1_3_2 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_2_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemEventFailureCycleCounterThreshold'))
                    node_1_3_2_2 = dom.createElement('VALUE')
                    node_1_3_2_2.appendChild(dom.createTextNode(str(int(float(g_uad_3[1])))))

                    node_1_3_3 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_3_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemEventKind'))
                    node_1_3_3_2 = dom.createElement('VALUE')
                    node_1_3_3_2.appendChild(dom.createTextNode(g_uad_4[1]))

                    node_1_3_4 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_4_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_4_1.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    node_1_3_4_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemFFPrestorageSupported'))
                    node_1_3_4_2 = dom.createElement('VALUE')
                    node_1_3_4_2.appendChild(dom.createTextNode(str(g_uad_5[1]).lower()))

                    node_1_4 = dom.createElement('REFERENCE-VALUES')

                    node_1_4_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemDTCRef'))
                    node_1_4_1_2 = dom.createElement('VALUE-REF')
                    node_1_4_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{g_uad_6[1]}'))

                    node_1_4_2 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_2_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemEnableConditionGroupRef'))
                    node_1_4_2_2 = dom.createElement('VALUE-REF')
                    node_1_4_2_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_2_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_7[1]}'))

                    node_1_4_3 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_3_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemOperationCycleRef'))
                    node_1_4_3_2 = dom.createElement('VALUE-REF')
                    node_1_4_3_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_3_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_8[1]}'))

                    node_1_5 = dom.createElement('SUB-CONTAINERS')
                    node_1_5_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_5_1_1 = dom.createElement('SHORT-NAME')
                    node_1_5_1_1.appendChild(dom.createTextNode('DemDebounceAlgorithmClass'))
                    node_1_5_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_5_1_2.setAttribute('DEST', "ECUC-CHOICE-CONTAINER-DEF")
                    node_1_5_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemDebounceAlgorithmClass'))
                    node_1_5_1_3 = dom.createElement('SUB-CONTAINERS')
                    node_1_5_1_3_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_5_1_3_1_1 = dom.createElement('SHORT-NAME')
                    node_1_5_1_3_1_1.appendChild(dom.createTextNode('DemDebounceCounterBased'))
                    node_1_5_1_3_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_5_1_3_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_5_1_3_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemDebounceAlgorithmClass/DemDebounceCounterBased'))
                    node_1_5_1_3_1_3 = dom.createElement('REFERENCE-VALUES')
                    node_1_5_1_3_1_3_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_5_1_3_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_5_1_3_1_3_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_5_1_3_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemDebounceAlgorithmClass/DemDebounceCounterBased/DemDebounceCounterBasedClassRef'))
                    node_1_5_1_3_1_3_1_2 = dom.createElement('VALUE-REF')
                    node_1_5_1_3_1_3_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_5_1_3_1_3_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{g_uad_9[1]}'))

                    node_1_5_2 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_5_2_1 = dom.createElement('SHORT-NAME')
                    node_1_5_2_1.appendChild(dom.createTextNode('DemRbEventClass'))
                    node_1_5_2_2 = dom.createElement('DEFINITION-REF')
                    node_1_5_2_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_5_2_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemRbEventClass'))
                    node_1_5_2_3 = dom.createElement('PARAMETER-VALUES')
                    node_1_5_2_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_5_2_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_5_2_3_1_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_5_2_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemEventParameter/DemRbEventClass/DemRbEventBufferTime'))
                    node_1_5_2_3_1_2 = dom.createElement('VALUE')
                    node_1_5_2_3_1_2.appendChild(dom.createTextNode('0'))

                    tmp_node = [
                        (node_1_5_2_3_1, [node_1_5_2_3_1_1, node_1_5_2_3_1_2]),
                        (node_1_5_2_3, [node_1_5_2_3_1]),
                        (node_1_5_2, [node_1_5_2_1, node_1_5_2_2, node_1_5_2_3]),
                        (node_1_5_1_3_1_3_1, [node_1_5_1_3_1_3_1_1, node_1_5_1_3_1_3_1_2]),
                        (node_1_5_1_3_1_3, [node_1_5_1_3_1_3_1]),
                        (node_1_5_1_3_1, [node_1_5_1_3_1_1, node_1_5_1_3_1_2, node_1_5_1_3_1_3]),
                        (node_1_5_1_3, [node_1_5_1_3_1]),
                        (node_1_5_1, [node_1_5_1_1, node_1_5_1_2, node_1_5_1_3]),
                        (node_1_4_3, [node_1_4_3_1, node_1_4_3_2]),
                        (node_1_4_2, [node_1_4_2_1, node_1_4_2_2]),
                        (node_1_4_1, [node_1_4_1_1, node_1_4_1_2]),
                        (node_1_3_4, [node_1_3_4_1, node_1_3_4_2]),
                        (node_1_3_3, [node_1_3_3_1, node_1_3_3_2]),
                        (node_1_3_2, [node_1_3_2_1, node_1_3_2_2]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_5, [node_1_5_1, node_1_5_2]),
                        (node_1_4, [node_1_4_1, node_1_4_2, node_1_4_3]),
                        (node_1_3, [node_1_3_1, node_1_3_2, node_1_3_3, node_1_3_4]),
                        (node_1, [node_1_1, node_1_2, node_1_3, node_1_4, node_1_5]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemEventParameters

    def parser_DemExtendedDataRecordClasss(self, root, dom, row: object = None, method: str = 'read'):
        global g_uad_1, g_uad_2, g_uad_3, g_uad_4, g_uad_5
        if method == 'update' or method == 'remove' or method == 'add':
            g_uad_1 = str(row['DemExtendedDataRecordClass']).split('/')
            g_uad_2 = str(row['DemExtendedDataRecordNumber*']).split('/')
            g_uad_3 = str(row['DemExtendedDataRecordTrigger*']).split('/')
            g_uad_4 = str(row['DemExtendedDataRecordUpdate*']).split('/')
            g_uad_5 = str(row['DemDataElementClassRef']).split('/')
            try:
                g_uad_1.append(g_uad_1[0])
                g_uad_2.append(g_uad_2[0])
                g_uad_3.append(g_uad_3[0])
                g_uad_4.append(g_uad_4[0])
                g_uad_5.append(g_uad_5[0])
                g_uad_2[1] = str(g_uad_2[1]).replace('0x', '')
                g_uad_2[1] = str(g_uad_2[1]).replace('0X', '')
                g_uad_2[1] = str(int(g_uad_2[1], 16))
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    node_1_3_1_2_text = ''
                    node_1_3_2_2_text = ''
                    node_1_3_3_2_text = ''
                    node_1_4_1_2_text = ''
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemExtendedDataRecordClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]
                        node_1_3_1 = node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        node_1_3_1_1 = node_1_3_1.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_1_2 = node_1_3_1.getElementsByTagName('VALUE')[0]
                        node_1_3_1_1_text = str(node_1_3_1_1.firstChild.data)
                        node_1_3_1_1_attr = node_1_3_1_1.getAttribute('DEST')
                        if node_1_3_1_1_text.find('/DemExtendedDataRecordNumber') != -1:
                            node_1_3_1_2_text = str(node_1_3_1_2.firstChild.data)

                        node_1_3_2 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                        node_1_3_2_1 = node_1_3_2.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_2_2 = node_1_3_2.getElementsByTagName('VALUE')[0]
                        node_1_3_2_1_text = str(node_1_3_2_1.firstChild.data)
                        node_1_3_2_1_attr = node_1_3_2_1.getAttribute('DEST')
                        if node_1_3_2_1_text.find('/DemExtendedDataRecordTrigger') != -1:
                            node_1_3_2_2_text = str(node_1_3_2_2.firstChild.data)

                        node_1_3_3 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[1]
                        node_1_3_3_1 = node_1_3_3.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_3_2 = node_1_3_3.getElementsByTagName('VALUE')[0]
                        node_1_3_3_1_text = str(node_1_3_3_1.firstChild.data)
                        node_1_3_3_1_attr = node_1_3_3_1.getAttribute('DEST')
                        if node_1_3_3_1_text.find('/DemExtendedDataRecordUpdate') != -1:
                            node_1_3_3_2_text = str(node_1_3_3_2.firstChild.data)

                        node_1_4 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        node_1_4_1 = node_1_4.getElementsByTagName('ECUC-REFERENCE-VALUE')[0]
                        node_1_4_1_1 = node_1_4_1.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_4_1_2 = node_1_4_1.getElementsByTagName('VALUE-REF')[0]
                        node_1_4_1_1_text = str(node_1_4_1_1.firstChild.data)
                        node_1_4_1_1_attr = node_1_4_1_1.getAttribute('DEST')
                        if node_1_4_1_1_text.find('/DemDataElementClassRef') != -1:
                            node_1_4_1_2_text = str(node_1_4_1_2.firstChild.data)

                        if method == 'read':
                            l_r_1 = node_1_1_text
                            l_r_2 = node_1_3_1_2_text
                            l_r_3 = node_1_3_2_2_text
                            l_r_4 = node_1_3_3_2_text
                            l_r_5 = node_1_4_1_2_text
                            atom = PrototypeDemExtendedDataRecordClasss(l_r_1, l_r_2, l_r_3, l_r_4, l_r_5)
                            self._DemExtendedDataRecordClasss += [atom]
                        if method == 'update' and node_1_1_text == g_uad_1[0]:
                            node_1_1.firstChild.data = g_uad_1[1]
                            node_1_3_1_2.firstChild.data = g_uad_2[1]
                            node_1_3_2_2.firstChild.data = g_uad_3[1]
                            node_1_3_3_2.firstChild.data = g_uad_4[1]
                            node_1_4_1_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_5[1]}'
                        if method == 'remove' and node_1_1_text == g_uad_1[0]:
                            target.removeChild(node_1)
                        if method == 'add' and node_1_1_text == g_uad_1[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(g_uad_1[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataRecordClass'))

                    node_1_3 = dom.createElement('PARAMETER-VALUES')

                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataRecordClass/DemExtendedDataRecordNumber'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(g_uad_2[1]))

                    node_1_3_2 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_2_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataRecordClass/DemExtendedDataRecordTrigger'))
                    node_1_3_2_2 = dom.createElement('VALUE')
                    node_1_3_2_2.appendChild(dom.createTextNode(g_uad_3[1]))

                    node_1_3_3 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_3_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataRecordClass/DemExtendedDataRecordUpdate'))
                    node_1_3_3_2 = dom.createElement('VALUE')
                    node_1_3_3_2.appendChild(dom.createTextNode(g_uad_4[1]))

                    node_1_4 = dom.createElement('REFERENCE-VALUES')

                    node_1_4_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataRecordClass/DemDataElementClassRef'))
                    node_1_4_1_2 = dom.createElement('VALUE-REF')
                    node_1_4_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{g_uad_5[1]}'))

                    tmp_node = [
                        (node_1_4_1, [node_1_4_1_1, node_1_4_1_2]),
                        (node_1_3_3, [node_1_3_3_1, node_1_3_3_2]),
                        (node_1_3_2, [node_1_3_2_1, node_1_3_2_2]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_4, [node_1_4_1]),
                        (node_1_3, [node_1_3_1, node_1_3_2, node_1_3_3]),
                        (node_1, [node_1_1, node_1_2, node_1_3, node_1_4]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemExtendedDataRecordClasss

    def parser_DemExtendedDataClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemExtendedDataClass, TMP_UAD_DemExtendedDataRecordClassRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemExtendedDataClass = str(row['DemExtendedDataClass']).split('/')
            TMP_UAD_DemExtendedDataRecordClassRef = str(row['DemExtendedDataRecordClassRef']).split('/')
        try:
            TMP_UAD_DemExtendedDataClass.append(TMP_UAD_DemExtendedDataClass[0])
            TMP_UAD_DemExtendedDataRecordClassRef.append(TMP_UAD_DemExtendedDataRecordClassRef[0])
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemExtendedDataClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3_1_2_text = list()
                        node_1_3 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE-REF')[0]
                            node_1_3_x_1_attr = node_1_3_x_1.getAttribute('DEST')
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemExtendedDataRecordClassRef') != -1:
                                node_1_3_1_2_text.append(str(node_1_3_x_2.firstChild.data).split('/')[-1])

                        if method == 'read':
                            TMP_R_DemExtendedDataClass = node_1_1_text
                            TMP_R_DemExtendedDataRecordClassRef = node_1_3_1_2_text

                            atom = PrototypeDemExtendedDataClasss(TMP_R_DemExtendedDataClass, TMP_R_DemExtendedDataRecordClassRef)
                            self._DemExtendedDataClasss += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemExtendedDataClass[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemExtendedDataClass[1]
                            node_1.removeChild(node_1_3)
                            node_1_3 = dom.createElement('REFERENCE-VALUES')
                            for ref in TMP_UAD_DemExtendedDataRecordClassRef[1].split('\n'):
                                if ref != '':
                                    node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                                    node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                                    node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                                    node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataClass/DemExtendedDataRecordClassRef'))
                                    node_1_3_x_2 = dom.createElement('VALUE-REF')
                                    node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                                    node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                                    node_1_3_x.appendChild(node_1_3_x_1)
                                    node_1_3_x.appendChild(node_1_3_x_2)
                                    node_1_3.appendChild(node_1_3_x)
                                node_1.appendChild(node_1_3)
                        if method == 'remove' and node_1_1_text == TMP_UAD_DemExtendedDataClass[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemExtendedDataClass[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemExtendedDataClass[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataClass'))

                    node_1_3 = dom.createElement('REFERENCE-VALUES')
                    for ref in TMP_UAD_DemExtendedDataRecordClassRef[1].split('\n'):
                        if ref != '':
                            node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                            node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                            node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                            node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemExtendedDataClass/DemExtendedDataRecordClassRef'))
                            node_1_3_x_2 = dom.createElement('VALUE-REF')
                            node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                            node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                            node_1_3_x.appendChild(node_1_3_x_1)
                            node_1_3_x.appendChild(node_1_3_x_2)
                            node_1_3.appendChild(node_1_3_x)
                    tmp_node = [
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemExtendedDataClasss

    def parser_DemFreezeFrameClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemFreezeFrameClass, TMP_UAD_DemDidClassRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemFreezeFrameClass = str(row['DemFreezeFrameClass']).split('/')
            TMP_UAD_DemDidClassRef = str(row['DemDidClassRef']).split('/')
        try:
            TMP_UAD_DemFreezeFrameClass.append(TMP_UAD_DemFreezeFrameClass[0])
            TMP_UAD_DemDidClassRef.append(TMP_UAD_DemDidClassRef[0])
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemFreezeFrameClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3_1_2_text = list()
                        node_1_3 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE-REF')[0]
                            node_1_3_x_1_attr = node_1_3_x_1.getAttribute('DEST')
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemDidClassRef') != -1:
                                node_1_3_1_2_text.append(str(node_1_3_x_2.firstChild.data).split('/')[-1])

                        if method == 'read':
                            TMP_R_DemFreezeFrameClass = node_1_1_text
                            TMP_R_DemDidClassRef = node_1_3_1_2_text

                            atom = PrototypeDemFreezeFrameClasss(TMP_R_DemFreezeFrameClass, TMP_R_DemDidClassRef)
                            self._DemFreezeFrameClasss += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemFreezeFrameClass[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemFreezeFrameClass[1]
                            node_1.removeChild(node_1_3)
                            node_1_3 = dom.createElement('REFERENCE-VALUES')
                            for ref in TMP_UAD_DemDidClassRef[1].split('\n'):
                                if ref != '':
                                    node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                                    node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                                    node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                                    node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameClass/DemDidClassRef'))
                                    node_1_3_x_2 = dom.createElement('VALUE-REF')
                                    node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                                    node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                                    node_1_3_x.appendChild(node_1_3_x_1)
                                    node_1_3_x.appendChild(node_1_3_x_2)
                                    node_1_3.appendChild(node_1_3_x)
                                node_1.appendChild(node_1_3)
                        if method == 'remove' and node_1_1_text == TMP_UAD_DemFreezeFrameClass[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemFreezeFrameClass[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameClass[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameClass'))

                    node_1_3 = dom.createElement('REFERENCE-VALUES')
                    for ref in TMP_UAD_DemDidClassRef[1].split('\n'):
                        if ref != '':
                            node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                            node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                            node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                            node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameClass/DemDidClassRef'))
                            node_1_3_x_2 = dom.createElement('VALUE-REF')
                            node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                            node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                            node_1_3_x.appendChild(node_1_3_x_1)
                            node_1_3_x.appendChild(node_1_3_x_2)
                            node_1_3.appendChild(node_1_3_x)
                    tmp_node = [
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemFreezeFrameClasss

    def parser_DemFreezeFrameRecordClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemFreezeFrameRecordClass, TMP_UAD_DemFreezeFrameRecordNumber, TMP_UAD_DemFreezeFrameRecordTrigger, TMP_UAD_DemFreezeFrameRecordUpdate
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemFreezeFrameRecordClass = str(row['DemFreezeFrameRecordClass']).split('/')
            TMP_UAD_DemFreezeFrameRecordNumber = str(row['DemFreezeFrameRecordNumber*']).split('/')
            TMP_UAD_DemFreezeFrameRecordTrigger = str(row['DemFreezeFrameRecordTrigger*']).split('/')
            TMP_UAD_DemFreezeFrameRecordUpdate = str(row['DemFreezeFrameRecordUpdate*']).split('/')
        try:
            TMP_UAD_DemFreezeFrameRecordClass.append(TMP_UAD_DemFreezeFrameRecordClass[0])
            TMP_UAD_DemFreezeFrameRecordNumber.append(TMP_UAD_DemFreezeFrameRecordNumber[0])
            TMP_UAD_DemFreezeFrameRecordTrigger.append(TMP_UAD_DemFreezeFrameRecordTrigger[0])
            TMP_UAD_DemFreezeFrameRecordUpdate.append(TMP_UAD_DemFreezeFrameRecordUpdate[0])

            TMP_UAD_DemFreezeFrameRecordNumber[1] = str(TMP_UAD_DemFreezeFrameRecordNumber[1]).replace('0x', '')
            TMP_UAD_DemFreezeFrameRecordNumber[1] = str(TMP_UAD_DemFreezeFrameRecordNumber[1]).replace('0X', '')
            TMP_UAD_DemFreezeFrameRecordNumber[1] = str(int(TMP_UAD_DemFreezeFrameRecordNumber[1], 16))
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    node_1_3_1_2_text = ''
                    node_1_3_2_2_text = ''
                    node_1_3_3_2_text = ''
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemFreezeFrameRecordClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]

                        node_1_3_1 = node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        node_1_3_1_1 = node_1_3_1.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_1_2 = node_1_3_1.getElementsByTagName('VALUE')[0]
                        node_1_3_1_1_text = str(node_1_3_1_1.firstChild.data)
                        node_1_3_1_1_attr = node_1_3_1_1.getAttribute('DEST')
                        if node_1_3_1_1_text.find('/DemFreezeFrameRecordNumber') != -1:
                            node_1_3_1_2_text = str(node_1_3_1_2.firstChild.data)

                        node_1_3_2 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                        node_1_3_2_1 = node_1_3_2.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_2_2 = node_1_3_2.getElementsByTagName('VALUE')[0]
                        node_1_3_2_1_text = str(node_1_3_2_1.firstChild.data)
                        node_1_3_2_1_attr = node_1_3_2_1.getAttribute('DEST')
                        if node_1_3_2_1_text.find('/DemFreezeFrameRecordTrigger') != -1:
                            node_1_3_2_2_text = str(node_1_3_2_2.firstChild.data)

                        node_1_3_3 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[1]
                        node_1_3_3_1 = node_1_3_3.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_3_2 = node_1_3_3.getElementsByTagName('VALUE')[0]
                        node_1_3_3_1_text = str(node_1_3_3_1.firstChild.data)
                        node_1_3_3_1_attr = node_1_3_3_1.getAttribute('DEST')
                        if node_1_3_3_1_text.find('/DemFreezeFrameRecordUpdate') != -1:
                            node_1_3_3_2_text = str(node_1_3_3_2.firstChild.data)

                        if method == 'read':
                            TMP_R_DemFreezeFrameRecordClass = node_1_1_text
                            TMP_R_DemFreezeFrameRecordNumber = node_1_3_1_2_text
                            TMP_R_DemFreezeFrameRecordTrigger = node_1_3_2_2_text
                            TMP_R_DemFreezeFrameRecordUpdate = node_1_3_3_2_text

                            atom = PrototypeDemFreezeFrameRecordClasss(TMP_R_DemFreezeFrameRecordClass, TMP_R_DemFreezeFrameRecordNumber, TMP_R_DemFreezeFrameRecordTrigger, TMP_R_DemFreezeFrameRecordUpdate)
                            self._DemFreezeFrameRecordClasss += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemFreezeFrameRecordClass[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemFreezeFrameRecordClass[1]
                            node_1_3_1_2.firstChild.data = TMP_UAD_DemFreezeFrameRecordNumber[1]
                            node_1_3_2_2.firstChild.data = TMP_UAD_DemFreezeFrameRecordTrigger[1]
                            node_1_3_3_2.firstChild.data = TMP_UAD_DemFreezeFrameRecordUpdate[1]

                        if method == 'remove' and node_1_1_text == TMP_UAD_DemFreezeFrameRecordClass[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemFreezeFrameRecordClass[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameRecordClass[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecordClass'))

                    node_1_3 = dom.createElement('PARAMETER-VALUES')

                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecordClass/DemFreezeFrameRecordNumber'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameRecordNumber[1]))

                    node_1_3_2 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_2_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecordClass/DemFreezeFrameRecordTrigger'))
                    node_1_3_2_2 = dom.createElement('VALUE')
                    node_1_3_2_2.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameRecordTrigger[1]))

                    node_1_3_3 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_3_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecordClass/DemFreezeFrameRecordUpdate'))
                    node_1_3_3_2 = dom.createElement('VALUE')
                    node_1_3_3_2.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameRecordUpdate[1]))

                    tmp_node = [
                        (node_1_3_3, [node_1_3_3_1, node_1_3_3_2]),
                        (node_1_3_2, [node_1_3_2_1, node_1_3_2_2]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_3, [node_1_3_1, node_1_3_2, node_1_3_3]),
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemFreezeFrameRecordClasss

    def parser_DemFreezeFrameRecNumClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemFreezeFrameRecNumClass, TMP_UAD_DemFreezeFrameRecordClassRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemFreezeFrameRecNumClass = str(row['DemFreezeFrameRecNumClass']).split('/')
            TMP_UAD_DemFreezeFrameRecordClassRef = str(row['DemFreezeFrameRecordClassRef']).split('/')
        try:
            TMP_UAD_DemFreezeFrameRecNumClass.append(TMP_UAD_DemFreezeFrameRecNumClass[0])
            TMP_UAD_DemFreezeFrameRecordClassRef.append(TMP_UAD_DemFreezeFrameRecordClassRef[0])
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemFreezeFrameRecNumClass') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3_1_2_text = list()
                        node_1_3 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE-REF')[0]
                            node_1_3_x_1_attr = node_1_3_x_1.getAttribute('DEST')
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemFreezeFrameRecordClassRef') != -1:
                                node_1_3_1_2_text.append(str(node_1_3_x_2.firstChild.data).split('/')[-1])

                        if method == 'read':
                            TMP_R_DemFreezeFrameRecNumClass = node_1_1_text
                            TMP_R_DemFreezeFrameRecordClassRef = node_1_3_1_2_text

                            atom = PrototypeDemFreezeFrameRecNumClasss(TMP_R_DemFreezeFrameRecNumClass, TMP_R_DemFreezeFrameRecordClassRef)
                            self._DemFreezeFrameRecNumClasss += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemFreezeFrameRecNumClass[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemFreezeFrameRecNumClass[1]
                            node_1.removeChild(node_1_3)
                            node_1_3 = dom.createElement('REFERENCE-VALUES')
                            for ref in TMP_UAD_DemFreezeFrameRecordClassRef[1].split('\n'):
                                if ref != '':
                                    node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                                    node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                                    node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                                    node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecNumClass/DemFreezeFrameRecordClassRef'))
                                    node_1_3_x_2 = dom.createElement('VALUE-REF')
                                    node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                                    node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                                    node_1_3_x.appendChild(node_1_3_x_1)
                                    node_1_3_x.appendChild(node_1_3_x_2)
                                    node_1_3.appendChild(node_1_3_x)
                                node_1.appendChild(node_1_3)
                        if method == 'remove' and node_1_1_text == TMP_UAD_DemFreezeFrameRecNumClass[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemFreezeFrameRecNumClass[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemFreezeFrameRecNumClass[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecNumClass'))

                    node_1_3 = dom.createElement('REFERENCE-VALUES')
                    for ref in TMP_UAD_DemFreezeFrameRecordClassRef[1].split('\n'):
                        if ref != '':
                            node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                            node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                            node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                            node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemFreezeFrameRecNumClass/DemFreezeFrameRecordClassRef'))
                            node_1_3_x_2 = dom.createElement('VALUE-REF')
                            node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                            node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                            node_1_3_x.appendChild(node_1_3_x_1)
                            node_1_3_x.appendChild(node_1_3_x_2)
                            node_1_3.appendChild(node_1_3_x)
                    tmp_node = [
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemFreezeFrameRecNumClasss

    def parser_DemEnableConditionGroups(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemEnableConditionGroup, TMP_UAD_DemEnableConditionRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemEnableConditionGroup = str(row['DemEnableConditionGroup']).split('/')
            TMP_UAD_DemEnableConditionRef = str(row['DemEnableConditionRef']).split('/')
        try:
            TMP_UAD_DemEnableConditionGroup.append(TMP_UAD_DemEnableConditionGroup[0])
            TMP_UAD_DemEnableConditionRef.append(TMP_UAD_DemEnableConditionRef[0])
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemEnableConditionGroup') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        node_1_3_1_2_text = list()
                        node_1_3 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                        for node_1_3_x in node_1_3.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                            node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE-REF')[0]
                            node_1_3_x_1_attr = node_1_3_x_1.getAttribute('DEST')
                            node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                            if node_1_3_x_1_text.find('/DemEnableConditionRef') != -1:
                                node_1_3_1_2_text.append(str(node_1_3_x_2.firstChild.data).split('/')[-1])

                        if method == 'read':
                            TMP_R_DemEnableConditionGroup = node_1_1_text
                            TMP_R_DemEnableConditionRef = node_1_3_1_2_text

                            atom = PrototypeDemEnableConditionGroups(TMP_R_DemEnableConditionGroup, TMP_R_DemEnableConditionRef)
                            self._DemEnableConditionGroups += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemEnableConditionGroup[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemEnableConditionGroup[1]
                            node_1.removeChild(node_1_3)
                            node_1_3 = dom.createElement('REFERENCE-VALUES')
                            for ref in TMP_UAD_DemEnableConditionRef[1].split('\n'):
                                if ref != '':
                                    node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                                    node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                                    node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                                    node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemEnableConditionGroup/DemEnableConditionRef'))
                                    node_1_3_x_2 = dom.createElement('VALUE-REF')
                                    node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                                    node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                                    node_1_3_x.appendChild(node_1_3_x_1)
                                    node_1_3_x.appendChild(node_1_3_x_2)
                                    node_1_3.appendChild(node_1_3_x)
                                node_1.appendChild(node_1_3)
                        if method == 'remove' and node_1_1_text == TMP_UAD_DemEnableConditionGroup[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemEnableConditionGroup[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemEnableConditionGroup[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemEnableConditionGroup'))

                    node_1_3 = dom.createElement('REFERENCE-VALUES')
                    for ref in TMP_UAD_DemEnableConditionRef[1].split('\n'):
                        if ref != '':
                            node_1_3_x = dom.createElement('ECUC-REFERENCE-VALUE')
                            node_1_3_x_1 = dom.createElement('DEFINITION-REF')
                            node_1_3_x_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                            node_1_3_x_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemEnableConditionGroup/DemEnableConditionRef'))
                            node_1_3_x_2 = dom.createElement('VALUE-REF')
                            node_1_3_x_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                            node_1_3_x_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{ref.strip()}'))
                            node_1_3_x.appendChild(node_1_3_x_1)
                            node_1_3_x.appendChild(node_1_3_x_2)
                            node_1_3.appendChild(node_1_3_x)
                    tmp_node = [
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemEnableConditionGroups

    def parser_DemEnableConditions(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemEnableCondition, TMP_UAD_DemEnableConditionStatus
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemEnableCondition = str(row['DemEnableCondition']).split('/')
            TMP_UAD_DemEnableConditionStatus = str(row['DemEnableConditionStatus*']).split('/')
        try:
            TMP_UAD_DemEnableCondition.append(TMP_UAD_DemEnableCondition[0])
            TMP_UAD_DemEnableConditionStatus.append(TMP_UAD_DemEnableConditionStatus[0])
        except Exception as e:
            pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemEnableCondition') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF' and node_1_2_text.find('/DemEnableConditionGroup') == -1:
                        # PARAMETER-VALUES
                        node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]
                        node_1_3_1 = node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        node_1_3_1_1 = node_1_3_1.getElementsByTagName('DEFINITION-REF')[0]
                        node_1_3_1_2 = node_1_3_1.getElementsByTagName('VALUE')[0]
                        node_1_3_1_2_text = str(node_1_3_1_2.firstChild.data)
                        if method == 'read':
                            TMP_R_DemEnableCondition = node_1_1_text
                            TMP_R_DemEnableConditionStatus = node_1_3_1_2_text

                            atom = PrototypeDemEnableConditions(TMP_R_DemEnableCondition, TMP_R_DemEnableConditionStatus)
                            self._DemEnableConditions += [atom]

                        if method == 'update' and node_1_1_text == TMP_UAD_DemEnableCondition[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemEnableCondition[1]
                            node_1_3_1_2.firstChild.data = TMP_UAD_DemEnableConditionStatus[1].lower()

                        if method == 'remove' and node_1_1_text == TMP_UAD_DemEnableCondition[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemEnableCondition[0]:
                            add_item_exist = True

                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemEnableCondition[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemEnableCondition'))

                    node_1_3 = dom.createElement('PARAMETER-VALUES')

                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemEnableCondition/DemEnableConditionStatus'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(TMP_UAD_DemEnableConditionStatus[1].lower()))

                    tmp_node = [
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_3, [node_1_3_1]),
                        (node_1, [node_1_1, node_1_2, node_1_3]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemEnableConditions

    def parser_DemDTCAttributess(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemDTCAttributes, TMP_UAD_DemAgingAllowed, TMP_UAD_DemAgingCycleCounterThreshold, TMP_UAD_DemDTCPriority, \
            TMP_UAD_DemDTCSignificance, TMP_UAD_DemImmediateNvStorage, TMP_UAD_DemMaxNumberFreezeFrameRecords, TMP_UAD_DemAgingCycleRef, \
            TMP_UAD_DemExtendedDataClassRef, TMP_UAD_DemFreezeFrameClassRef, TMP_UAD_DemFreezeFrameRecNumClassRef, TMP_UAD_DemMemoryDestinationRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemDTCAttributes = str(row['DemDTCAttributes']).split('/')
            TMP_UAD_DemAgingAllowed = str(row['DemAgingAllowed*']).split('/')
            TMP_UAD_DemAgingCycleCounterThreshold = str(row['DemAgingCycleCounterThreshold']).split('/')
            TMP_UAD_DemDTCPriority = str(row['DemDTCPriority*']).split('/')
            TMP_UAD_DemDTCSignificance = str(row['DemDTCSignificance']).split('/')
            TMP_UAD_DemImmediateNvStorage = str(row['DemImmediateNvStorage*']).split('/')
            TMP_UAD_DemMaxNumberFreezeFrameRecords = str(row['DemMaxNumberFreezeFrameRecords']).split('/')
            TMP_UAD_DemAgingCycleRef = str(row['DemAgingCycleRef*']).split('/')
            TMP_UAD_DemExtendedDataClassRef = str(row['DemExtendedDataClassRef']).split('/')
            TMP_UAD_DemFreezeFrameClassRef = str(row['DemFreezeFrameClassRef']).split('/')
            TMP_UAD_DemFreezeFrameRecNumClassRef = str(row['DemFreezeFrameRecNumClassRef']).split('/')
            TMP_UAD_DemMemoryDestinationRef = str(row['DemMemoryDestinationRef']).split('/')
            try:
                for item in [TMP_UAD_DemDTCAttributes, TMP_UAD_DemAgingAllowed, TMP_UAD_DemAgingCycleCounterThreshold, TMP_UAD_DemDTCPriority,
                             TMP_UAD_DemDTCSignificance, TMP_UAD_DemImmediateNvStorage, TMP_UAD_DemMaxNumberFreezeFrameRecords, TMP_UAD_DemAgingCycleRef,
                             TMP_UAD_DemExtendedDataClassRef, TMP_UAD_DemFreezeFrameClassRef, TMP_UAD_DemFreezeFrameRecNumClassRef, TMP_UAD_DemMemoryDestinationRef]:
                    item.append(item[0])
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemConfigSet':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for node_1 in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    node_1_1 = node_1.getElementsByTagName('SHORT-NAME')[0]
                    node_1_2 = node_1.getElementsByTagName('DEFINITION-REF')[0]
                    node_1_1_text = str(node_1_1.firstChild.data)
                    node_1_2_attr = node_1_2.getAttribute('DEST')
                    node_1_2_text = str(node_1_2.firstChild.data)
                    if node_1_2_text.find('/DemDTCAttributes') != -1 and node_1_2_attr == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        # Nodes
                        node_1_3_1_2 = object
                        node_1_3_2_2 = object
                        node_1_3_3_2 = object
                        node_1_3_4_2 = object
                        node_1_3_5_2 = object
                        node_1_3_6_2 = object
                        node_1_4_1_2 = object
                        node_1_4_2_2 = object
                        node_1_4_3_2 = object
                        node_1_4_4_2 = object
                        node_1_4_5_2 = object
                        # PARAMETER-VALUES
                        node_1_3_1_2_text = ''
                        node_1_3_2_2_text = ''
                        node_1_3_3_2_text = ''
                        node_1_3_4_2_text = ''
                        node_1_3_5_2_text = ''
                        # REFERENCE-VALUES
                        node_1_4_1_2_text = ''
                        node_1_4_2_2_text = ''
                        node_1_4_3_2_text = ''
                        node_1_4_4_2_text = ''
                        node_1_4_5_2_text = ''
                        try:
                            node_1_3 = node_1.getElementsByTagName('PARAMETER-VALUES')[0]
                            for node_1_3_x in node_1_3.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE'):
                                node_1_3_x_1 = node_1_3_x.getElementsByTagName('DEFINITION-REF')[0]
                                node_1_3_x_2 = node_1_3_x.getElementsByTagName('VALUE')[0]
                                node_1_3_x_1_text = str(node_1_3_x_1.firstChild.data)
                                if node_1_3_x_1_text.find('/DemAgingAllowed') != -1:
                                    node_1_3_1_2 = node_1_3_x_2
                                    node_1_3_1_2_text = str(node_1_3_x_2.firstChild.data)
                                if node_1_3_x_1_text.find('/DemAgingCycleCounterThreshold') != -1:
                                    node_1_3_2_2 = node_1_3_x_2
                                    node_1_3_2_2_text = str(node_1_3_x_2.firstChild.data)
                                if node_1_3_x_1_text.find('/DemDTCPriority') != -1:
                                    node_1_3_3_2 = node_1_3_x_2
                                    node_1_3_3_2_text = str(node_1_3_x_2.firstChild.data)
                                if node_1_3_x_1_text.find('/DemImmediateNvStorage') != -1:
                                    node_1_3_5_2 = node_1_3_x_2
                                    node_1_3_5_2_text = str(node_1_3_x_2.firstChild.data)
                                if node_1_3_x_1_text.find('/DemMaxNumberFreezeFrameRecords') != -1:
                                    node_1_3_6_2 = node_1_3_x_2
                                    node_1_3_6_2_text = str(node_1_3_x_2.firstChild.data)
                            node_1_3_4 = node_1_3.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                            node_1_3_4_1 = node_1_3_4.getElementsByTagName('DEFINITION-REF')[0]
                            node_1_3_4_2 = node_1_3_4.getElementsByTagName('VALUE')[0]
                            node_1_3_4_1_text = str(node_1_3_4_1.firstChild.data)
                            if node_1_3_4_1_text.find('/DemDTCSignificance') != -1:
                                node_1_3_4_2_text = str(node_1_3_4_2.firstChild.data)

                            node_1_4 = node_1.getElementsByTagName('REFERENCE-VALUES')[0]
                            for node_1_4_x in node_1_4.getElementsByTagName('ECUC-REFERENCE-VALUE'):
                                node_1_4_x_1 = node_1_4_x.getElementsByTagName('DEFINITION-REF')[0]
                                node_1_4_x_2 = node_1_4_x.getElementsByTagName('VALUE-REF')[0]
                                node_1_4_x_1_text = str(node_1_4_x_1.firstChild.data)
                                if node_1_4_x_1_text.find('/DemAgingCycleRef') != -1:
                                    node_1_4_1_2 = node_1_4_x_2
                                    node_1_4_1_2_text = str(node_1_4_x_2.firstChild.data)
                                if node_1_4_x_1_text.find('/DemExtendedDataClassRef') != -1:
                                    node_1_4_2_2 = node_1_4_x_2
                                    node_1_4_2_2_text = str(node_1_4_x_2.firstChild.data)
                                if node_1_4_x_1_text.find('/DemFreezeFrameClassRef') != -1:
                                    node_1_4_3_2 = node_1_4_x_2
                                    node_1_4_3_2_text = str(node_1_4_x_2.firstChild.data)
                                if node_1_4_x_1_text.find('/DemFreezeFrameRecNumClassRef') != -1:
                                    node_1_4_4_2 = node_1_4_x_2
                                    node_1_4_4_2_text = str(node_1_4_x_2.firstChild.data)
                                if node_1_4_x_1_text.find('/DemMemoryDestinationRef') != -1:
                                    node_1_4_5_2 = node_1_4_x_2
                                    node_1_4_5_2_text = str(node_1_4_x_2.firstChild.data)
                        except Exception as e:
                            pass

                        if method == 'read':
                            TMP_R_DemDTCAttributes = node_1_1_text
                            TMP_R_DemAgingAllowed = node_1_3_1_2_text
                            TMP_R_DemAgingCycleCounterThreshold = int(node_1_3_2_2_text)
                            TMP_R_DemDTCPriority = int(node_1_3_3_2_text)
                            TMP_R_DemDTCSignificance = node_1_3_4_2_text
                            TMP_R_DemImmediateNvStorage = node_1_3_5_2_text
                            TMP_R_DemMaxNumberFreezeFrameRecords = int(node_1_3_6_2_text)
                            TMP_R_DemAgingCycleRef = node_1_4_1_2_text
                            TMP_R_DemExtendedDataClassRef = node_1_4_2_2_text
                            TMP_R_DemFreezeFrameClassRef = node_1_4_3_2_text
                            TMP_R_DemFreezeFrameRecNumClassRef = node_1_4_4_2_text
                            TMP_R_DemMemoryDestinationRef = node_1_4_5_2_text

                            atom = PrototypeDemDTCAttributess(TMP_R_DemDTCAttributes, TMP_R_DemAgingAllowed, TMP_R_DemAgingCycleCounterThreshold, TMP_R_DemDTCPriority,
                                                              TMP_R_DemDTCSignificance, TMP_R_DemImmediateNvStorage, TMP_R_DemMaxNumberFreezeFrameRecords, TMP_R_DemAgingCycleRef,
                                                              TMP_R_DemExtendedDataClassRef, TMP_R_DemFreezeFrameClassRef, TMP_R_DemFreezeFrameRecNumClassRef, TMP_R_DemMemoryDestinationRef)
                            self._DemDTCAttributes += [atom]
                        if method == 'update' and node_1_1_text == TMP_UAD_DemDTCAttributes[0]:
                            node_1_1.firstChild.data = TMP_UAD_DemDTCAttributes[1]
                            node_1_3_1_2.firstChild.data = TMP_UAD_DemAgingAllowed[1].lower()
                            node_1_3_2_2.firstChild.data = int(float(TMP_UAD_DemAgingCycleCounterThreshold[1]))
                            node_1_3_3_2.firstChild.data = int(float(TMP_UAD_DemDTCPriority[1]))
                            node_1_3_4_2.firstChild.data = TMP_UAD_DemDTCSignificance[1].upper()
                            node_1_3_5_2.firstChild.data = TMP_UAD_DemImmediateNvStorage[1].lower()
                            node_1_3_6_2.firstChild.data = int(float(TMP_UAD_DemMaxNumberFreezeFrameRecords[1]))
                            node_1_4_1_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemAgingCycleRef[1]}'
                            node_1_4_2_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemExtendedDataClassRef[1]}'
                            node_1_4_3_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemFreezeFrameClassRef[1]}'
                            node_1_4_4_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemFreezeFrameRecNumClassRef[1]}'
                            node_1_4_5_2.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemMemoryDestinationRef[1]}'
                        if method == 'remove' and node_1_1_text == TMP_UAD_DemDTCAttributes[0]:
                            target.removeChild(node_1)

                        if method == 'add' and node_1_1_text == TMP_UAD_DemDTCAttributes[0]:
                            add_item_exist = True
                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemDTCAttributes[1]))

                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes'))

                    node_1_3 = dom.createElement('PARAMETER-VALUES')

                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemAgingAllowed'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(TMP_UAD_DemAgingAllowed[1].lower()))

                    node_1_3_2 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_2_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemAgingCycleCounterThreshold'))
                    node_1_3_2_2 = dom.createElement('VALUE')
                    node_1_3_2_2.appendChild(dom.createTextNode(str(int(float(TMP_UAD_DemAgingCycleCounterThreshold[1])))))

                    node_1_3_3 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_3_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemDTCPriority'))
                    node_1_3_3_2 = dom.createElement('VALUE')
                    node_1_3_3_2.appendChild(dom.createTextNode(str(int(float(TMP_UAD_DemDTCPriority[1])))))

                    node_1_3_4 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_4_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_4_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_4_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemDTCSignificance'))
                    node_1_3_4_2 = dom.createElement('VALUE')
                    node_1_3_4_2.appendChild(dom.createTextNode(TMP_UAD_DemDTCSignificance[1].upper()))

                    node_1_3_5 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_5_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_5_1.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    node_1_3_5_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemImmediateNvStorage'))
                    node_1_3_5_2 = dom.createElement('VALUE')
                    node_1_3_5_2.appendChild(dom.createTextNode(TMP_UAD_DemImmediateNvStorage[1].lower()))

                    node_1_3_6 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_6_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_6_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_6_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemMaxNumberFreezeFrameRecords'))
                    node_1_3_6_2 = dom.createElement('VALUE')
                    node_1_3_6_2.appendChild(dom.createTextNode(str(int(float(TMP_UAD_DemMaxNumberFreezeFrameRecords[1])))))

                    node_1_4 = dom.createElement('REFERENCE-VALUES')

                    node_1_4_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemAgingCycleRef'))
                    node_1_4_1_2 = dom.createElement('VALUE-REF')
                    node_1_4_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemAgingCycleRef[1]}'))

                    node_1_4_2 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_2_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemExtendedDataClassRef'))
                    node_1_4_2_2 = dom.createElement('VALUE-REF')
                    node_1_4_2_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_2_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemExtendedDataClassRef[1]}'))

                    node_1_4_3 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_3_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_3_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_3_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemFreezeFrameClassRef'))
                    node_1_4_3_2 = dom.createElement('VALUE-REF')
                    node_1_4_3_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_3_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemFreezeFrameClassRef[1]}'))

                    node_1_4_4 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_4_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_4_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_4_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemFreezeFrameRecNumClassRef'))
                    node_1_4_4_2 = dom.createElement('VALUE-REF')
                    node_1_4_4_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_4_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemFreezeFrameRecNumClassRef[1]}'))

                    node_1_4_5 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_5_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_5_1.setAttribute('DEST', "ECUC-CHOICE-REFERENCE-DEF")
                    node_1_4_5_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTCAttributes/DemMemoryDestinationRef'))
                    node_1_4_5_2 = dom.createElement('VALUE-REF')
                    node_1_4_5_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_5_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemMemoryDestinationRef[1]}'))

                    tmp_node = [
                        (node_1_4_5, [node_1_4_5_1, node_1_4_5_2]),
                        (node_1_4_4, [node_1_4_4_1, node_1_4_4_2]),
                        (node_1_4_3, [node_1_4_3_1, node_1_4_3_2]),
                        (node_1_4_2, [node_1_4_2_1, node_1_4_2_2]),
                        (node_1_4_1, [node_1_4_1_1, node_1_4_1_2]),
                        (node_1_4, [node_1_4_1, node_1_4_2, node_1_4_3, node_1_4_4, node_1_4_5]),
                        (node_1_3_6, [node_1_3_6_1, node_1_3_6_2]),
                        (node_1_3_5, [node_1_3_5_1, node_1_3_5_2]),
                        (node_1_3_4, [node_1_3_4_1, node_1_3_4_2]),
                        (node_1_3_3, [node_1_3_3_1, node_1_3_3_2]),
                        (node_1_3_2, [node_1_3_2_1, node_1_3_2_2]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_3, [node_1_3_1, node_1_3_2, node_1_3_3, node_1_3_4, node_1_3_5, node_1_3_6]),
                        (node_1, [node_1_1, node_1_2, node_1_3, node_1_4]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
                    # del empty ref
                    if TMP_UAD_DemAgingCycleRef[1] == 'nan':
                        node_1_4.removeChild(node_1_4_1)
                    if TMP_UAD_DemExtendedDataClassRef[1] == 'nan':
                        node_1_4.removeChild(node_1_4_2)
                    if TMP_UAD_DemFreezeFrameClassRef[1] == 'nan':
                        node_1_4.removeChild(node_1_4_3)
                    if TMP_UAD_DemFreezeFrameRecNumClassRef[1] == 'nan':
                        node_1_4.removeChild(node_1_4_4)
                    if TMP_UAD_DemMemoryDestinationRef[1] == 'nan':
                        node_1_4.removeChild(node_1_4_5)
        return self._DemDTCAttributes

    def parser_DemDTCs(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemDTC, TMP_UAD_DemDTCSeverity, TMP_UAD_DemDtcValue, TMP_UAD_DemDTCAttributesRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemDTC = str(row['DemDTC']).split('/')
            TMP_UAD_DemDTCSeverity = str(row['DemDTCSeverity']).split('/')
            TMP_UAD_DemDtcValue = str(row['DemDtcValue']).split('/')
            TMP_UAD_DemDTCAttributesRef = str(row['DemDTCAttributesRef*']).split('/')
            try:
                TMP_UAD_DemDTC.append(TMP_UAD_DemDTC[0])
                TMP_UAD_DemDTCSeverity.append(TMP_UAD_DemDTCSeverity[0])
                TMP_UAD_DemDtcValue.append(TMP_UAD_DemDtcValue[0])
                TMP_UAD_DemDTCAttributesRef.append(TMP_UAD_DemDTCAttributesRef[0])

                TMP_UAD_DemDtcValue[1] = str(TMP_UAD_DemDtcValue[1]).replace('0x', '')
                TMP_UAD_DemDtcValue[1] = str(TMP_UAD_DemDtcValue[1]).replace('0X', '')
                TMP_UAD_DemDtcValue[1] = str(int(TMP_UAD_DemDtcValue[1], 16))
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemConfigSet':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                for DemDTCs in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    SHORT_NAME = DemDTCs.getElementsByTagName('SHORT-NAME')[0]
                    DEFINITION_REF = DemDTCs.getElementsByTagName('DEFINITION-REF')[0]
                    DEFINITION_REF_ATTR_DEST = DEFINITION_REF.getAttribute('DEST')
                    SHORT_NAME_TEXT = str(SHORT_NAME.firstChild.data)
                    DEFINITION_REF_TEXT = str(DEFINITION_REF.firstChild.data)

                    if DEFINITION_REF_TEXT.find('/DemDTC') != -1 and DEFINITION_REF_ATTR_DEST == 'ECUC-PARAM-CONF-CONTAINER-DEF' and DEFINITION_REF_TEXT.find('/DemDTCAttributes') == -1:
                        # PARAMETER-VALUES
                        PARAMETER_VALUE = DemDTCs.getElementsByTagName('PARAMETER-VALUES')[0]
                        # PARAMETER-VALUES / ECUC-TEXTUAL-PARAM-VALUE
                        TEXTUAL_PARAM = PARAMETER_VALUE.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                        TEXTUAL_DEFINITION_REF = TEXTUAL_PARAM.getElementsByTagName('DEFINITION-REF')[0]
                        TEXTUAL_VALUE = TEXTUAL_PARAM.getElementsByTagName('VALUE')[0]
                        TEXTUAL_VALUE_TEXT = str(TEXTUAL_VALUE.firstChild.data)
                        # PARAMETER-VALUES / ECUC-NUMERICAL-PARAM-VALUE
                        NUMERICAL_PARAM = PARAMETER_VALUE.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        NUMERICAL_DEFINITION_REF = NUMERICAL_PARAM.getElementsByTagName('DEFINITION-REF')[0]
                        NUMERICAL_VALUE = NUMERICAL_PARAM.getElementsByTagName('VALUE')[0]
                        NUMERICAL_VALUE_TEXT = str(NUMERICAL_VALUE.firstChild.data)
                        # REFERENCE-VALUES
                        REFERENCE_VALUE = DemDTCs.getElementsByTagName('REFERENCE-VALUES')[0]
                        REFERENCE_PARAM = REFERENCE_VALUE.getElementsByTagName('ECUC-REFERENCE-VALUE')[0]
                        REFERENCE_DEFINITION_REF = REFERENCE_PARAM.getElementsByTagName('DEFINITION-REF')[0]
                        REFERENCE_VALUE_REF = REFERENCE_PARAM.getElementsByTagName('VALUE-REF')[0]
                        REFERENCE_VALUE_REF_TEXT = str(REFERENCE_VALUE_REF.firstChild.data)

                        if method == 'read':
                            TMP_R_DemDTC = SHORT_NAME_TEXT
                            TMP_R_DemDTCSeverity = TEXTUAL_VALUE_TEXT
                            TMP_R_DemDtcValue = NUMERICAL_VALUE_TEXT
                            TMP_R_DemDTCAttributesRef = REFERENCE_VALUE_REF_TEXT.split('/')[-1]

                            atom = PrototypeDemDTCs(TMP_R_DemDTC, TMP_R_DemDTCSeverity, TMP_R_DemDtcValue, TMP_R_DemDTCAttributesRef)
                            self._DemDTCs += [atom]

                        if method == 'update' and SHORT_NAME_TEXT == TMP_UAD_DemDTC[0]:
                            SHORT_NAME.firstChild.data = TMP_UAD_DemDTC[1]
                            TEXTUAL_VALUE.firstChild.data = TMP_UAD_DemDTCSeverity[1]
                            NUMERICAL_VALUE.firstChild.data = TMP_UAD_DemDtcValue[1]
                            REFERENCE_VALUE_REF.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{TMP_UAD_DemDTCAttributesRef[1]}'

                        if method == 'remove' and SHORT_NAME_TEXT == TMP_UAD_DemDTC[0]:
                            target.removeChild(DemDTCs)

                        if method == 'add' and SHORT_NAME_TEXT == TMP_UAD_DemDTC[0]:
                            add_item_exist = True
                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemDTC[1]))
                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTC'))
                    node_1_3 = dom.createElement('PARAMETER-VALUES')
                    node_1_3_1 = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTC/DemDTCSeverity'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(TMP_UAD_DemDTCSeverity[1]))
                    node_1_3_2 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_2_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_2_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_2_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTC/DemDtcValue'))
                    node_1_3_2_2 = dom.createElement('VALUE')
                    node_1_3_2_2.appendChild(dom.createTextNode(TMP_UAD_DemDtcValue[1]))
                    node_1_4 = dom.createElement('REFERENCE-VALUES')
                    node_1_4_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemConfigSet/DemDTC/DemDTCAttributesRef'))
                    node_1_4_1_2 = dom.createElement('VALUE-REF')
                    node_1_4_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemConfigSet/{TMP_UAD_DemDTCAttributesRef[1]}'))

                    tmp_node = [
                        (node_1_4_1, [node_1_4_1_1, node_1_4_1_2]),
                        (node_1_4, [node_1_4_1]),
                        (node_1_3_2, [node_1_3_2_1, node_1_3_2_2]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_3, [node_1_3_1, node_1_3_2]),
                        (node_1, [node_1_1, node_1_2, node_1_3, node_1_4]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemDTCs

    def parser_DemDidClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemDidClass, TMP_UAD_DemDidIdentifier, TMP_UAD_DemDidDataElementClassRef
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemDidClass = str(row['DemDidClass']).split('/')
            TMP_UAD_DemDidIdentifier = str(row['DemDidIdentifier*']).split('/')
            TMP_UAD_DemDidDataElementClassRef = str(row['DemDidDataElementClassRef']).split('/')
            try:
                TMP_UAD_DemDidClass.append(TMP_UAD_DemDidClass[0])
                TMP_UAD_DemDidIdentifier.append(TMP_UAD_DemDidIdentifier[0])
                TMP_UAD_DemDidDataElementClassRef.append(TMP_UAD_DemDidDataElementClassRef[0])

                TMP_UAD_DemDidIdentifier[1] = str(TMP_UAD_DemDidIdentifier[1]).replace('0x', '')
                TMP_UAD_DemDidIdentifier[1] = str(TMP_UAD_DemDidIdentifier[1]).replace('0X', '')
                TMP_UAD_DemDidIdentifier[1] = str(int(TMP_UAD_DemDidIdentifier[1], 16))
            except Exception as e:
                pass

        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                TMP_R_DemDidClass = ''
                TMP_R_DemDidIdentifier = ''
                TMP_R_DemDidDataElementClassRef = ''
                for DemDidClass in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    # SHORT-NAME
                    SHORT_NAME = DemDidClass.getElementsByTagName('SHORT-NAME')[0]
                    DEFINITION_REF = DemDidClass.getElementsByTagName('DEFINITION-REF')[0]
                    DEFINITION_REF_ATTR_DEST = DEFINITION_REF.getAttribute('DEST')
                    SHORT_NAME_TEXT = str(SHORT_NAME.firstChild.data)
                    DEFINITION_REF_TEXT = str(DEFINITION_REF.firstChild.data)

                    if DEFINITION_REF_TEXT.find('/DemDidClass') != -1 and DEFINITION_REF_ATTR_DEST == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        # PARAMETER-VALUES
                        PARAMETER_VALUE = DemDidClass.getElementsByTagName('PARAMETER-VALUES')[0]
                        NUMERICAL_PARAM = PARAMETER_VALUE.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        NUMERICAL_DEFINITION_REF = NUMERICAL_PARAM.getElementsByTagName('DEFINITION-REF')[0]
                        NUMERICAL_VALUE = NUMERICAL_PARAM.getElementsByTagName('VALUE')[0]
                        NUMERICAL_VALUE_TEXT = str(NUMERICAL_VALUE.firstChild.data)
                        # REFERENCE-VALUES
                        REFERENCE_VALUE = DemDidClass.getElementsByTagName('REFERENCE-VALUES')[0]
                        REFERENCE_PARAM = REFERENCE_VALUE.getElementsByTagName('ECUC-REFERENCE-VALUE')[0]
                        REFERENCE_DEFINITION_REF = REFERENCE_PARAM.getElementsByTagName('DEFINITION-REF')[0]
                        REFERENCE_VALUE_REF = REFERENCE_PARAM.getElementsByTagName('VALUE-REF')[0]
                        REFERENCE_VALUE_REF_TEXT = str(REFERENCE_VALUE_REF.firstChild.data)

                        if method == 'read':
                            TMP_R_DemDidClass = SHORT_NAME_TEXT
                            TMP_R_DemDidIdentifier = NUMERICAL_VALUE_TEXT
                            TMP_R_DemDidDataElementClassRef = REFERENCE_VALUE_REF_TEXT.split('/')[-1]

                            atom = PrototypeDemDidClasss(TMP_R_DemDidClass, TMP_R_DemDidIdentifier, TMP_R_DemDidDataElementClassRef)
                            self._DemDidClasss += [atom]

                        if method == 'update' and SHORT_NAME_TEXT == TMP_UAD_DemDidClass[0]:
                            SHORT_NAME.firstChild.data = TMP_UAD_DemDidClass[1]
                            NUMERICAL_VALUE.firstChild.data = TMP_UAD_DemDidIdentifier[1]
                            REFERENCE_VALUE_REF.firstChild.data = f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemDidDataElementClassRef[1]}'

                        if method == 'remove' and SHORT_NAME_TEXT == TMP_UAD_DemDidClass[0]:
                            target.removeChild(DemDidClass)

                        if method == 'add' and SHORT_NAME_TEXT == TMP_UAD_DemDidClass[0]:
                            add_item_exist = True
                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    node_1 = dom.createElement('ECUC-CONTAINER-VALUE')
                    node_1_1 = dom.createElement('SHORT-NAME')
                    node_1_1.appendChild(dom.createTextNode(TMP_UAD_DemDidClass[1]))
                    node_1_2 = dom.createElement('DEFINITION-REF')
                    node_1_2.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    node_1_2.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDidClass'))
                    node_1_3 = dom.createElement('PARAMETER-VALUES')
                    node_1_3_1 = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    node_1_3_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_3_1_1.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    node_1_3_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDidClass/DemDidIdentifier'))
                    node_1_3_1_2 = dom.createElement('VALUE')
                    node_1_3_1_2.appendChild(dom.createTextNode(TMP_UAD_DemDidIdentifier[1]))
                    node_1_4 = dom.createElement('REFERENCE-VALUES')
                    node_1_4_1 = dom.createElement('ECUC-REFERENCE-VALUE')
                    node_1_4_1_1 = dom.createElement('DEFINITION-REF')
                    node_1_4_1_1.setAttribute('DEST', "ECUC-REFERENCE-DEF")
                    node_1_4_1_1.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDidClass/DemDidDataElementClassRef'))
                    node_1_4_1_2 = dom.createElement('VALUE-REF')
                    node_1_4_1_2.setAttribute('DEST', "ECUC-CONTAINER-VALUE")
                    node_1_4_1_2.appendChild(dom.createTextNode(f'/ETAS_Project/EcucModuleConfigurationValuess/Dem/DemGeneral/{TMP_UAD_DemDidDataElementClassRef[1]}'))

                    tmp_node = [
                        (node_1_4_1, [node_1_4_1_1, node_1_4_1_2]),
                        (node_1_4, [node_1_4_1]),
                        (node_1_3_1, [node_1_3_1_1, node_1_3_1_2]),
                        (node_1_3, [node_1_3_1]),
                        (node_1, [node_1_1, node_1_2, node_1_3, node_1_4]),
                        (target, [node_1]),
                    ]
                    for item in tmp_node:
                        for each in item[1]:
                            item[0].appendChild(each)
        return self._DemDidClasss

    def parser_DemDataElementClasss(self, root, dom, row: object = None, method: str = 'read'):
        global TMP_UAD_DemDataElementClasss, TMP_UAD_DemDataElementClassType, TMP_UAD_DemDataElementClassTypeName, TMP_UAD_DemDataElementDataSize, TMP_UAD_DemDataElementUsePort, TMP_UAD_DemInternalDataElement
        if method == 'update' or method == 'remove' or method == 'add':
            TMP_UAD_DemDataElementClasss = str(row['DemDataElementClasss']).split('/')
            TMP_UAD_DemDataElementClassType = str(row['DemDataElementClassType']).split('/')
            TMP_UAD_DemDataElementClassTypeName = str(row['DemDataElementClassTypeName']).split('/')
            TMP_UAD_DemDataElementDataSize = str(row['DemDataElementDataSize']).split('/')
            TMP_UAD_DemDataElementUsePort = str(row['DemDataElementUsePort']).split('/')
            TMP_UAD_DemInternalDataElement = str(row['DemInternalDataElement']).split('/')
            try:
                TMP_UAD_DemDataElementClasss.append(TMP_UAD_DemDataElementClasss[0])
                TMP_UAD_DemDataElementClassType.append(TMP_UAD_DemDataElementClassType[0])
                TMP_UAD_DemDataElementClassTypeName.append(TMP_UAD_DemDataElementClassTypeName[0])
                TMP_UAD_DemDataElementDataSize.append(TMP_UAD_DemDataElementDataSize[0])
                TMP_UAD_DemDataElementUsePort.append(TMP_UAD_DemDataElementUsePort[0])
                TMP_UAD_DemInternalDataElement.append(TMP_UAD_DemInternalDataElement[0])
            except Exception as e:
                pass

        update_processing_state = False
        add_item_exist = False
        for node in root.getElementsByTagName("SHORT-NAME"):
            if str(node.firstChild.data) == 'DemGeneral':
                target = node.parentNode.getElementsByTagName("SUB-CONTAINERS")[0]
                TMP_R_DemDataElementClasss = ''
                TMP_R_DemDataElementClassTypeName = ''
                TMP_R_DemDataElementClassType = ''
                for DemDataEleClass in target.getElementsByTagName("ECUC-CONTAINER-VALUE"):
                    TMP_R_DemDataElementDataSize = ''
                    TMP_R_DemDataElementUsePort = ''
                    TMP_R_DemInternalDataElement = ''
                    SHORT_NAME = DemDataEleClass.getElementsByTagName('SHORT-NAME')[0]
                    DEFINITION_REF = DemDataEleClass.getElementsByTagName('DEFINITION-REF')[0]
                    DEFINITION_REF_ATTR_DEST = DEFINITION_REF.getAttribute('DEST')
                    SHORT_NAME_TEXT = str(SHORT_NAME.firstChild.data)
                    DEFINITION_REF_TEXT = str(DEFINITION_REF.firstChild.data)

                    if DEFINITION_REF_TEXT.find('/DemDataElementClass') != -1 and DEFINITION_REF_ATTR_DEST == 'ECUC-CHOICE-CONTAINER-DEF':
                        if method == 'read':
                            TMP_R_DemDataElementClasss = SHORT_NAME_TEXT
                        if method == 'update' and SHORT_NAME_TEXT == TMP_UAD_DemDataElementClasss[0]:
                            SHORT_NAME.firstChild.data = TMP_UAD_DemDataElementClasss[1]
                            update_processing_state = True
                        if method == 'remove' and SHORT_NAME_TEXT == TMP_UAD_DemDataElementClasss[0]:
                            target.removeChild(DemDataEleClass)
                        if method == 'add' and SHORT_NAME_TEXT == TMP_UAD_DemDataElementClasss[0]:
                            add_item_exist = True

                    if DEFINITION_REF_TEXT.find('/DemExternalCSDataElementClass') != -1 and DEFINITION_REF_ATTR_DEST == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        if method == 'read':
                            TMP_R_DemDataElementClassTypeName = SHORT_NAME_TEXT
                            TMP_R_DemDataElementClassType = 'DemExternalCSDataElementClass'
                        if method == 'update' and update_processing_state is True:
                            # DEFINITION_REF.firstChild.data = '' # [Disable] DemDataElementClassTypeREF Update
                            SHORT_NAME.firstChild.data = TMP_UAD_DemDataElementClassTypeName[1]

                        PARAM = DemDataEleClass.getElementsByTagName('PARAMETER-VALUES')[0]
                        for item in PARAM.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE'):
                            PARAM_VALUE = item.getElementsByTagName('VALUE')[0]
                            PARAM_DEFINITION_REF = item.getElementsByTagName('DEFINITION-REF')[0]
                            PARAM_DEFINITION_REF_ATTR_DEST = PARAM_DEFINITION_REF.getAttribute('DEST')
                            PARAM_VALUE_TEXT = str(PARAM_VALUE.firstChild.data)
                            PARAM_DEFINITION_REF_TEXT = str(PARAM_DEFINITION_REF.firstChild.data)
                            if PARAM_DEFINITION_REF_TEXT.find('/DemDataElementDataSize') != -1 and PARAM_DEFINITION_REF_ATTR_DEST == 'ECUC-INTEGER-PARAM-DEF':
                                if method == 'read':
                                    TMP_R_DemDataElementDataSize = PARAM_VALUE_TEXT
                                if method == 'update' and update_processing_state is True:
                                    PARAM_VALUE.firstChild.data = int(float(TMP_UAD_DemDataElementDataSize[1]))

                            if PARAM_DEFINITION_REF_TEXT.find('/DemDataElementUsePort') != -1 and PARAM_DEFINITION_REF_ATTR_DEST == 'ECUC-BOOLEAN-PARAM-DEF':
                                if method == 'read':
                                    TMP_R_DemDataElementUsePort = PARAM_VALUE_TEXT
                                if method == 'update' and update_processing_state is True:
                                    PARAM_VALUE.firstChild.data = TMP_UAD_DemDataElementUsePort[1]
                                    update_processing_state = False

                        if method == 'read':
                            atom = PrototypeDemDataElementClasss(TMP_R_DemDataElementClasss, TMP_R_DemDataElementClassTypeName, TMP_R_DemDataElementClassType, TMP_R_DemDataElementDataSize, TMP_R_DemDataElementUsePort,
                                                                 TMP_R_DemInternalDataElement)
                            self._DemDataElementClasss += [atom]

                    if DEFINITION_REF_TEXT.find('/DemInternalDataElementClass') != -1 and DEFINITION_REF_ATTR_DEST == 'ECUC-PARAM-CONF-CONTAINER-DEF':
                        TMP_R_DemDataElementClassTypeName = SHORT_NAME_TEXT
                        TMP_R_DemDataElementClassType = 'DemInternalDataElementClass'

                        PARAM = DemDataEleClass.getElementsByTagName('PARAMETER-VALUES')[0]
                        NUMERICAL = PARAM.getElementsByTagName('ECUC-NUMERICAL-PARAM-VALUE')[0]
                        NUMERICAL_PARAM_VALUE = NUMERICAL.getElementsByTagName('VALUE')[0]
                        NUMERICAL_PARAM_DEFINITION_REF = NUMERICAL.getElementsByTagName('DEFINITION-REF')[0]
                        NUMERICAL_PARAM_DEFINITION_REF_ATTR_DEST = NUMERICAL_PARAM_DEFINITION_REF.getAttribute('DEST')
                        NUMERICAL_PARAM_VALUE_TEXT = str(NUMERICAL_PARAM_VALUE.firstChild.data)
                        NUMERICAL_PARAM_DEFINITION_REF_TEXT = str(NUMERICAL_PARAM_DEFINITION_REF.firstChild.data)

                        TEXTUAL = PARAM.getElementsByTagName('ECUC-TEXTUAL-PARAM-VALUE')[0]
                        TEXTUAL_PARAM_VALUE = TEXTUAL.getElementsByTagName('VALUE')[0]
                        TEXTUAL_PARAM_DEFINITION_REF = TEXTUAL.getElementsByTagName('DEFINITION-REF')[0]
                        TEXTUAL_PARAM_DEFINITION_REF_ATTR_DEST = TEXTUAL_PARAM_DEFINITION_REF.getAttribute('DEST')
                        TEXTUAL_PARAM_VALUE_TEXT = str(TEXTUAL_PARAM_VALUE.firstChild.data)
                        TEXTUAL_PARAM_DEFINITION_REF_TEXT = str(TEXTUAL_PARAM_DEFINITION_REF.firstChild.data)
                        if NUMERICAL_PARAM_DEFINITION_REF_TEXT.find('/DemDataElementDataSize') != -1 and NUMERICAL_PARAM_DEFINITION_REF_ATTR_DEST == 'ECUC-INTEGER-PARAM-DEF':
                            if method == 'read':
                                TMP_R_DemDataElementDataSize = NUMERICAL_PARAM_VALUE_TEXT
                            if method == 'update' and update_processing_state is True:
                                NUMERICAL_PARAM_VALUE.firstChild.data = int(float(TMP_UAD_DemDataElementDataSize[1]))

                        if TEXTUAL_PARAM_DEFINITION_REF_TEXT.find('/DemInternalDataElement') != -1 and TEXTUAL_PARAM_DEFINITION_REF_ATTR_DEST == 'ECUC-ENUMERATION-PARAM-DEF':
                            if method == 'read':
                                TMP_R_DemInternalDataElement = TEXTUAL_PARAM_VALUE_TEXT
                            if method == 'update' and update_processing_state is True:
                                TEXTUAL_PARAM_DEFINITION_REF.firstChild.data = TMP_UAD_DemInternalDataElement[1]
                                update_processing_state = False

                        if method == 'read':
                            atom = PrototypeDemDataElementClasss(TMP_R_DemDataElementClasss, TMP_R_DemDataElementClassTypeName, TMP_R_DemDataElementClassType, TMP_R_DemDataElementDataSize, TMP_R_DemDataElementUsePort,
                                                                 TMP_R_DemInternalDataElement)
                            self._DemDataElementClasss += [atom]
                if method == 'add' and add_item_exist is False:
                    # ECUC-CONTAINER-VALUE
                    NODE_ADD_ECUC_CONTAINER_VALUE = dom.createElement('ECUC-CONTAINER-VALUE')
                    # ECUC-CONTAINER-VALUE / SHORT-NAME
                    NODE_ADD_SHORT_NAME = dom.createElement('SHORT-NAME')
                    NODE_ADD_SHORT_NAME.appendChild(dom.createTextNode(TMP_UAD_DemDataElementClasss[0]))
                    # ECUC-CONTAINER-VALUE / DEFINITION-REF
                    NODE_ADD_DEFINITION_REF = dom.createElement('DEFINITION-REF')
                    NODE_ADD_DEFINITION_REF.setAttribute('DEST', "ECUC-CHOICE-CONTAINER-DEF")
                    NODE_ADD_DEFINITION_REF.appendChild(dom.createTextNode('/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDataElementClass'))
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS
                    NODE_ADD_SUB_CONTAINERS = dom.createElement('SUB-CONTAINERS')
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS /
                    NODE_ADD_SUB_CONTAINERS_CONTAINER = dom.createElement('ECUC-CONTAINER-VALUE')
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / SHORT-NAME
                    NODE_ADD_SUB_CONTAINERS_SHORT_NAME = dom.createElement('SHORT-NAME')
                    NODE_ADD_SUB_CONTAINERS_SHORT_NAME.appendChild(dom.createTextNode(TMP_UAD_DemDataElementClassTypeName[0]))
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / DEFINITION-REF
                    NODE_ADD_SUB_CONTAINERS_DEFINITION_REF = dom.createElement('DEFINITION-REF')
                    NODE_ADD_SUB_CONTAINERS_DEFINITION_REF.setAttribute('DEST', "ECUC-PARAM-CONF-CONTAINER-DEF")
                    NODE_ADD_SUB_CONTAINERS_DEFINITION_REF.appendChild(dom.createTextNode(f'/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDataElementClass/{TMP_UAD_DemDataElementClassType[0]}'))
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / PARAMETER-VALUES
                    NODE_ADD_SUB_CONTAINERS_PARAMETER_VALUES = dom.createElement('PARAMETER-VALUES')
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / PARAMETER-VALUES / ECUC-NUMERICAL-PARAM-VALUE size
                    NODE_ADD_NUMERICAL_Size = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    NODE_ADD_NUMERICAL_Size_DEFINITION_REF = dom.createElement('DEFINITION-REF')
                    NODE_ADD_NUMERICAL_Size_DEFINITION_REF.setAttribute('DEST', "ECUC-INTEGER-PARAM-DEF")
                    NODE_ADD_NUMERICAL_Size_DEFINITION_REF.appendChild(dom.createTextNode(f'/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDataElementClass/{TMP_UAD_DemDataElementClassType[0]}/DemDataElementDataSize'))
                    NODE_ADD_NUMERICAL_Size_VALUE = dom.createElement('VALUE')
                    NODE_ADD_NUMERICAL_Size_VALUE.appendChild(dom.createTextNode(str(int(float(TMP_UAD_DemDataElementDataSize[0])))))
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / PARAMETER-VALUES / ECUC-NUMERICAL-PARAM-VALUE port
                    NODE_ADD_NUMERICAL_port = dom.createElement('ECUC-NUMERICAL-PARAM-VALUE')
                    NODE_ADD_NUMERICAL_port_DEFINITION_REF = dom.createElement('DEFINITION-REF')
                    NODE_ADD_NUMERICAL_port_DEFINITION_REF.setAttribute('DEST', "ECUC-BOOLEAN-PARAM-DEF")
                    NODE_ADD_NUMERICAL_port_DEFINITION_REF.appendChild(dom.createTextNode(f'/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDataElementClass/{TMP_UAD_DemDataElementClassType[0]}/DemDataElementUsePort'))
                    NODE_ADD_NUMERICAL_port_VALUE = dom.createElement('VALUE')
                    NODE_ADD_NUMERICAL_port_VALUE.appendChild(dom.createTextNode(TMP_UAD_DemDataElementUsePort[0]))
                    # ECUC-CONTAINER-VALUE / SUB-CONTAINERS / PARAMETER-VALUES / ECUC-TEXTUAL-PARAM-VALUE
                    NODE_ADD_TEXTUAL = dom.createElement('ECUC-TEXTUAL-PARAM-VALUE')
                    NODE_ADD_TEXTUAL_DEFINITION_REF = dom.createElement('DEFINITION-REF')
                    NODE_ADD_TEXTUAL_DEFINITION_REF.setAttribute('DEST', "ECUC-ENUMERATION-PARAM-DEF")
                    NODE_ADD_TEXTUAL_DEFINITION_REF.appendChild(dom.createTextNode(f'/AUTOSAR_Dem/EcucModuleDefs/Dem/DemGeneral/DemDataElementClass/{TMP_UAD_DemDataElementClassType[0]}/DemInternalDataElement'))
                    NODE_ADD_TEXTUAL_VALUE = dom.createElement('VALUE')
                    NODE_ADD_TEXTUAL_VALUE.appendChild(dom.createTextNode(TMP_UAD_DemInternalDataElement[0]))

                    tmp_node_Internal = [
                        (NODE_ADD_NUMERICAL_Size, [NODE_ADD_NUMERICAL_Size_DEFINITION_REF, NODE_ADD_NUMERICAL_Size_VALUE]),
                        (NODE_ADD_TEXTUAL, [NODE_ADD_TEXTUAL_DEFINITION_REF, NODE_ADD_TEXTUAL_VALUE]),
                        (NODE_ADD_SUB_CONTAINERS_PARAMETER_VALUES, [NODE_ADD_TEXTUAL, NODE_ADD_NUMERICAL_Size]),
                        (NODE_ADD_SUB_CONTAINERS_CONTAINER, [NODE_ADD_SUB_CONTAINERS_SHORT_NAME, NODE_ADD_SUB_CONTAINERS_DEFINITION_REF, NODE_ADD_SUB_CONTAINERS_PARAMETER_VALUES]),
                        (NODE_ADD_SUB_CONTAINERS, [NODE_ADD_SUB_CONTAINERS_CONTAINER]),
                        (NODE_ADD_ECUC_CONTAINER_VALUE, [NODE_ADD_SHORT_NAME, NODE_ADD_DEFINITION_REF, NODE_ADD_SUB_CONTAINERS]),
                        (target, [NODE_ADD_ECUC_CONTAINER_VALUE]),
                    ]

                    tmp_node_ExternalCS = [
                        (NODE_ADD_NUMERICAL_Size, [NODE_ADD_NUMERICAL_Size_DEFINITION_REF, NODE_ADD_NUMERICAL_Size_VALUE]),
                        (NODE_ADD_NUMERICAL_port, [NODE_ADD_NUMERICAL_port_DEFINITION_REF, NODE_ADD_NUMERICAL_port_VALUE]),
                        (NODE_ADD_SUB_CONTAINERS_PARAMETER_VALUES, [NODE_ADD_NUMERICAL_Size, NODE_ADD_NUMERICAL_port]),
                        (NODE_ADD_SUB_CONTAINERS_CONTAINER, [NODE_ADD_SUB_CONTAINERS_SHORT_NAME, NODE_ADD_SUB_CONTAINERS_DEFINITION_REF, NODE_ADD_SUB_CONTAINERS_PARAMETER_VALUES]),
                        (NODE_ADD_SUB_CONTAINERS, [NODE_ADD_SUB_CONTAINERS_CONTAINER]),
                        (NODE_ADD_ECUC_CONTAINER_VALUE, [NODE_ADD_SHORT_NAME, NODE_ADD_DEFINITION_REF, NODE_ADD_SUB_CONTAINERS]),
                        (target, [NODE_ADD_ECUC_CONTAINER_VALUE]),
                    ]
                    if TMP_UAD_DemDataElementClassType[0] == 'DemExternalCSDataElementClass':
                        for item in tmp_node_ExternalCS:
                            for each in item[1]:
                                item[0].appendChild(each)
                    elif TMP_UAD_DemDataElementClassType[0] == 'DemInternalDataElementClass':
                        for item in tmp_node_Internal:
                            for each in item[1]:
                                item[0].appendChild(each)
        return self._DemDataElementClasss

    def processing_arxml2xlsx(self, arxml_path, sheetName, fn_cbk, RE_SN: str = "RE_FltM_Main_10ms"):
        if sheetName == 'DemDataElementClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemDataElementClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemDataElementClasss', entity.get_itemName),
                        ('D', 'DemDataElementClassType', entity.get_itemType),
                        ('E', 'DemDataElementClassTypeName', entity.get_itemTypeName),
                        ('F', 'DemDataElementDataSize', entity.get_itemSize),
                        ('G', 'DemDataElementUsePort', entity.get_itemPort),
                        ('H', 'DemInternalDataElement', entity.get_itemInternal)
                    ]
                return rtn_back

        if sheetName == 'DemDidClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemDidClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemDidClass', entity.get_itemName),
                        ('D', 'DemDidIdentifier*', entity.get_itemID),
                        ('E', 'DemDidDataElementClassRef', entity.get_itemRef)
                    ]
                return rtn_back

        if sheetName == 'DemDTCs':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemDTCs):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemDTC', entity.get_itemDtc),
                        ('E', 'DemDTCSeverity', entity.get_itemDtcSeverity),
                        ('F', 'DemDtcValue', entity.get_itemDtcId),
                        ('I', 'DemDTCAttributesRef*', entity.get_itemDtrAttrRef)
                    ]
                return rtn_back

        if sheetName == 'DemDTCAttributess':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemDTCAttributess):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemDTCAttributes', entity.get_itemDTCAttributes),
                        ('D', 'DemAgingAllowed*', entity.get_itemAgingAllowed),
                        ('E', 'DemAgingCycleCounterThreshold', entity.get_itemAgingCycleCounterThreshold),
                        ('G', 'DemDTCPriority*', entity.get_itemDTCPriority),
                        ('H', 'DemDTCSignificance', entity.get_itemDTCSignificance),
                        ('J', 'DemImmediateNvStorage*', entity.get_itemImmediateNvStorage),
                        ('K', 'DemMaxNumberFreezeFrameRecords', entity.get_itemMaxNumberFreezeFrameRecords),
                        ('L', 'DemAgingCycleRef*', entity.get_itemAgingCycleRef),
                        ('M', 'DemExtendedDataClassRef', entity.get_itemExtendedDataClassRef),
                        ('N', 'DemFreezeFrameClassRef', entity.get_itemFreezeFrameClassRef),
                        ('O', 'DemFreezeFrameRecNumClassRef', entity.get_itemFreezeFrameRecNumClassRef),
                        ('S', 'DemMemoryDestinationRef', entity.get_itemMemoryDestinationRef)
                    ]
                return rtn_back

        if sheetName == 'DemEnableConditions':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemEnableConditions):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemEnableCondition', entity.get_itemEnaCondition),
                        ('D', 'DemEnableConditionStatus*', entity.get_itemConditionStatus)
                    ]
                return rtn_back

        if sheetName == 'DemEnableConditionGroups':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemEnableConditionGroups):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemEnableConditionGroup', entity.get_itemEnableConditionGroup),
                        ('D', 'DemEnableConditionRef', entity.get_itemEnableConditionRef)
                    ]
                return rtn_back

        if sheetName == 'DemFreezeFrameRecNumClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemFreezeFrameRecNumClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemFreezeFrameRecNumClass', entity.get_itemDemFreezeFrameRecNumClass),
                        ('D', 'DemFreezeFrameRecordClassRef', entity.get_itemDemFreezeFrameRecordClassRef)
                    ]
                return rtn_back

        if sheetName == 'DemFreezeFrameRecordClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemFreezeFrameRecordClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemFreezeFrameRecordClass', entity.get_itemDemFreezeFrameRecordClass),
                        ('D', 'DemFreezeFrameRecordNumber*', entity.get_itemDemFreezeFrameRecordNumber),
                        ('E', 'DemFreezeFrameRecordTrigger*', entity.get_itemDemFreezeFrameRecordTrigger),
                        ('F', 'DemFreezeFrameRecordUpdate*', entity.get_itemDemFreezeFrameRecordUpdate)
                    ]
                return rtn_back

        if sheetName == 'DemFreezeFrameClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemFreezeFrameClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemFreezeFrameClass', entity.get_itemDemFreezeFrameClass),
                        ('D', 'DemDidClassRef', entity.get_itemDemDidClassRef)
                    ]
                return rtn_back

        if sheetName == 'DemExtendedDataClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemExtendedDataClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemExtendedDataClass', entity.get_itemDemExtendedDataClass),
                        ('D', 'DemExtendedDataRecordClassRef', entity.get_itemDemExtendedDataRecordClassRef)
                    ]
                return rtn_back

        if sheetName == 'DemExtendedDataRecordClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemExtendedDataRecordClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemExtendedDataRecordClass', entity.get_itemDemExtendedDataRecordClass),
                        ('D', 'DemExtendedDataRecordNumber*', entity.get_itemDemExtendedDataRecordNumber),
                        ('E', 'DemExtendedDataRecordTrigger*', entity.get_itemDemExtendedDataRecordTrigger),
                        ('F', 'DemExtendedDataRecordUpdate*', entity.get_itemDemExtendedDataRecordUpdate),
                        ('G', 'DemDataElementClassRef', entity.get_itemDemDataElementClassRef)
                    ]
                return rtn_back

        if sheetName == 'DemEventParameters':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemEventParameters):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemEventParameter', entity.get_itemDemEventParameter),
                        ('E', 'DemEventAvailable*', entity.get_itemDemEventAvailable),
                        ('F', 'DemEventFailureCycleCounterThreshold*', entity.get_itemDemEventFailureCycleCounterThreshold),
                        ('G', 'DemEventKind*', entity.get_itemDemEventKind),
                        ('H', 'DemFFPrestorageSupported*', entity.get_itemDemFFPrestorageSupported),
                        ('M', 'DemDTCRef', entity.get_itemDemDTCRef),
                        ('N', 'DemEnableConditionGroupRef', entity.get_itemDemEnableConditionGroupRef),
                        ('P', 'DemOperationCycleRef*', entity.get_itemDemOperationCycleRef),
                        ('R', 'DemDebounceCounterBasedClassRef', entity.get_itemDemDebounceCounterBasedClassRef)
                    ]
                return rtn_back

        if sheetName == 'DemDebounceCounterBasedClasss':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeDemDebounceCounterBasedClasss):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'DemDebounceCounterBasedClass', entity.get_itemDemDebounceCounterBasedClass),
                        ('D', 'DemDebounceBehavior*', entity.get_itemDemDebounceBehavior),
                        ('E', 'DemDebounceCounterDecrementStepSize*', entity.get_itemDemDebounceCounterDecrementStepSize),
                        ('F', 'DemDebounceCounterFailedThreshold', entity.get_itemDemDebounceCounterFailedThreshold),
                        ('G', 'DemDebounceCounterIncrementStepSize*', entity.get_itemDemDebounceCounterIncrementStepSize),
                        ('H', 'DemDebounceCounterJumpDown', entity.get_itemDemDebounceCounterJumpDown),
                        ('I', 'DemDebounceCounterJumpDownValue', entity.get_itemDemDebounceCounterJumpDownValue),
                        ('J', 'DemDebounceCounterJumpUp', entity.get_itemDemDebounceCounterJumpUp),
                        ('K', 'DemDebounceCounterJumpUpValue', entity.get_itemDemDebounceCounterJumpUpValue),
                        ('L', 'DemDebounceCounterPassedThreshold', entity.get_itemDemDebounceCounterPassedThreshold),
                        ('M', 'DemDebounceCounterStorage*', entity.get_itemDemDebounceCounterStorage),
                        ('O', 'DemRbDebounceCounterFdcThresholdStorageValue', entity.get_itemDemRbDebounceCounterFdcThresholdStorageValue),
                        ('Q', 'DemRbDebounceCounterJumpDownAlternative', entity.get_itemDemRbDebounceCounterJumpDownAlternative),
                        ('S', 'DemRbDebounceCounterJumpUpAlternative', entity.get_itemDemRbDebounceCounterJumpUpAlternative)
                    ]
                return rtn_back

        if sheetName == 'FltMPRPort':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeFltMPRPort):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'Type', entity.get_itemPortType),
                        ('D', 'Short Name', entity.get_itemShortName),
                        ('E', 'Interface', entity.get_itemInterface),
                        ('F', 'Port InterfacePath', entity.get_itemPortInterfacePath),
                        ('G', 'Port Interface', entity.get_itemPortInterface),
                        ('H', 'Is Service', entity.get_itemIsService),
                    ]
                return rtn_back

        if sheetName == 'FltMRE':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeFltMRE):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'Runnable Entity', entity.get_itemRunnableEntity),
                        ('D', 'Function Name', entity.get_itemFunctionName),
                    ]
                return rtn_back

        if sheetName == 'FltMEvents':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeFltMEvents):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'Rte Event Type', entity.get_itemRteEventType),
                        ('D', 'Event Name', entity.get_itemEventName),
                        ('E', 'Start Runnable Entity', entity.get_itemStartRunnableEntity),
                        ('F', 'Timing Period(s)', entity.get_itemTimingPeriod),
                        ('H', 'Target', entity.get_itemEventTarget),
                        ('I', 'PortPath', entity.get_itemEventPortPath),
                        ('J', 'Port', entity.get_itemEventPort)
                    ]
                return rtn_back

        if sheetName == 'FltMREServerCallPoint':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeFltMREServerCallPoint):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'Server Call Point', entity.get_itemServerCallPoint),
                        ('D', 'Short Name', entity.get_itemShortName),
                        ('E', 'OperationPath', entity.get_itemOperationPath),
                        ('F', 'Operation', entity.get_itemOperation),
                        ('G', 'Client Port', entity.get_itemClientPort)
                    ]
                return rtn_back

        if sheetName == 'Composition':
            def a2x(entity, index_num):
                rtn_back = list()
                if isinstance(entity, PrototypeComposition):
                    rtn_back = [
                        ('A', 'index', index_num + 1),
                        ('C', 'Component Prototype', entity.get_itemComponentPrototype),
                        ('D', 'Connector Name', entity.get_itemConnectorName),
                        ('E', 'PortPath', entity.get_itemPortPath),
                        ('F', 'Port', entity.get_itemPort),
                        ('G', 'Connected Composition', entity.get_itemConnectedComposition),
                        ('H', 'Connected PortPath', entity.get_itemConnectedPortPath),
                        ('I', 'Connected Port', entity.get_itemConnectedPort)
                    ]
                return rtn_back

        try:
            dom = xml.dom.minidom.parse(arxml_path)
            minidom.Element.writexml = self.fixed_writexml
            wb = openpyxl.load_workbook(self._xlsxPath)
            sheet = wb[sheetName]
        except OSError:
            print("Invalid file path specified.")
            print(arxml_path)
            print(self._xlsxPath)
            return
        root = dom.documentElement
        for row in sheet.iter_rows(min_row=4):
            for cell in row:
                cell.value = ''
        # 组合数据为DF
        if sheetName == 'FltMREServerCallPoint':
            rtn = fn_cbk(root, dom, RE_SN=RE_SN)
        else:
            rtn = fn_cbk(root, dom)
        dynamicCreateListDone = False  # 确定动态创建list是否结束,仅第一遍循环执行,格式如list = [list(),list(),...]
        dynamicCreateList = []  # 容器组合 list = [list(),list(),...]
        rtn_a2x = list()  # 提升等级为局部变量
        # 动态创建list,动态存储数据
        for each in rtn:
            rtn_a2x = a2x(each, rtn.index(each))
            index = 0
            for row in rtn_a2x:
                if dynamicCreateListDone is False:
                    dynamicCreateList.append(list())
                dynamicCreateList[index].append(row[2])
                index += 1
            dynamicCreateListDone = True

        dynamicCreateDict_0 = {}
        for item in rtn_a2x:
            dynamicCreateDict_0[item[1]] = dynamicCreateList[rtn_a2x.index(item)]
        dynamicCreateDF = pd.DataFrame(dynamicCreateDict_0)

        # 输出DF到xlsx
        for index, row in dynamicCreateDF.iterrows():
            index += 4
            for item in rtn_a2x:
                sheet[f'{item[0]}{str(index)}'].value = row[f'{item[1]}']
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_xlsx2arxml(self, arxml_path, sheetName, fn_cbk, RE_SN: str = "RE_FltM_Main_10ms", cp_reindex: bool = False):
        try:
            dom = xml.dom.minidom.parse(arxml_path)
            minidom.Element.writexml = self.fixed_writexml
        except OSError:
            print("Invalid file path specified.")
            print(arxml_path)
            return
        root = dom.documentElement

        df = pd.read_excel(io=self._xlsxPath, sheet_name=sheetName, header=0, index_col=None)
        df = df.dropna(axis=0, how='all')

        if cp_reindex is True:
            # 仅parser_FltMREServerCallPoint可以使用,否则会出错
            fn_cbk(root, dom, method='', RE_SN=RE_SN, auto_reindex=True)

        for index, row in df.iterrows():
            if str(row[1]).find('u') != -1:
                fn_cbk(root, dom, row, 'update')
            if str(row[1]).find('d') != -1:
                fn_cbk(root, dom, row, 'remove')
            # 有一个BUG就是，DF为空的时候是NaN，被识别认为a
            if str(row[1]).find('add') != -1 and str(row[1]) != "nan":
                fn_cbk(root, dom, row, 'add')

        # 输出文件
        fh = codecs.open(arxml_path, 'w', 'UTF-8')
        dom.writexml(fh, addindent='  ', newl='\n', encoding='UTF-8')
        fh.close()

    def processing_EditDem2ws(self, tgt_sheetName: str, task: str):
        try:
            df = pd.read_excel(io=self._xlsxPath, sheet_name="EditDem", header=0, index_col=None)
            df = df.dropna(axis=0, how='all')
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_tgt = wb[tgt_sheetName]
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return
        for row in ws_tgt.iter_rows(min_row=4):
            for cell in row:
                cell.value = ''
        # 输出DF到xlsx
        links = []
        for index, row in df.iterrows():
            index += 2
            if task == "DemDTCs":
                links = [
                    # (col, col_name, prefix, value)
                    ("A", "目录", "", str(index - 3)),
                    ("C", "DemDTC", "DemDTC_DTC_", row["DID"]),
                    ("E", "DemDTCSeverity", "", "DEM_SEVERITY_CHECK_IMMEDIATELY"),
                    ("F", "DemDtcValue", "0x", row["DID"]),
                    ("I", "DemDTCAttributesRef*", "DemDTCAttributes_DTC_", row["DID"])
                ]
            if task == "DemDTCAttributess":
                if row["DtcType"] == "ExtDTC":
                    val_DemEDCPrefix = "DemEDC_E_"
                    val_DemEDCRef = "FFFFFF"
                    val_FFCPrefix = "DemFFC_"
                    val_FFCRef = "20"
                    val_FFRNCPrefix = "DemFFRNC_"
                    val_FFRNCRef = "20_21"
                    val_memDestRef = "DemPrimaryMemory"
                elif row["DtcType"] == "InnDTC":
                    val_DemEDCPrefix = "DemEDC_I_"
                    val_DemEDCRef = row["DID"]
                    val_FFCPrefix = ""
                    val_FFCRef = ""
                    val_FFRNCPrefix = ""
                    val_FFRNCRef = ""
                    val_memDestRef = "DemSecondaryMemory"
                else:
                    val_DemEDCPrefix = ""
                    val_DemEDCRef = ""
                    val_FFCPrefix = ""
                    val_FFCRef = ""
                    val_FFRNCPrefix = ""
                    val_FFRNCRef = ""
                    val_memDestRef = ''
                links = [
                    # (col, col_name, prefix, value)
                    ("A", "目录", '', index - 3),
                    ("C", "DemDTCAttributes", "DemDTCAttributes_DTC_", row["DID"]),
                    ("D", "DemAgingAllowed*", "", "true"),
                    ("E", "DemAgingCycleCounterThreshold", "", row["DemAgingCycleCounterThreshold"]),
                    ("G", "DemDTCPriority*", '', 1),
                    ("H", "DemDTCSignificance", "", "DEM_EVENT_SIGNIFICANCE_FAULT"),
                    ("J", "DemImmediateNvStorage*", "", "true"),
                    ("K", "DemMaxNumberFreezeFrameRecords", '', 2),
                    ("L", "DemAgingCycleRef*", "", "DemOperationCycle_0"),
                    ("M", "DemExtendedDataClassRef", val_DemEDCPrefix, val_DemEDCRef),  # 待定
                    ("N", "DemFreezeFrameClassRef", val_FFCPrefix, val_FFCRef),  # 待定
                    ("O", "DemFreezeFrameRecNumClassRef", val_FFRNCPrefix, val_FFRNCRef),  # 待定
                    ("S", "DemMemoryDestinationRef", "", val_memDestRef)
                ]
            for item in links:
                ws_tgt[f'{item[0]}{str(index)}'].value = f"{item[2]}{item[3]}"
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_EditDem2DemEventParameters(self):
        """
        Method: 修改,如果DemEventParameters没有对应的EventParam则无法修改
        从 EditDem 读取 EventName 到 DemEventParameters 中,判断标准则为 EventName , 然后拼装出 前缀 + EventID + EventName
        :return:
        """
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_src = wb['EditDem']
            ws_tgt = wb['DemEventParameters']
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return
        # 遍历DemEventParameters表格C列
        # # 1.0 获取表格所在范围
        # print(ws_src.dimensions)
        # # 2.0 获取某个单元格的具体内容
        # print(ws_src['A1'].value)
        # print(ws_src.cell(row=2, column=3).value)
        # # 3.0 获取某个单元格的行列坐标
        # cell = ws_src.cell(row=2, column=3)
        # print(cell.row, cell.column, cell.coordinate)
        # # 4.0 获取多个格子的值
        # # 指定坐标范围的值
        # print(ws_src['A1:B5'])
        # # 指定列的值
        # print(ws_src['A'])
        # print(ws_src['A:C'])
        # # 指定行的值
        # print(ws_src[5])
        # print(ws_src[6])
        # # 打印上述值
        # for cell in ws_src[5]:
        #     print(cell.value)
        for rowTgt in ws_tgt.iter_rows(min_row=4, min_col=3, max_col=3):
            for cellTgt in rowTgt:
                # 遍历EditDem表格C列
                for rowSrc in ws_src.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cellSrc in rowSrc:
                        if str(ws_tgt[cellTgt.coordinate].value).removeprefix("DemEventParameter_") == ws_src[cellSrc.coordinate].value:  # 判断两个表格EventParam名称是否一致
                            column_d_cell = ws_src.cell(row=cellSrc.row, column=cellSrc.column + 1).value  # 获取ID列的信息
                            cellTgt.value = f"{cellTgt.value}/DemEventParameter{column_d_cell}_{str(cellSrc.value).removeprefix('v01_')}"
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_h2EditDem(self):
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            sheet = wb['EditDem']
            for row in sheet.iter_rows(min_row=4, min_col=1, max_col=5):
                for cell in row:
                    cell.value = ''
            file_h_event = open(self._eventId_path, 'r', encoding='UTF-8')
            file_h_dtc = open(self._dtcId_path, 'r', encoding='UTF-8')
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            print(self._eventId_path)
            print(self._dtcId_path)
            return
        # 组合xlsx表格layout
        eventID_list = [
            ('A', 'index', list()),
            ('C', 'EventName', list()),
            ('D', 'EventID', list()),
            ('E', 'DtcType', list()),
            ('F', 'DID', list())
        ]
        # 从Dem_Cfg_EventId.h文件读取EventID到xlsx
        event_param_index = 0
        for line in file_h_event:
            line = line.rstrip('\n')
            if line.startswith("#define DemConf_DemEventParameter_DemEventParameter_"):
                event_param_index += 1
                items = line.split()
                eventID_list[0][2].append(event_param_index)
                eventID_list[1][2].append(items[1].removeprefix('DemConf_DemEventParameter_DemEventParameter_'))
                eventID_list[2][2].append(f"_{int(items[2].removesuffix('u')):03d}")
        file_h_event.close()
        # 从Dem_Cfg_DtcId.h文件读取DtcID到xlsx
        dtc_param_index = 0
        for line in file_h_dtc:
            line = line.rstrip('\n')
            if line.startswith("   ,DemConf_DemDTCClass_DemDTC_") or line.startswith("   ,DEM_DTCID_INVALID"):
                items = line.split()
                # 判断EventID跟从Dem_Cfg_EventId中读出来的是否一致,如果一致则将DTC ID赋值给对应的list
                if eventID_list[1][2][dtc_param_index] == items[2].removeprefix('DemEventParameter_'):
                    if items[0].removeprefix(',DemConf_DemDTCClass_DemDTC_DTC_').replace(",DEM_DTCID_INVALID", '0').startswith('40'):
                        eventID_list[3][2].append('InnDTC')
                    else:
                        eventID_list[3][2].append('ExtDTC')
                    eventID_list[4][2].append(items[0].removeprefix(',DemConf_DemDTCClass_DemDTC_DTC_').replace(",DEM_DTCID_INVALID", '0'))
                dtc_param_index += 1
        file_h_dtc.close()
        # 组合成DF
        dynamicCreateDict_0 = {}
        for item in eventID_list:
            dynamicCreateDict_0[item[1]] = item[2]
        dynamicCreateDF = pd.DataFrame(dynamicCreateDict_0)
        # 输出DF到xlsx
        for index, row in dynamicCreateDF.iterrows():
            index += 4
            for item in eventID_list:
                sheet[f'{item[0]}{str(index)}'].value = row[f'{item[1]}']
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_DemDTCAttributess2EditDem(self):
        """
        Method: 读取,重写
        从DemDTCAttributess读取DemAgingCycleCounterThreshold到EditDem中,判断标准则为DID,将对应DID的值填入
        :return:
        """
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_src = wb['DemDTCAttributess']
            ws_tgt = wb['EditDem']
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return

        for rowTgt in ws_tgt.iter_rows(min_row=4, min_col=6, max_col=6):
            for cellTgt in rowTgt:
                for rowSrc in ws_src.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cellSrc in rowSrc:
                        # 判断两个表格DID名称是否一致,str不可少,否则纯数字ID可能会出错, [-6:]为了取最后命名的DID
                        if str(ws_tgt[cellTgt.coordinate].value) == str(ws_src[cellSrc.coordinate].value)[-6:]:
                            match_link = [
                                # TgtBase:6(F)           SrcBase: 3(C)
                                # (tgtCol, tgtColName, tgtOffset, srcCol, srcColName, srcOffset)
                                ("I", "DemAgingCycleCounterThreshold", 3, "E", "DemAgingCycleCounterThreshold", 2),
                                ("J", "DemDTCPriority*", 4, "G", "DemDTCPriority*", 4),
                            ]
                            for item in match_link:
                                ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + item[2]).value = ws_src.cell(row=cellSrc.row, column=cellSrc.column + item[5]).value
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_DemEventParameters2EditDem(self):
        """
        Method: 读取,重写
        从DemDTCAttributess读取DemAgingCycleCounterThreshold到EditDem中,判断标准则为DID,将对应DID的值填入
        :return:
        """
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_src = wb['DemEventParameters']
            ws_tgt = wb['EditDem']
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return
        for rowTgt in ws_tgt.iter_rows(min_row=4, min_col=3, max_col=3):
            for cellTgt in rowTgt:
                for rowSrc in ws_src.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cellSrc in rowSrc:
                        # 判断两个表格DID名称是否一致,str不可少,否则纯数字ID可能会出错
                        if str(ws_tgt[cellTgt.coordinate].value) == str(ws_src[cellSrc.coordinate].value).removeprefix("DemEventParameter_"):
                            match_link = [
                                # TgtBase:3(C)           SrcBase: 3(C)
                                # (tgtCol, tgtColName, tgtOffset, srcCol, srcColName, srcOffset)
                                ("K", "DemEventFailureCycleCounterThreshold*", 8, "F", "DemEventFailureCycleCounterThreshold*", 3),
                                ("L", "DemEventKind*", 9, "G", "DemEventKind*", 4),
                                ("M", "DemEnableConditionGroupRef", 10, "N", "DemEnableConditionGroupRef", 11),
                                ("N", "DemDebounceCounterFailedThreshold", 11, "R", "DemDebounceCounterBasedClassRef", 15),
                                ("O", "DemDebounceCounterDecrementStepSize*", 12, "R", "DemDebounceCounterBasedClassRef", 15),
                                ("P", "DemDebounceCounterIncrementStepSize*", 13, "R", "DemDebounceCounterBasedClassRef", 15),
                            ]
                            for item in match_link:
                                ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + item[2]).value = ws_src.cell(row=cellSrc.row, column=cellSrc.column + item[5]).value
                            ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + 11).value = int(str(ws_src.cell(row=cellSrc.row, column=cellSrc.column + 15).value).removeprefix("DemDebounceCounterBasedClass_").split("_")[0])
                            ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + 12).value = int(str(ws_src.cell(row=cellSrc.row, column=cellSrc.column + 15).value).removeprefix("DemDebounceCounterBasedClass_").split("_")[1])
                            ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + 13).value = int(str(ws_src.cell(row=cellSrc.row, column=cellSrc.column + 15).value).removeprefix("DemDebounceCounterBasedClass_").split("_")[2])
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_DemExtendedDataRecordClasss2EditDem(self):
        """
        Method: 读取,改写
        从 DemExtendedDataRecordClasss 读取 DemExtendedDataRecordNumber* 到EditDem中,判断标准则为EventName,将对应 EventName 的值填入
        :return:
        """
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_src = wb['DemExtendedDataRecordClasss']
            ws_tgt = wb['EditDem']
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return
        for rowTgt in ws_tgt.iter_rows(min_row=4, min_col=3, max_col=3):
            for cellTgt in rowTgt:
                for rowSrc in ws_src.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cellSrc in rowSrc:
                        # 判断两个表格DID名称是否一致,str不可少,否则纯数字ID可能会出错
                        if str(ws_src[cellSrc.coordinate].value).find(str(ws_tgt[cellTgt.coordinate].value)) != -1:
                            match_link = [
                                # TgtBase:3(C)           SrcBase: 3(C)
                                # (tgtCol, tgtColName, tgtOffset, srcCol, srcColName, srcOffset)
                                ("G", "DemExtDataID", 4, "D", "DemExtendedDataRecordNumber*", 1),
                            ]
                            for item in match_link:
                                ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + item[2]).value = ws_src.cell(row=cellSrc.row, column=cellSrc.column + item[5]).value
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def processing_DemDataElementClasss2EditDem(self):
        """
        Method: 读取,改写
        从 DemDataElementClasss 读取 DemDataElementDataSize 到EditDem中,判断标准则为EventName,将对应 EventName 的值填入
        :return:
        """
        try:
            wb = openpyxl.load_workbook(self._xlsxPath)
            ws_src = wb['DemDataElementClasss']
            ws_tgt = wb['EditDem']
        except OSError:
            print("Invalid file path specified.")
            print(self._xlsxPath)
            return
        for rowTgt in ws_tgt.iter_rows(min_row=4, min_col=3, max_col=3):
            for cellTgt in rowTgt:
                for rowSrc in ws_src.iter_rows(min_row=4, min_col=3, max_col=3):
                    for cellSrc in rowSrc:
                        # 判断两个表格DID名称是否一致,str不可少,否则纯数字ID可能会出错
                        if str(ws_src[cellSrc.coordinate].value).find(str(ws_tgt[cellTgt.coordinate].value)) != -1:
                            match_link = [
                                # TgtBase:3(C)           SrcBase: 3(C)
                                # (tgtCol, tgtColName, tgtOffset, srcCol, srcColName, srcOffset)
                                ("H", "DemExtDataLen", 5, "F", "DemExtendedDataRecordNumber*", 3),
                            ]
                            for item in match_link:
                                ws_tgt.cell(row=cellTgt.row, column=cellTgt.column + item[2]).value = ws_src.cell(row=cellSrc.row, column=cellSrc.column + item[5]).value
        if self.check_xlsx_is_open() is True:
            wb.save(self._xlsxPath)

    def loading(self, task: str = '3'):
        if task == "1":
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemDataElementClasss', fn_cbk=self.parser_DemDataElementClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemDidClasss', fn_cbk=self.parser_DemDidClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemDTCs', fn_cbk=self.parser_DemDTCs)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemDTCAttributess', fn_cbk=self.parser_DemDTCAttributess)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemEnableConditions', fn_cbk=self.parser_DemEnableConditions)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemEnableConditionGroups', fn_cbk=self.parser_DemEnableConditionGroups)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameRecNumClasss', fn_cbk=self.parser_DemFreezeFrameRecNumClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameRecordClasss', fn_cbk=self.parser_DemFreezeFrameRecordClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameClasss', fn_cbk=self.parser_DemFreezeFrameClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemExtendedDataClasss', fn_cbk=self.parser_DemExtendedDataClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemExtendedDataRecordClasss', fn_cbk=self.parser_DemExtendedDataRecordClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemEventParameters', fn_cbk=self.parser_DemEventParameters)
            self.processing_arxml2xlsx(arxml_path=self._arxmlDemPath, sheetName='DemDebounceCounterBasedClasss', fn_cbk=self.parser_DemDebounceCounterBasedClasss)
            self.processing_arxml2xlsx(arxml_path=self._arxmlFltMPath, sheetName='FltMPRPort', fn_cbk=self.parser_FltMPRPort)
            self.processing_arxml2xlsx(arxml_path=self._arxmlFltMPath, sheetName='FltMRE', fn_cbk=self.parser_FltMRE)
            self.processing_arxml2xlsx(arxml_path=self._arxmlFltMPath, sheetName='FltMEvents', fn_cbk=self.parser_FltMEvents)
            self.processing_arxml2xlsx(arxml_path=self._arxmlFltMPath, sheetName='FltMREServerCallPoint', fn_cbk=self.parser_FltMREServerCallPoint, RE_SN="RE_FltM_Main_10ms")
            self.processing_arxml2xlsx(arxml_path=self._arxmlCompositionPath, sheetName='Composition', fn_cbk=self.parser_Composition)
        elif task == '2':
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemDataElementClasss', fn_cbk=self.parser_DemDataElementClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemDidClasss', fn_cbk=self.parser_DemDidClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemDTCs', fn_cbk=self.parser_DemDTCs)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemDTCAttributess', fn_cbk=self.parser_DemDTCAttributess)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemEnableConditions', fn_cbk=self.parser_DemEnableConditions)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemEnableConditionGroups', fn_cbk=self.parser_DemEnableConditionGroups)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameRecNumClasss', fn_cbk=self.parser_DemFreezeFrameRecNumClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameRecordClasss', fn_cbk=self.parser_DemFreezeFrameRecordClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemFreezeFrameClasss', fn_cbk=self.parser_DemFreezeFrameClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemExtendedDataClasss', fn_cbk=self.parser_DemExtendedDataClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemExtendedDataRecordClasss', fn_cbk=self.parser_DemExtendedDataRecordClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemEventParameters', fn_cbk=self.parser_DemEventParameters)
            self.processing_xlsx2arxml(arxml_path=self._arxmlDemPath, sheetName='DemDebounceCounterBasedClasss', fn_cbk=self.parser_DemDebounceCounterBasedClasss)
            self.processing_xlsx2arxml(arxml_path=self._arxmlFltMPath, sheetName='FltMPRPort', fn_cbk=self.parser_FltMPRPort)
            self.processing_xlsx2arxml(arxml_path=self._arxmlFltMPath, sheetName='FltMRE', fn_cbk=self.parser_FltMRE)
            self.processing_xlsx2arxml(arxml_path=self._arxmlFltMPath, sheetName='FltMEvents', fn_cbk=self.parser_FltMEvents)
            self.processing_xlsx2arxml(arxml_path=self._arxmlFltMPath, sheetName='FltMREServerCallPoint', fn_cbk=self.parser_FltMREServerCallPoint, cp_reindex=True, RE_SN="RE_FltM_Main_10ms")
            self.processing_xlsx2arxml(arxml_path=self._arxmlCompositionPath, sheetName='Composition', fn_cbk=self.parser_Composition)
        elif task == '3':
            self.processing_h2EditDem()
            self.processing_DemDTCAttributess2EditDem()
            self.processing_DemEventParameters2EditDem()
            self.processing_DemExtendedDataRecordClasss2EditDem()
            self.processing_DemDataElementClasss2EditDem()
        elif task == '4':
            self.processing_EditDem2DemEventParameters()
            self.processing_EditDem2ws(tgt_sheetName="DemDTCs", task="DemDTCs")
            self.processing_EditDem2ws(tgt_sheetName="DemDTCAttributess", task="DemDTCAttributess")


__version__ = "20230309.0.0.1"
app = typer.Typer()


@app.command()
def status():
    print("==========================================================================")
    print("[bold green]Fast Dem Console[/bold green](ADC20 TC397) ")
    print("==========================================================================")
    print("Num       Status")
    print("--------------------------------------------------------------------------")
    print(" 0        Exit this exe loop! ")
    print(" x        Print detail of status")
    print("--------------------------------------------------------------------------")
    dtc_status = "0xFF"
    while dtc_status != "0":
        dtc_status = typer.prompt(f"({time.ctime()})Input Dtc.Status to run")
        print("\n\n")
        DTCSTATUS = ISO14229DTCSTATUS()
        DTCSTATUS.set_status_code(dtc_status)
    else:
        print("Terminal Task Done!!!")

@app.command()
def config():
    # 工具预期放在Build中,下面将会按照./Build为base取相对路径
    if getattr(sys, 'frozen', False):
        # 获取应用程序exe的路径
        path = os.path.dirname(sys.executable)
    elif __file__:
        # 获取脚本程序的路径
        path = os.path.dirname(__file__)
    # path = f"{path}\\fastdem.ini"
    path = f".\\fastdem.ini"
    print("==========================================================================")
    print("[bold green]Fast Dem Console[/bold green](ADC20 TC397) ")
    print("==========================================================================")
    print("Num       Task")
    print("--------------------------------------------------------------------------")
    print(" 0        Exit this exe loop! ")
    print(" 1        Config the base path")
    print(" 2        Cat the base path")
    print("--------------------------------------------------------------------------")
    task = "1"
    while task != "0":
        task = typer.prompt(f"({time.ctime()})Input Task.Num to run")
        if task == "1":
            PATH_ARXML_DEM_BSW = typer.prompt("Setting PATH_ARXML_DEM_BSW", default=r"../ConfigEnv/IsolarConf/ecu_config/bsw/static/Dia/Dem_EcucValues.arxml", show_default=True)
            PATH_ARXML_DEM_SWC = typer.prompt("Setting PATH_ARXML_DEM_SWC", default=r"../ConfigEnv/IsolarConf/swc_config/FltM_SWC/FltM_SWC.arxml", show_default=True)
            PATH_ARXML_DEM_RTE = typer.prompt("Setting PATH_ARXML_DEM_RTE", default=r"../ConfigEnv/IsolarConf/swc_config/Composition.arxml", show_default=True)
            PATH_XLSX_DEM_ETAS = typer.prompt("Setting PATH_XLSX_DEM_ETAS", default=r"./fastdem.xlsx", show_default=True)
            PATH_H_DEM_EventID = typer.prompt("Setting PATH_H_DEM_EventID", default=r"../Bsw/Dem/Dem_Cfg_EventId.h", show_default=True)
            PATH_H_DEM_DtcID = typer.prompt("Setting PATH_H_DEM_DtcID  ", default=r"../Bsw/Dem/Dem_Cfg_DtcId.h", show_default=True)
            with open(path, "w", encoding='utf-8') as f:
                f.writelines(f'PATH_ARXML_DEM_BSW : {PATH_ARXML_DEM_BSW}\n')
                f.writelines(f'PATH_ARXML_DEM_SWC : {PATH_ARXML_DEM_SWC}\n')
                f.writelines(f'PATH_ARXML_DEM_RTE : {PATH_ARXML_DEM_RTE}\n')
                f.writelines(f'PATH_XLSX_DEM_ETAS : {PATH_XLSX_DEM_ETAS}\n')
                f.writelines(f'PATH_H_DEM_EventID : {PATH_H_DEM_EventID}\n')
                f.writelines(f'PATH_H_DEM_DtcID   : {PATH_H_DEM_DtcID}\n')
            f.close()
            print(f"The config file create at >> [green]{path}[/green]")
            print("Input [red]0[/red] to Exit")
        if task == "2":
            try:
                with open(path, "r", encoding='utf-8') as f:
                    PATH_ARXML_DEM_BSW = ''
                    PATH_ARXML_DEM_SWC = ''
                    PATH_ARXML_DEM_RTE = ''
                    PATH_XLSX_DEM_ETAS = ''
                    PATH_H_DEM_EventID = ''
                    PATH_H_DEM_DtcID = ''
                    for line in f:
                        line = line.replace('\n', '').split(':')
                        line[0] = line[0].strip()
                        line[1] = line[1].strip()
                        if line[0] == "PATH_ARXML_DEM_BSW":
                            PATH_ARXML_DEM_BSW = line[-1]
                        if line[0] == "PATH_ARXML_DEM_SWC":
                            PATH_ARXML_DEM_SWC = line[-1]
                        if line[0] == "PATH_ARXML_DEM_RTE":
                            PATH_ARXML_DEM_RTE = line[-1]
                        if line[0] == "PATH_XLSX_DEM_ETAS":
                            PATH_XLSX_DEM_ETAS = line[-1]
                        if line[0] == "PATH_H_DEM_EventID":
                            PATH_H_DEM_EventID = line[-1]
                        if line[0] == "PATH_H_DEM_DtcID":
                            PATH_H_DEM_DtcID = line[-1]
                f.close()
                print(f'PATH_ARXML_DEM_BSW : {PATH_ARXML_DEM_BSW}')
                print(f'PATH_ARXML_DEM_SWC : {PATH_ARXML_DEM_SWC}')
                print(f'PATH_ARXML_DEM_RTE : {PATH_ARXML_DEM_RTE}')
                print(f'PATH_XLSX_DEM_ETAS : {PATH_XLSX_DEM_ETAS}')
                print(f'PATH_H_DEM_EventID : {PATH_H_DEM_EventID}')
                print(f'PATH_H_DEM_DtcID   : {PATH_H_DEM_DtcID}')
                print("Input [red]0[/red] to Exit")
            except Exception as e:
                print(e)
                return
    else:
        print("Terminal Task Done!!!")


@app.command()
def run():
    # 工具预期放在Build中,下面将会按照./Build为base取相对路径
    if getattr(sys, 'frozen', False):
        # 获取应用程序exe的路径
        path = os.path.dirname(sys.executable)
    elif __file__:
        # 获取脚本程序的路径
        path = os.path.dirname(__file__)
    # path = f"{path}\\fastdem.ini"
    path = f".\\fastdem.ini"
    with open(path, "r", encoding='utf-8') as f:
        PATH_ARXML_DEM_BSW = ''
        PATH_ARXML_DEM_SWC = ''
        PATH_ARXML_DEM_RTE = ''
        PATH_XLSX_DEM_ETAS = ''
        PATH_H_DEM_EventID = ''
        PATH_H_DEM_DtcID = ''
        for line in f:
            line = line.replace('\n', '').split(':')
            line[0] = line[0].strip()
            line[1] = line[1].strip()
            if line[0] == "PATH_ARXML_DEM_BSW":
                PATH_ARXML_DEM_BSW = line[-1]
            if line[0] == "PATH_ARXML_DEM_SWC":
                PATH_ARXML_DEM_SWC = line[-1]
            if line[0] == "PATH_ARXML_DEM_RTE":
                PATH_ARXML_DEM_RTE = line[-1]
            if line[0] == "PATH_XLSX_DEM_ETAS":
                PATH_XLSX_DEM_ETAS = line[-1]
            if line[0] == "PATH_H_DEM_EventID":
                PATH_H_DEM_EventID = line[-1]
            if line[0] == "PATH_H_DEM_DtcID":
                PATH_H_DEM_DtcID = line[-1]
    etas_entity = EtasDem(arxmlDemPath=PATH_ARXML_DEM_BSW,
                          arxmlFltMPath=PATH_ARXML_DEM_SWC,
                          arxmlCompositionPath=PATH_ARXML_DEM_RTE,
                          xlsx_path=PATH_XLSX_DEM_ETAS,
                          eventId_path=PATH_H_DEM_EventID,
                          dtcId_path=PATH_H_DEM_DtcID)
    print("==========================================================================")
    print("[bold green]Fast Dem Console[/bold green](ADC20 TC397) ")
    print("==========================================================================")
    print("Num       Task")
    print("--------------------------------------------------------------------------")
    print(" 0        Exit this exe loop! ")
    print(" 1        read [red]arxml data[/red]  write to: xlsx")
    print(" 2        read [red]xlsx  data[/red]  write to: arxml, support add,u,d ")
    print(" 3        read [red]h file and other sheet data[/red] to Sheet['DemEdit']")
    print(" 4        read Sheet['DemEdit'] data  write to: other Sheet")
    print(" 5        open fastdem.xlsx")
    print(" 6        open fastdem.ini")
    print("--------------------------------------------------------------------------")
    task = "1"
    while task != "0":
        task = typer.prompt(f"({time.ctime()})Input Task.Num to run")
        with Progress(SpinnerColumn(),
                      TextColumn("[progress.description]{task.description}"),
                      transient=True) as progress:
            progress.add_task(description="Processing...", total=None)
            progress.add_task(description="Preparing...", total=None)
            if task == "5":
                try:
                    os.startfile(os.path.abspath(r"./fastdem.xlsx"))
                except Exception as e:
                    print(f"Invalid Path : {os.path.abspath(r'./fastdem.xlsx')}")
            elif task == "6":
                try:
                    os.startfile(os.path.abspath(r"./fastdem.ini"))
                except Exception as e:
                    print(f"Invalid Path : {os.path.abspath(r'./fastdem.ini')}")
            else:
                etas_entity.loading(task=task)
    else:
        print("Terminal Task Done!!!")


def development():
    dev = True
    fact = False
    run_method = dev
    file_1 = 'Dem_EcucValues.arxml'
    file_2 = 'fastdem.xlsx'
    file_3 = 'Dem_Cfg_EventId.h'
    file_4 = 'Dem_Cfg_DtcId.h'
    file_5 = 'FltM_SWC.arxml'
    file_6 = 'Composition.arxml'
    if run_method is True:
        workspace_i1 = f'./example/input/{file_1}'
        workspace_i2 = f'./example/input/{file_2}'
        workspace_i3 = f'./example/input/{file_3}'
        workspace_i4 = f'./example/input/{file_4}'
        workspace_i5 = f'./example/input/{file_5}'
        workspace_i6 = f'./example/input/{file_6}'
        workspace_o1 = f'./example/output/{file_1}'
    else:
        tmp_date = '20221028'
        workspace_i1 = f'./workspace/{tmp_date}/input/{file_1}'
        workspace_i2 = f'./workspace/{tmp_date}/input/{file_2}'
        workspace_i3 = f'./workspace/{tmp_date}/input/{file_3}'
        workspace_i4 = f'./workspace/{tmp_date}/input/{file_4}'
        workspace_i5 = f'./workspace/{tmp_date}/input/{file_5}'
        workspace_i6 = f'./workspace/{tmp_date}/input/{file_6}'
        workspace_o1 = f'./workspace/{tmp_date}/output/{file_1}'
    etas_entity = EtasDem(arxmlDemPath=workspace_i1,
                          arxmlFltMPath=workspace_i5,
                          arxmlCompositionPath=workspace_i6,
                          xlsx_path=workspace_i2,
                          eventId_path=workspace_i3,
                          dtcId_path=workspace_i4)
    etas_entity.loading()


def fastdem_gui():
    pass


if __name__ == '__main__':
    app()
    # development()
